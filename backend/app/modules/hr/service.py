"""HR domain services: org sections, employee documents, import/export."""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.common.codegen import next_tenant_code
from app.models import Department, Designation, Employee, EmployeeDocument, EmployeeStatusHistory, HrSection, Tenant, User


async def next_section_code(db: AsyncSession, tenant_id: int) -> str:
    return await next_tenant_code(db, model=HrSection, tenant_id=tenant_id, prefix="SEC-", width=4)


async def get_section_or_404(db: AsyncSession, tenant_id: int, section_id: int) -> HrSection:
    row = await db.get(HrSection, section_id)
    if not row or row.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Section not found")
    return row


def section_to_dict(row: HrSection) -> dict[str, Any]:
    return {
        "id": row.id,
        "tenant_id": row.tenant_id,
        "code": row.code,
        "name": row.name,
        "section_type": row.section_type,
        "parent_section_id": row.parent_section_id,
        "department_id": row.department_id,
        "head_employee_id": row.head_employee_id,
        "is_active": row.is_active,
        "created_at": row.created_at.isoformat(),
        "updated_at": row.updated_at.isoformat(),
    }


async def export_employees_excel(db: AsyncSession, tenant_id: int) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = [
        "employee_code",
        "first_name",
        "last_name",
        "email",
        "phone",
        "department_id",
        "designation_id",
        "section_id",
        "employee_category",
        "joining_date",
        "employment_type",
        "is_active",
    ]
    ws.append(headers)
    result = await db.execute(select(Employee).where(Employee.tenant_id == tenant_id).order_by(Employee.employee_code))
    for emp in result.scalars().all():
        ws.append(
            [
                emp.employee_code,
                emp.first_name,
                emp.last_name or "",
                emp.email or "",
                emp.phone or "",
                emp.department_id or "",
                emp.designation_id or "",
                emp.section_id or "",
                emp.employee_category or "",
                emp.joining_date.isoformat() if emp.joining_date else "",
                emp.employment_type or "",
                "Y" if emp.is_active else "N",
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def import_employees_excel(db: AsyncSession, tenant: Tenant, user: User, file: UploadFile) -> dict[str, Any]:
    from openpyxl import load_workbook

    raw = await file.read()
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    updated = 0
    errors: list[str] = []
    for idx, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        try:
            code = str(row[0] or "").strip()
            if not code:
                errors.append(f"Row {idx}: missing employee_code")
                continue
            first_name = str(row[1] or "").strip()
            if not first_name:
                errors.append(f"Row {idx}: missing first_name")
                continue
            existing = (
                await db.execute(select(Employee).where(Employee.tenant_id == tenant.id, Employee.employee_code == code))
            ).scalar_one_or_none()
            dept_id = int(row[5]) if row[5] not in (None, "") else None
            desig_id = int(row[6]) if row[6] not in (None, "") else None
            sec_id = int(row[7]) if row[7] not in (None, "") else None
            if dept_id is not None:
                d = await db.get(Department, dept_id)
                if not d or d.tenant_id != tenant.id:
                    dept_id = None
            if desig_id is not None:
                d = await db.get(Designation, desig_id)
                if not d or d.tenant_id != tenant.id:
                    desig_id = None
            if sec_id is not None:
                s = await db.get(HrSection, sec_id)
                if not s or s.tenant_id != tenant.id:
                    sec_id = None
            cat = str(row[8] or "").strip() or None
            join_raw = row[9]
            join_date: date | None = None
            if join_raw:
                if isinstance(join_raw, datetime):
                    join_date = join_raw.date()
                elif isinstance(join_raw, date):
                    join_date = join_raw
                else:
                    from datetime import datetime as dtmod

                    try:
                        join_date = dtmod.fromisoformat(str(join_raw)[:10]).date()
                    except ValueError:
                        join_date = None
            emp_type = str(row[10] or "").strip() or None
            active = str(row[11] or "Y").upper() in {"Y", "YES", "TRUE", "1"}
            if existing:
                existing.first_name = first_name
                existing.last_name = str(row[2] or "").strip() or None
                existing.email = str(row[3] or "").strip() or None
                existing.phone = str(row[4] or "").strip() or None
                existing.department_id = dept_id
                existing.designation_id = desig_id
                existing.section_id = sec_id
                existing.employee_category = cat
                existing.joining_date = join_date
                existing.employment_type = emp_type
                existing.is_active = active
                updated += 1
            else:
                db.add(
                    Employee(
                        tenant_id=tenant.id,
                        employee_code=code,
                        first_name=first_name,
                        last_name=str(row[2] or "").strip() or None,
                        email=str(row[3] or "").strip() or None,
                        phone=str(row[4] or "").strip() or None,
                        department_id=dept_id,
                        designation_id=desig_id,
                        section_id=sec_id,
                        employee_category=cat,
                        joining_date=join_date,
                        employment_type=emp_type,
                        is_active=active,
                    )
                )
                created += 1
        except Exception as e:  # noqa: BLE001
            errors.append(f"Row {idx}: {e!s}")
    try:
        await db.commit()
    except IntegrityError as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=f"Import failed: {e!s}") from e
    return {"created": created, "updated": updated, "errors": errors[:50]}
