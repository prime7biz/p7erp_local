"""
Merchandising linked module (PrimeX parity slice):
- styles + components + colorways + size scales
- boms + bom items
- consumption plans + plan items
- order followups and pipeline/alerts aggregates
"""
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from pathlib import Path

from io import BytesIO

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, Response, UploadFile, status
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import and_, case, delete, false, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.common.auth import get_current_user
from app.common.storage import FileStorageService
from app.common.codegen import next_tenant_code
from app.common.pagination import MAX_PAGE_SIZE
from app.common.tenant import require_tenant
from app.common.workflow import (
    BOM_TRANSITIONS,
    INQUIRY_TRANSITIONS,
    ORDER_TRANSITIONS,
    QUOTATION_TRANSITIONS,
    next_status_options,
    validate_transition,
)
from app.database import get_db
from app.modules.audit.service import log_action
from app.modules.merch import constants as merch_c
from app.modules.merch.deps import (
    ensure_tenant,
    normalize_optional_choice,
    normalize_style_stage,
    to_decimal,
    to_float_safe,
)
from app.modules.merch.permissions import (
    MERCH_PERMISSION_ALERT_ASSIGN,
    MERCH_PERMISSION_ALERT_SCAN,
    MERCH_PERMISSION_BOM_APPROVE,
    MERCH_PERMISSION_BOM_FREEZE,
    MERCH_PERMISSION_PO_GENERATE,
    MERCH_PERMISSION_STYLE_MANAGE,
    MERCH_PERMISSION_TNA_MANAGE,
    MERCH_PERMISSION_WASTAGE_MANAGE,
    MERCH_PERMISSION_ALERT_DEFINITIONS,
    require_merch_permission,
)

from app.models import (
    AlertDefinition,
    AlertInstance,
    AlertHistory,
    AlertComment,
    AlertRelatedEntity,
    AlertEscalationLog,
    AlertSavedView,
    Bom,
    BomItem,
    ConsumptionPlan,
    ConsumptionPlanItem,
    Customer,
    Followup,
    FollowupActionTemplate,
    FollowupActionRejectionLog,
    OrderFollowupAction,
    FollowupActionComment,
    GarmentStyle,
    Inquiry,
    Item,
    ItemCategory,
    ItemUnit,
    Order,
    ProformaInvoice,
    ProformaInvoiceOrder,
    PurchaseOrder,
    PurchaseOrderItem,
    Quotation,
    FxReceipt,
    Shipment,
    StockMovement,
    StyleColorway,
    Warehouse,
    StyleComponent,
    StyleSizeScale,
    Tenant,
    TradeCase,
    User,
    Vendor,
    WastageReason,
    WastageTransaction,
    WastageThresholdRule,
    WastageOrderSummary,
    WastageSavedView,
)

router = APIRouter(prefix="/merch", tags=["merch"])
STYLE_LIFECYCLE_STAGES = merch_c.STYLE_LIFECYCLE_STAGES
STYLE_PRIORITY_VALUES = merch_c.STYLE_PRIORITY_VALUES
STYLE_RISK_VALUES = merch_c.STYLE_RISK_VALUES

_ensure_tenant = ensure_tenant
_to_decimal = to_decimal
_normalize_style_stage = normalize_style_stage
_normalize_optional_choice = normalize_optional_choice
_to_float_safe = to_float_safe


class StyleCreate(BaseModel):
    style_code: str | None = None
    name: str
    buyer_customer_id: int | None = None
    season: str | None = None
    department: str | None = None
    product_type: str | None = None
    fabric_type: str | None = None
    gsm: str | None = None
    fit_type: str | None = None
    wash_type: str | None = None
    brand: str | None = None
    buyer_style_ref: str | None = None
    hs_code: str | None = None
    uom: str | None = None
    target_fob: str | None = None
    currency: str | None = None
    sample_lead_days: int | None = None
    production_lead_days: int | None = None
    is_active_for_new_orders: bool = True
    lifecycle_stage: str = "INQUIRY"
    priority: str | None = None
    risk_level: str | None = None
    style_image_url: str | None = None
    status: str = "ACTIVE"
    notes: str | None = None


class StyleUpdate(BaseModel):
    style_code: str | None = None
    name: str | None = None
    buyer_customer_id: int | None = None
    season: str | None = None
    department: str | None = None
    product_type: str | None = None
    fabric_type: str | None = None
    gsm: str | None = None
    fit_type: str | None = None
    wash_type: str | None = None
    brand: str | None = None
    buyer_style_ref: str | None = None
    hs_code: str | None = None
    uom: str | None = None
    target_fob: str | None = None
    currency: str | None = None
    sample_lead_days: int | None = None
    production_lead_days: int | None = None
    is_active_for_new_orders: bool | None = None
    lifecycle_stage: str | None = None
    priority: str | None = None
    risk_level: str | None = None
    style_image_url: str | None = None
    status: str | None = None
    notes: str | None = None


class StyleImageUploadResponse(BaseModel):
    style_image_url: str
    filename: str
    size_bytes: int


class StyleComponentBody(BaseModel):
    component_name: str
    sequence_no: int = 1
    notes: str | None = None


class StyleColorwayBody(BaseModel):
    color_name: str
    color_code: str | None = None
    notes: str | None = None


class StyleSizeScaleBody(BaseModel):
    scale_name: str
    sizes_csv: str | None = None
    notes: str | None = None


class BomCreate(BaseModel):
    style_id: int
    version_no: int = 1
    status: str = "DRAFT"
    notes: str | None = None


class BomUpdate(BaseModel):
    version_no: int | None = None
    status: str | None = None
    notes: str | None = None


class BomItemBody(BaseModel):
    item_id: int | None = None
    category: str
    item_code: str | None = None
    description: str | None = None
    uom: str | None = None
    base_consumption: str
    wastage_pct: str | None = None


class ConsumptionPlanCreate(BaseModel):
    order_id: int
    status: str = "PLANNED"


class ConsumptionPlanUpdate(BaseModel):
    status: str | None = None


class ConsumptionPlanItemBody(BaseModel):
    item_code: str | None = None
    required_qty: str
    uom: str | None = None


class FollowupCreate(BaseModel):
    order_id: int
    title: str
    due_date: date | None = None
    status: str = "OPEN"
    severity: str | None = None
    notes: str | None = None


class FollowupUpdate(BaseModel):
    title: str | None = None
    due_date: date | None = None
    status: str | None = None
    severity: str | None = None
    notes: str | None = None


class StyleSummaryResponse(BaseModel):
    style_id: int
    inquiry_count: int
    quotation_count: int
    order_count: int
    open_followup_actions: int
    overdue_followup_actions: int
    shipment_count: int
    shipped_order_qty: int
    pending_order_qty: int
    invoice_amount: str
    received_amount: str
    due_amount: str
    last_event_at: datetime | None = None
    next_due_at: date | None = None


class StyleTimelineEvent(BaseModel):
    event_type: str
    reference: str
    status: str | None = None
    event_at: datetime
    notes: str | None = None


class StyleReportRow(BaseModel):
    style_id: int
    style_code: str
    style_name: str
    lifecycle_stage: str
    priority: str | None = None
    risk_level: str | None = None
    open_followup_actions: int
    overdue_followup_actions: int
    invoice_amount: str
    received_amount: str
    due_amount: str
    last_event_at: datetime | None = None
    next_due_at: date | None = None


async def _resolve_style_order_ids(db: AsyncSession, tenant_id: int, style: GarmentStyle) -> list[int]:
    quotation_ids = (
        await db.execute(
            select(Quotation.id).where(
                Quotation.tenant_id == tenant_id,
                Quotation.style_id == style.id,
            )
        )
    ).scalars().all()
    style_code_lower = style.style_code.lower()
    style_name_lower = style.name.lower()
    order_ids = (
        await db.execute(
            select(Order.id)
            .where(Order.tenant_id == tenant_id)
            .where(
                or_(
                    Order.quotation_id.in_(quotation_ids) if quotation_ids else false(),
                    func.lower(func.coalesce(Order.style_ref, "")) == style_code_lower,
                    func.lower(func.coalesce(Order.style_ref, "")) == style_name_lower,
                )
            )
        )
    ).scalars().all()
    return list({oid for oid in order_ids})


async def _resolve_style_order_ids_batch(
    db: AsyncSession, tenant_id: int, styles: list[GarmentStyle]
) -> dict[int, list[int]]:
    """Resolve order ids for many styles with one quotations query and one orders query (avoids N+1)."""
    if not styles:
        return {}
    style_ids = [s.id for s in styles]
    qres = await db.execute(
        select(Quotation.id, Quotation.style_id).where(
            Quotation.tenant_id == tenant_id,
            Quotation.style_id.in_(style_ids),
        )
    )
    quot_by_style: dict[int, list[int]] = defaultdict(list)
    all_qids: list[int] = []
    for qid, sid in qres.all():
        quot_by_style[sid].append(qid)
        all_qids.append(qid)
    ref_set: set[str] = set()
    for s in styles:
        ref_set.add((s.style_code or "").lower())
        ref_set.add((s.name or "").lower())
    conditions = []
    if all_qids:
        conditions.append(Order.quotation_id.in_(all_qids))
    conditions.append(func.lower(func.coalesce(Order.style_ref, "")).in_(list(ref_set)))
    ores = await db.execute(select(Order).where(Order.tenant_id == tenant_id, or_(*conditions)))
    orders = list(ores.scalars().all())
    out: dict[int, list[int]] = {}
    for s in styles:
        qids = set(quot_by_style.get(s.id, []))
        cl = (s.style_code or "").lower()
        nl = (s.name or "").lower()
        seen: set[int] = set()
        oid_list: list[int] = []
        for o in orders:
            if o.quotation_id in qids or (o.style_ref or "").lower() == cl or (o.style_ref or "").lower() == nl:
                if o.id not in seen:
                    seen.add(o.id)
                    oid_list.append(o.id)
        out[s.id] = oid_list
    return out


async def _build_style_summary(
    db: AsyncSession,
    tenant_id: int,
    style: GarmentStyle,
    *,
    inquiry_count: int | None = None,
    quotation_count: int | None = None,
    order_ids: list[int] | None = None,
) -> StyleSummaryResponse:
    if inquiry_count is None:
        inquiry_count = int(
            (
                await db.execute(
                    select(func.count(Inquiry.id)).where(Inquiry.tenant_id == tenant_id, Inquiry.style_id == style.id)
                )
            ).scalar_one()
        )
    if quotation_count is None:
        quotation_count = int(
            (
                await db.execute(
                    select(func.count(Quotation.id)).where(Quotation.tenant_id == tenant_id, Quotation.style_id == style.id)
                )
            ).scalar_one()
        )
    if order_ids is None:
        order_ids = await _resolve_style_order_ids(db, tenant_id, style)
    order_count = len(order_ids)
    open_followup_actions = 0
    overdue_followup_actions = 0
    shipment_count = 0
    shipped_order_qty = 0
    pending_order_qty = 0
    invoice_amount = Decimal("0")
    received_amount = Decimal("0")
    last_event_at: datetime | None = style.updated_at
    next_due_at: date | None = None

    if order_ids:
        actions = (
            await db.execute(
                select(OrderFollowupAction).where(
                    OrderFollowupAction.tenant_id == tenant_id,
                    OrderFollowupAction.order_id.in_(order_ids),
                    OrderFollowupAction.is_active.is_(True),
                )
            )
        ).scalars().all()
        today = date.today()
        for action in actions:
            status_text = (action.status or "").strip().lower()
            if status_text not in {"done", "completed", "closed"}:
                open_followup_actions += 1
            if action.planned_date and action.planned_date < today and status_text not in {"done", "completed", "closed"}:
                overdue_followup_actions += 1
            if action.planned_date and (next_due_at is None or action.planned_date < next_due_at):
                next_due_at = action.planned_date
            if action.updated_at and (last_event_at is None or action.updated_at > last_event_at):
                last_event_at = action.updated_at

        orders = (await db.execute(select(Order).where(Order.id.in_(order_ids)))).scalars().all()
        trade_cases = (
            await db.execute(select(TradeCase).where(TradeCase.tenant_id == tenant_id, TradeCase.order_id.in_(order_ids)))
        ).scalars().all()
        trade_case_by_order = {tc.order_id: tc for tc in trade_cases if tc.order_id is not None}
        trade_case_ids = [tc.id for tc in trade_cases]
        shipments: list[Shipment] = []
        if trade_case_ids:
            shipments = (
                await db.execute(
                    select(Shipment).where(Shipment.tenant_id == tenant_id, Shipment.trade_case_id.in_(trade_case_ids))
                )
            ).scalars().all()
        shipped_trade_case_ids = {s.trade_case_id for s in shipments if (s.status or "").upper() in {"SHIPPED", "DELIVERED", "CLOSED"}}
        shipment_count = len(shipments)
        for order in orders:
            qty = order.quantity or 0
            trade_case = trade_case_by_order.get(order.id)
            if trade_case and trade_case.id in shipped_trade_case_ids:
                shipped_order_qty += qty
            else:
                pending_order_qty += qty
            if order.updated_at and (last_event_at is None or order.updated_at > last_event_at):
                last_event_at = order.updated_at

        invoice_rows = (
            await db.execute(
                select(ProformaInvoice)
                .join(ProformaInvoiceOrder, ProformaInvoiceOrder.proforma_invoice_id == ProformaInvoice.id)
                .where(
                    ProformaInvoice.tenant_id == tenant_id,
                    ProformaInvoiceOrder.order_id.in_(order_ids),
                )
            )
        ).scalars().all()
        invoice_refs = {inv.reference for inv in invoice_rows if inv.reference}
        for invoice in invoice_rows:
            invoice_amount += _to_decimal(invoice.amount)
            if invoice.updated_at and (last_event_at is None or invoice.updated_at > last_event_at):
                last_event_at = invoice.updated_at

        if invoice_refs:
            receipts = (
                await db.execute(
                    select(FxReceipt).where(
                        FxReceipt.tenant_id == tenant_id,
                        FxReceipt.source_ref.in_(list(invoice_refs)),
                    )
                )
            ).scalars().all()
            for receipt in receipts:
                received_amount += _to_decimal(receipt.base_amount)
                if receipt.created_at and (last_event_at is None or receipt.created_at > last_event_at):
                    last_event_at = receipt.created_at

    due_amount = invoice_amount - received_amount
    if due_amount < Decimal("0"):
        due_amount = Decimal("0")
    return StyleSummaryResponse(
        style_id=style.id,
        inquiry_count=inquiry_count,
        quotation_count=quotation_count,
        order_count=order_count,
        open_followup_actions=open_followup_actions,
        overdue_followup_actions=overdue_followup_actions,
        shipment_count=shipment_count,
        shipped_order_qty=shipped_order_qty,
        pending_order_qty=pending_order_qty,
        invoice_amount=str(invoice_amount.quantize(Decimal("0.01"))),
        received_amount=str(received_amount.quantize(Decimal("0.01"))),
        due_amount=str(due_amount.quantize(Decimal("0.01"))),
        last_event_at=last_event_at,
        next_due_at=next_due_at,
    )


def _garment_style_list_base_stmt(
    tenant_id: int,
    *,
    search: str | None = None,
    status_filter: str | None = None,
    buyer_customer_id: int | None = None,
    season: str | None = None,
    department: str | None = None,
    lifecycle_stage: str | None = None,
    active_for_orders: bool | None = None,
    priority: str | None = None,
    risk_level: str | None = None,
    style_ids: list[int] | None = None,
):
    """Shared SQL filters for garment style list and summary-report (parity with GET /styles)."""
    stmt = select(GarmentStyle).where(GarmentStyle.tenant_id == tenant_id)
    if style_ids:
        stmt = stmt.where(GarmentStyle.id.in_(style_ids))
    if search:
        pattern = f"%{search.strip().lower()}%"
        stmt = stmt.where(
            or_(
                func.lower(GarmentStyle.style_code).like(pattern),
                func.lower(GarmentStyle.name).like(pattern),
                func.lower(func.coalesce(GarmentStyle.buyer_style_ref, "")).like(pattern),
                func.lower(func.coalesce(GarmentStyle.product_type, "")).like(pattern),
            )
        )
    if status_filter:
        stmt = stmt.where(GarmentStyle.status == status_filter)
    if buyer_customer_id is not None:
        stmt = stmt.where(GarmentStyle.buyer_customer_id == buyer_customer_id)
    if season:
        stmt = stmt.where(func.lower(func.coalesce(GarmentStyle.season, "")) == season.strip().lower())
    if department:
        stmt = stmt.where(func.lower(func.coalesce(GarmentStyle.department, "")) == department.strip().lower())
    if lifecycle_stage:
        stmt = stmt.where(GarmentStyle.lifecycle_stage == _normalize_style_stage(lifecycle_stage))
    if active_for_orders is not None:
        stmt = stmt.where(GarmentStyle.is_active_for_new_orders == active_for_orders)
    normalized_priority = _normalize_optional_choice(priority, STYLE_PRIORITY_VALUES, "priority")
    if normalized_priority:
        stmt = stmt.where(GarmentStyle.priority == normalized_priority)
    normalized_risk = _normalize_optional_choice(risk_level, STYLE_RISK_VALUES, "risk_level")
    if normalized_risk:
        stmt = stmt.where(GarmentStyle.risk_level == normalized_risk)
    return stmt


async def _style_report_rows_for_styles(
    db: AsyncSession,
    tenant_id: int,
    styles: list[GarmentStyle],
    *,
    critical_only: bool,
    normalized_saved_view: str,
) -> list[StyleReportRow]:
    rows: list[StyleReportRow] = []
    if not styles:
        return rows
    sid_list = [s.id for s in styles]
    inquiry_counts: dict[int, int] = {}
    quotation_counts: dict[int, int] = {}
    for sid, cnt in (
        await db.execute(
            select(Inquiry.style_id, func.count(Inquiry.id))
            .where(Inquiry.tenant_id == tenant_id, Inquiry.style_id.in_(sid_list))
            .group_by(Inquiry.style_id)
        )
    ).all():
        inquiry_counts[int(sid)] = int(cnt)
    for sid, cnt in (
        await db.execute(
            select(Quotation.style_id, func.count(Quotation.id))
            .where(Quotation.tenant_id == tenant_id, Quotation.style_id.in_(sid_list))
            .group_by(Quotation.style_id)
        )
    ).all():
        quotation_counts[int(sid)] = int(cnt)
    order_ids_by_style = await _resolve_style_order_ids_batch(db, tenant_id, styles)
    for style in styles:
        summary = await _build_style_summary(
            db,
            tenant_id,
            style,
            inquiry_count=inquiry_counts.get(style.id, 0),
            quotation_count=quotation_counts.get(style.id, 0),
            order_ids=order_ids_by_style.get(style.id, []),
        )
        is_payment_overdue = _to_decimal(summary.due_amount) > Decimal("0")
        has_overdue_milestone = summary.overdue_followup_actions > 0
        if critical_only and not (has_overdue_milestone or is_payment_overdue):
            continue
        if normalized_saved_view == "critical_styles" and not (has_overdue_milestone or is_payment_overdue):
            continue
        if normalized_saved_view == "shipment_due_week":
            if summary.next_due_at is None:
                continue
            days_to_due = (summary.next_due_at - date.today()).days
            if days_to_due < 0 or days_to_due > 7:
                continue
        if normalized_saved_view == "payment_overdue" and not is_payment_overdue:
            continue
        rows.append(
            StyleReportRow(
                style_id=style.id,
                style_code=style.style_code,
                style_name=style.name,
                lifecycle_stage=style.lifecycle_stage,
                priority=style.priority,
                risk_level=style.risk_level,
                open_followup_actions=summary.open_followup_actions,
                overdue_followup_actions=summary.overdue_followup_actions,
                invoice_amount=summary.invoice_amount,
                received_amount=summary.received_amount,
                due_amount=summary.due_amount,
                last_event_at=summary.last_event_at,
                next_due_at=summary.next_due_at,
            )
        )
    return rows


@router.get("/styles")
async def list_styles(
    response: Response,
    search: str | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    buyer_customer_id: int | None = Query(default=None),
    season: str | None = Query(default=None),
    department: str | None = Query(default=None),
    lifecycle_stage: str | None = Query(default=None),
    active_for_orders: bool | None = Query(default=None),
    priority: str | None = Query(default=None),
    risk_level: str | None = Query(default=None),
    style_ids: list[int] | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    stmt = _garment_style_list_base_stmt(
        tenant.id,
        search=search,
        status_filter=status_filter,
        buyer_customer_id=buyer_customer_id,
        season=season,
        department=department,
        lifecycle_stage=lifecycle_stage,
        active_for_orders=active_for_orders,
        priority=priority,
        risk_level=risk_level,
        style_ids=style_ids,
    )
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = int((await db.execute(count_stmt)).scalar() or 0)
    if style_ids:
        order_expr = case(
            *[(GarmentStyle.id == sid, pos) for pos, sid in enumerate(style_ids)],
            else_=len(style_ids),
        )
        stmt_ordered = stmt.order_by(order_expr.asc())
    else:
        stmt_ordered = stmt.order_by(GarmentStyle.updated_at.desc(), GarmentStyle.id.desc())
    result = await db.execute(stmt_ordered.offset(offset).limit(limit))
    rows = result.scalars().all()
    response.headers["X-Total-Count"] = str(total)
    return rows


@router.post("/styles", status_code=201)
async def create_style(
    body: StyleCreate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_merch_permission(MERCH_PERMISSION_STYLE_MANAGE)),
):
    _ensure_tenant(user, tenant)
    if body.buyer_customer_id is not None:
        bc = await db.get(Customer, body.buyer_customer_id)
        if not bc or bc.tenant_id != tenant.id:
            raise HTTPException(status_code=404, detail="Buyer customer not found")
    style_code = (body.style_code or "").strip().upper()
    if not style_code:
        style_code = await next_tenant_code(
            db,
            model=GarmentStyle,
            tenant_id=tenant.id,
            prefix="STY-",
            width=4,
        )
    existing = await db.execute(
        select(GarmentStyle.id).where(
            GarmentStyle.tenant_id == tenant.id,
            GarmentStyle.style_code == style_code,
        )
    )
    if existing.scalar_one_or_none() is not None:
        raise HTTPException(status_code=409, detail="Style code already exists for this tenant")

    lifecycle_stage = _normalize_style_stage(body.lifecycle_stage)
    normalized_priority = _normalize_optional_choice(body.priority, STYLE_PRIORITY_VALUES, "priority")
    normalized_risk = _normalize_optional_choice(body.risk_level, STYLE_RISK_VALUES, "risk_level")
    row = GarmentStyle(
        tenant_id=tenant.id,
        style_code=style_code,
        name=body.name,
        buyer_customer_id=body.buyer_customer_id,
        season=body.season,
        department=body.department,
        product_type=body.product_type,
        fabric_type=body.fabric_type,
        gsm=body.gsm,
        fit_type=body.fit_type,
        wash_type=body.wash_type,
        brand=body.brand,
        buyer_style_ref=body.buyer_style_ref,
        hs_code=body.hs_code,
        uom=body.uom,
        target_fob=body.target_fob,
        currency=body.currency,
        sample_lead_days=body.sample_lead_days,
        production_lead_days=body.production_lead_days,
        is_active_for_new_orders=body.is_active_for_new_orders,
        lifecycle_stage=lifecycle_stage,
        priority=normalized_priority,
        risk_level=normalized_risk,
        style_image_url=body.style_image_url,
        status=body.status,
        notes=body.notes,
    )
    db.add(row)
    await db.flush()
    await log_action(
        db,
        tenant_id=tenant.id,
        user_id=user.id,
        action="STYLE_CREATED",
        resource="garment_style",
        details=f"style_id={row.id}, style_code={row.style_code}",
    )
    await db.refresh(row)
    return row


@router.get("/styles/summary-report", response_model=list[StyleReportRow])
async def list_style_summary_report(
    response: Response,
    search: str | None = Query(default=None),
    lifecycle_stage: str | None = Query(default=None),
    critical_only: bool = Query(default=False),
    saved_view: str | None = Query(default=None),
    style_ids: list[int] | None = Query(default=None),
    report_limit: int | None = Query(default=None, ge=1, le=MAX_PAGE_SIZE),
    report_offset: int = Query(default=0, ge=0),
    status_filter: str | None = Query(default=None, alias="status"),
    buyer_customer_id: int | None = Query(default=None),
    season: str | None = Query(default=None),
    department: str | None = Query(default=None),
    active_for_orders: bool | None = Query(default=None),
    priority: str | None = Query(default=None),
    risk_level: str | None = Query(default=None),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    stmt = _garment_style_list_base_stmt(
        tenant.id,
        search=search,
        status_filter=status_filter,
        buyer_customer_id=buyer_customer_id,
        season=season,
        department=department,
        lifecycle_stage=lifecycle_stage,
        active_for_orders=active_for_orders,
        priority=priority,
        risk_level=risk_level,
        style_ids=style_ids,
    )
    normalized_saved_view = (saved_view or "").strip().lower()
    needs_post_summary_filter = (
        critical_only
        or normalized_saved_view == "critical_styles"
        or normalized_saved_view == "shipment_due_week"
        or normalized_saved_view == "payment_overdue"
    )

    def _ordered_stmt(base):
        if style_ids:
            order_expr = case(
                *[(GarmentStyle.id == sid, pos) for pos, sid in enumerate(style_ids)],
                else_=len(style_ids),
            )
            return base.order_by(order_expr.asc())
        return base.order_by(GarmentStyle.updated_at.desc(), GarmentStyle.id.desc())

    # Fast path: SQL-level paging; summaries only for the current page (no critical/saved-view filters).
    if not needs_post_summary_filter and report_limit is not None:
        count_stmt = select(func.count()).select_from(stmt.subquery())
        total = int((await db.execute(count_stmt)).scalar() or 0)
        result = await db.execute(_ordered_stmt(stmt).offset(report_offset).limit(report_limit))
        page_styles = list(result.scalars().all())
        rows = await _style_report_rows_for_styles(
            db,
            tenant.id,
            page_styles,
            critical_only=False,
            normalized_saved_view="",
        )
        response.headers["X-Total-Count"] = str(total)
        return rows

    # Complex path: filter after building summaries; optional slice at end.
    styles = list((await db.execute(_ordered_stmt(stmt))).scalars().all())
    rows = await _style_report_rows_for_styles(
        db,
        tenant.id,
        styles,
        critical_only=critical_only,
        normalized_saved_view=normalized_saved_view,
    )
    total_count = len(rows)
    if report_limit is not None:
        rows = rows[report_offset : report_offset + report_limit]
    response.headers["X-Total-Count"] = str(total_count)
    return rows


@router.get("/styles/{style_id}")
async def get_style(
    style_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(GarmentStyle, style_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Style not found")
    return row


@router.patch("/styles/{style_id}")
async def update_style(
    style_id: int,
    body: StyleUpdate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(GarmentStyle, style_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Style not found")
    changes: list[str] = []
    if body.style_code is not None:
        new_code = body.style_code.strip().upper()
        if not new_code:
            raise HTTPException(status_code=422, detail="style_code cannot be empty")
        existing = await db.execute(
            select(GarmentStyle.id).where(
                GarmentStyle.tenant_id == tenant.id,
                GarmentStyle.style_code == new_code,
                GarmentStyle.id != style_id,
            )
        )
        if existing.scalar_one_or_none() is not None:
            raise HTTPException(status_code=409, detail="Style code already exists for this tenant")
        if row.style_code != new_code:
            row.style_code = new_code
            changes.append("style_code")

    if body.lifecycle_stage is not None:
        new_stage = _normalize_style_stage(body.lifecycle_stage)
        if row.lifecycle_stage != new_stage:
            row.lifecycle_stage = new_stage
            changes.append("lifecycle_stage")

    if body.priority is not None:
        normalized_priority = _normalize_optional_choice(body.priority, STYLE_PRIORITY_VALUES, "priority")
        if row.priority != normalized_priority:
            row.priority = normalized_priority
            changes.append("priority")

    if body.risk_level is not None:
        normalized_risk = _normalize_optional_choice(body.risk_level, STYLE_RISK_VALUES, "risk_level")
        if row.risk_level != normalized_risk:
            row.risk_level = normalized_risk
            changes.append("risk_level")

    for field in (
        "name",
        "buyer_customer_id",
        "season",
        "department",
        "product_type",
        "fabric_type",
        "gsm",
        "fit_type",
        "wash_type",
        "brand",
        "buyer_style_ref",
        "hs_code",
        "uom",
        "target_fob",
        "currency",
        "sample_lead_days",
        "production_lead_days",
        "is_active_for_new_orders",
        "style_image_url",
        "status",
        "notes",
    ):
        value = getattr(body, field)
        if value is not None:
            if getattr(row, field) != value:
                changes.append(field)
            setattr(row, field, value)
    await db.flush()
    if changes:
        sensitive_fields = {"status", "lifecycle_stage", "target_fob", "currency", "risk_level", "priority"}
        touched_sensitive = sorted(set(changes).intersection(sensitive_fields))
        await log_action(
            db,
            tenant_id=tenant.id,
            user_id=user.id,
            action="STYLE_UPDATED",
            resource="garment_style",
            details=f"style_id={row.id}, changed={','.join(sorted(set(changes)))}, sensitive={','.join(touched_sensitive) if touched_sensitive else 'none'}",
        )
    await db.refresh(row)
    return row


@router.post(
    "/styles/{style_id}/upload-picture",
    response_model=StyleImageUploadResponse,
)
async def upload_style_picture(
    style_id: int,
    *,
    file: UploadFile = File(...),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    style = await db.get(GarmentStyle, style_id)
    if not style or style.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Style not found")

    safe_filename, api_url, disk_path = await FileStorageService.save_file(file, tenant.id, "style_pictures")
    style.style_image_url = api_url
    await db.flush()

    size_bytes = Path(disk_path).stat().st_size if Path(disk_path).exists() else 0

    return StyleImageUploadResponse(
        style_image_url=style.style_image_url,
        filename=safe_filename,
        size_bytes=size_bytes,
    )


@router.delete("/styles/{style_id}", status_code=204)
async def delete_style(
    style_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(GarmentStyle, style_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Style not found")
    inquiry_link_count = (
        await db.execute(select(func.count(Inquiry.id)).where(Inquiry.tenant_id == tenant.id, Inquiry.style_id == style_id))
    ).scalar_one()
    quotation_link_count = (
        await db.execute(select(func.count(Quotation.id)).where(Quotation.tenant_id == tenant.id, Quotation.style_id == style_id))
    ).scalar_one()
    order_link_count = (
        await db.execute(
            select(func.count(Order.id))
            .select_from(Order)
            .join(Quotation, Quotation.id == Order.quotation_id, isouter=True)
            .where(
                Order.tenant_id == tenant.id,
                or_(
                    Quotation.style_id == style_id,
                    func.lower(func.coalesce(Order.style_ref, "")) == func.lower(row.style_code),
                ),
            )
        )
    ).scalar_one()
    if inquiry_link_count or quotation_link_count or order_link_count:
        raise HTTPException(
            status_code=409,
            detail="Style is linked with inquiry/quotation/order records. Archive the style instead of deleting.",
        )
    await db.delete(row)
    await db.flush()
    await log_action(
        db,
        tenant_id=tenant.id,
        user_id=user.id,
        action="STYLE_DELETED",
        resource="garment_style",
        details=f"style_id={style_id}, style_code={row.style_code}",
    )


@router.get("/styles/{style_id}/summary", response_model=StyleSummaryResponse)
async def get_style_summary(
    style_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(GarmentStyle, style_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Style not found")
    return await _build_style_summary(db, tenant.id, row)


@router.get("/styles/{style_id}/timeline", response_model=list[StyleTimelineEvent])
async def list_style_timeline(
    style_id: int,
    limit: int = Query(default=100, ge=1, le=500),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    style = await db.get(GarmentStyle, style_id)
    if not style or style.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Style not found")

    events: list[StyleTimelineEvent] = []
    inquiries = (
        await db.execute(select(Inquiry).where(Inquiry.tenant_id == tenant.id, Inquiry.style_id == style_id))
    ).scalars().all()
    for inq in inquiries:
        events.append(
            StyleTimelineEvent(
                event_type="INQUIRY",
                reference=inq.inquiry_code,
                status=inq.status,
                event_at=inq.updated_at,
                notes=inq.notes,
            )
        )
    quotations = (
        await db.execute(select(Quotation).where(Quotation.tenant_id == tenant.id, Quotation.style_id == style_id))
    ).scalars().all()
    quotation_ids = [q.id for q in quotations]
    for q in quotations:
        events.append(
            StyleTimelineEvent(
                event_type="QUOTATION",
                reference=q.quotation_code,
                status=q.status,
                event_at=q.updated_at,
                notes=q.notes,
            )
        )

    order_stmt = select(Order).where(Order.tenant_id == tenant.id)
    if quotation_ids:
        order_stmt = order_stmt.where(
            or_(
                Order.quotation_id.in_(quotation_ids),
                func.lower(func.coalesce(Order.style_ref, "")) == style.style_code.lower(),
            )
        )
    else:
        order_stmt = order_stmt.where(
            func.lower(func.coalesce(Order.style_ref, "")) == style.style_code.lower(),
        )
    orders = (await db.execute(order_stmt)).scalars().all()
    order_ids = [o.id for o in orders]
    for order in orders:
        events.append(
            StyleTimelineEvent(
                event_type="ORDER",
                reference=order.order_code,
                status=order.status,
                event_at=order.updated_at,
                notes=order.remarks,
            )
        )

    if order_ids:
        actions = (
            await db.execute(
                select(OrderFollowupAction).where(
                    OrderFollowupAction.tenant_id == tenant.id,
                    OrderFollowupAction.order_id.in_(order_ids),
                )
            )
        ).scalars().all()
        for action in actions:
            events.append(
                StyleTimelineEvent(
                    event_type="FOLLOWUP",
                    reference=action.title,
                    status=action.status,
                    event_at=action.updated_at,
                    notes=action.remarks,
                )
            )

        shipments = (
            await db.execute(
                select(Shipment)
                .join(TradeCase, TradeCase.id == Shipment.trade_case_id)
                .where(Shipment.tenant_id == tenant.id, TradeCase.order_id.in_(order_ids))
            )
        ).scalars().all()
        for shipment in shipments:
            events.append(
                StyleTimelineEvent(
                    event_type="SHIPMENT",
                    reference=shipment.reference,
                    status=shipment.status,
                    event_at=shipment.updated_at,
                    notes=shipment.notes,
                )
            )

        invoices = (
            await db.execute(
                select(ProformaInvoice)
                .join(ProformaInvoiceOrder, ProformaInvoiceOrder.proforma_invoice_id == ProformaInvoice.id)
                .where(ProformaInvoice.tenant_id == tenant.id, ProformaInvoiceOrder.order_id.in_(order_ids))
            )
        ).scalars().all()
        invoice_refs = []
        for invoice in invoices:
            if invoice.reference:
                invoice_refs.append(invoice.reference)
            events.append(
                StyleTimelineEvent(
                    event_type="INVOICE",
                    reference=invoice.reference,
                    status=invoice.status,
                    event_at=invoice.updated_at,
                    notes=invoice.terms_of_payment,
                )
            )
        if invoice_refs:
            receipts = (
                await db.execute(
                    select(FxReceipt).where(
                        FxReceipt.tenant_id == tenant.id,
                        FxReceipt.source_ref.in_(invoice_refs),
                    )
                )
            ).scalars().all()
            for receipt in receipts:
                events.append(
                    StyleTimelineEvent(
                        event_type="PAYMENT_RECEIPT",
                        reference=receipt.receipt_no,
                        status=receipt.status,
                        event_at=receipt.created_at,
                        notes=receipt.source_ref,
                    )
                )

    events.sort(key=lambda item: item.event_at, reverse=True)
    return events[:limit]


@router.get("/styles/{style_id}/components")
async def list_style_components(
    style_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    result = await db.execute(
        select(StyleComponent)
        .where(StyleComponent.style_id == style_id, StyleComponent.tenant_id == tenant.id)
        .order_by(StyleComponent.sequence_no, StyleComponent.id)
    )
    return result.scalars().all()


@router.post("/styles/{style_id}/components", status_code=201)
async def create_style_component(
    style_id: int,
    body: StyleComponentBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = StyleComponent(tenant_id=tenant.id, style_id=style_id, **body.model_dump())
    db.add(row)
    await db.flush()
    await db.refresh(row)
    return row


@router.patch("/styles/{style_id}/components/{component_id}")
async def update_style_component(
    style_id: int,
    component_id: int,
    body: StyleComponentBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(StyleComponent, component_id)
    if not row or row.tenant_id != tenant.id or row.style_id != style_id:
        raise HTTPException(status_code=404, detail="Component not found")
    row.component_name = body.component_name
    row.sequence_no = body.sequence_no
    row.notes = body.notes
    await db.flush()
    await db.refresh(row)
    return row


@router.delete("/styles/{style_id}/components/{component_id}", status_code=204)
async def delete_style_component(
    style_id: int,
    component_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(StyleComponent, component_id)
    if not row or row.tenant_id != tenant.id or row.style_id != style_id:
        raise HTTPException(status_code=404, detail="Component not found")
    await db.delete(row)
    await db.flush()


@router.get("/styles/{style_id}/colorways")
async def list_style_colorways(
    style_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    result = await db.execute(
        select(StyleColorway)
        .where(StyleColorway.style_id == style_id, StyleColorway.tenant_id == tenant.id)
        .order_by(StyleColorway.id)
    )
    return result.scalars().all()


@router.post("/styles/{style_id}/colorways", status_code=201)
async def create_style_colorway(
    style_id: int,
    body: StyleColorwayBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = StyleColorway(tenant_id=tenant.id, style_id=style_id, **body.model_dump())
    db.add(row)
    await db.flush()
    await db.refresh(row)
    return row


@router.patch("/styles/{style_id}/colorways/{colorway_id}")
async def update_style_colorway(
    style_id: int,
    colorway_id: int,
    body: StyleColorwayBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(StyleColorway, colorway_id)
    if not row or row.tenant_id != tenant.id or row.style_id != style_id:
        raise HTTPException(status_code=404, detail="Colorway not found")
    row.color_name = body.color_name
    row.color_code = body.color_code
    row.notes = body.notes
    await db.flush()
    await db.refresh(row)
    return row


@router.delete("/styles/{style_id}/colorways/{colorway_id}", status_code=204)
async def delete_style_colorway(
    style_id: int,
    colorway_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(StyleColorway, colorway_id)
    if not row or row.tenant_id != tenant.id or row.style_id != style_id:
        raise HTTPException(status_code=404, detail="Colorway not found")
    await db.delete(row)
    await db.flush()


@router.get("/styles/{style_id}/size-scales")
async def list_style_size_scales(
    style_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    result = await db.execute(
        select(StyleSizeScale)
        .where(StyleSizeScale.style_id == style_id, StyleSizeScale.tenant_id == tenant.id)
        .order_by(StyleSizeScale.id)
    )
    return result.scalars().all()


@router.post("/styles/{style_id}/size-scales", status_code=201)
async def create_style_size_scale(
    style_id: int,
    body: StyleSizeScaleBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = StyleSizeScale(tenant_id=tenant.id, style_id=style_id, **body.model_dump())
    db.add(row)
    await db.flush()
    await db.refresh(row)
    return row


@router.patch("/styles/{style_id}/size-scales/{scale_id}")
async def update_style_size_scale(
    style_id: int,
    scale_id: int,
    body: StyleSizeScaleBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(StyleSizeScale, scale_id)
    if not row or row.tenant_id != tenant.id or row.style_id != style_id:
        raise HTTPException(status_code=404, detail="Size scale not found")
    row.scale_name = body.scale_name
    row.sizes_csv = body.sizes_csv
    row.notes = body.notes
    await db.flush()
    await db.refresh(row)
    return row


@router.delete("/styles/{style_id}/size-scales/{scale_id}", status_code=204)
async def delete_style_size_scale(
    style_id: int,
    scale_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(StyleSizeScale, scale_id)
    if not row or row.tenant_id != tenant.id or row.style_id != style_id:
        raise HTTPException(status_code=404, detail="Size scale not found")
    await db.delete(row)
    await db.flush()


@router.get("/boms")
async def list_boms(
    response: Response,
    style_id: int | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=MAX_PAGE_SIZE),
    offset: int = Query(default=0, ge=0),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    stmt = select(Bom).where(Bom.tenant_id == tenant.id)
    if style_id is not None:
        stmt = stmt.where(Bom.style_id == style_id)
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = int((await db.execute(count_stmt)).scalar() or 0)
    result = await db.execute(stmt.order_by(Bom.created_at.desc()).offset(offset).limit(limit))
    response.headers["X-Total-Count"] = str(total)
    return result.scalars().all()


@router.post("/boms", status_code=201)
async def create_bom(
    body: BomCreate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    payload = body.model_dump()
    payload["status"] = validate_transition(
        BOM_TRANSITIONS,
        "DRAFT",
        payload.get("status") or "DRAFT",
        fallback="DRAFT",
        entity_label="bom",
    )
    row = Bom(tenant_id=tenant.id, **payload)
    db.add(row)
    await db.flush()
    await db.refresh(row)
    return row


@router.get("/boms/{bom_id}")
async def get_bom(
    bom_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    bom = await db.get(Bom, bom_id)
    if not bom or bom.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="BOM not found")
    items = await db.execute(
        select(BomItem)
        .where(BomItem.tenant_id == tenant.id, BomItem.bom_id == bom_id)
        .order_by(BomItem.id)
    )
    return {"bom": bom, "items": items.scalars().all()}


@router.patch("/boms/{bom_id}")
async def update_bom(
    bom_id: int,
    body: BomUpdate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(Bom, bom_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="BOM not found")
    if body.version_no is not None:
        row.version_no = body.version_no
    if body.status is not None:
        row.status = validate_transition(
            BOM_TRANSITIONS,
            row.status,
            body.status,
            fallback="DRAFT",
            entity_label="bom",
        )
    if body.notes is not None:
        row.notes = body.notes
    await db.flush()
    await db.refresh(row)
    return row


@router.delete("/boms/{bom_id}", status_code=204)
async def delete_bom(
    bom_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(Bom, bom_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="BOM not found")
    if (row.status or "").upper() in GOVERNED_BOM_STATUSES:
        raise HTTPException(
            status_code=400,
            detail="Approved/Frozen BOM cannot be deleted. Create a new BOM version instead.",
        )
    await db.delete(row)
    await db.flush()


@router.post("/boms/{bom_id}/submit")
async def submit_bom(
    bom_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(Bom, bom_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="BOM not found")
    row.status = validate_transition(
        BOM_TRANSITIONS,
        row.status,
        "SUBMITTED",
        fallback="DRAFT",
        entity_label="bom",
    )
    await db.flush()
    await db.refresh(row)
    return row


@router.post("/boms/{bom_id}/approve")
async def approve_bom(
    bom_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_merch_permission(MERCH_PERMISSION_BOM_APPROVE)),
):
    _ensure_tenant(user, tenant)
    row = await db.get(Bom, bom_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="BOM not found")
    row.status = validate_transition(
        BOM_TRANSITIONS,
        row.status,
        "APPROVED",
        fallback="DRAFT",
        entity_label="bom",
    )
    await db.flush()
    await db.refresh(row)
    return row


@router.post("/boms/{bom_id}/freeze")
async def freeze_bom(
    bom_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_merch_permission(MERCH_PERMISSION_BOM_FREEZE)),
):
    _ensure_tenant(user, tenant)
    row = await db.get(Bom, bom_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="BOM not found")
    row.status = validate_transition(
        BOM_TRANSITIONS,
        row.status,
        "FROZEN",
        fallback="DRAFT",
        entity_label="bom",
    )
    await db.flush()
    await db.refresh(row)
    return row


@router.post("/boms/{bom_id}/items", status_code=201)
async def create_bom_item(
    bom_id: int,
    body: BomItemBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    bom = await db.get(Bom, bom_id)
    if not bom or bom.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="BOM not found")
    if (bom.status or "").upper() in GOVERNED_BOM_STATUSES:
        raise HTTPException(status_code=400, detail="BOM is approved/frozen and cannot be edited.")
    payload = body.model_dump()
    item_id = payload.get("item_id")
    if item_id is not None:
        item = await db.get(Item, item_id)
        if not item or item.tenant_id != tenant.id:
            raise HTTPException(status_code=404, detail="Item not found or not in tenant")
        if payload.get("item_code") is None:
            payload["item_code"] = item.item_code
        if payload.get("description") is None:
            payload["description"] = item.name or item.description
        if payload.get("uom") is None:
            unit = await db.get(ItemUnit, item.unit_id) if item.unit_id else None
            if unit and unit.tenant_id != tenant.id:
                unit = None
            payload["uom"] = unit.unit_code if unit else None
    row = BomItem(tenant_id=tenant.id, bom_id=bom_id, **payload)
    db.add(row)
    await db.flush()
    await db.refresh(row)
    return row


@router.patch("/boms/{bom_id}/items/{item_id}")
async def update_bom_item(
    bom_id: int,
    item_id: int,
    body: BomItemBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    bom = await db.get(Bom, bom_id)
    if not bom or bom.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="BOM not found")
    if (bom.status or "").upper() in GOVERNED_BOM_STATUSES:
        raise HTTPException(status_code=400, detail="BOM is approved/frozen and cannot be edited.")
    row = await db.get(BomItem, item_id)
    if not row or row.tenant_id != tenant.id or row.bom_id != bom_id:
        raise HTTPException(status_code=404, detail="BOM item not found")
    for key, value in body.model_dump().items():
        setattr(row, key, value)
    await db.flush()
    await db.refresh(row)
    return row


@router.delete("/boms/{bom_id}/items/{item_id}", status_code=204)
async def delete_bom_item(
    bom_id: int,
    item_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    bom = await db.get(Bom, bom_id)
    if not bom or bom.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="BOM not found")
    if (bom.status or "").upper() in GOVERNED_BOM_STATUSES:
        raise HTTPException(status_code=400, detail="BOM is approved/frozen and cannot be edited.")
    row = await db.get(BomItem, item_id)
    if not row or row.tenant_id != tenant.id or row.bom_id != bom_id:
        raise HTTPException(status_code=404, detail="BOM item not found")
    await db.delete(row)
    await db.flush()


class GeneratePOFromBOMBody(BaseModel):
    quantity: float
    supplier_name: str | None = None
    vendor_id: int | None = None


@router.post("/boms/{bom_id}/generate-purchase-order")
async def generate_purchase_order_from_bom(
    bom_id: int,
    body: GeneratePOFromBOMBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_merch_permission(MERCH_PERMISSION_PO_GENERATE)),
):
    """Create a draft purchase order from BOM lines that have item_id set. Qty = quantity × base_consumption × (1 + wastage_pct/100)."""
    _ensure_tenant(user, tenant)
    bom = await db.get(Bom, bom_id)
    if not bom or bom.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="BOM not found")
    if (bom.status or "").upper() not in GOVERNED_BOM_STATUSES:
        raise HTTPException(
            status_code=400,
            detail="Only APPROVED/FROZEN BOM can generate purchase order.",
        )
    result = await db.execute(
        select(BomItem)
        .where(
            BomItem.tenant_id == tenant.id,
            BomItem.bom_id == bom_id,
            BomItem.item_id.isnot(None),
        )
        .order_by(BomItem.id)
    )
    bom_lines = list(result.scalars().all())
    if not bom_lines:
        raise HTTPException(
            status_code=400,
            detail="BOM has no lines linked to inventory items. Link BOM lines to items first.",
        )
    quantity = body.quantity
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be positive")

    supplier_name = (body.supplier_name or "").strip() or "From BOM"
    vendor_id = body.vendor_id
    if vendor_id is not None:
        vendor = await db.get(Vendor, vendor_id)
        if not vendor or vendor.tenant_id != tenant.id:
            raise HTTPException(status_code=404, detail="Vendor not found")
        supplier_name = vendor.name
    else:
        vendor_id = None

    po_code = await next_tenant_code(
        db,
        model=PurchaseOrder,
        tenant_id=tenant.id,
        prefix="PO-",
        width=4,
    )
    warnings: list[str] = []

    po = PurchaseOrder(
        tenant_id=tenant.id,
        po_code=po_code,
        vendor_id=vendor_id,
        supplier_name=supplier_name,
        status="DRAFT",
        source_bom_id=bom_id,
        notes=f"Generated from BOM #{bom_id} (Style {bom.style_id}), quantity={quantity}",
    )
    db.add(po)
    await db.flush()

    for line in bom_lines:
        item = await db.get(Item, line.item_id)
        if not item or item.tenant_id != tenant.id:
            warnings.append(f"Skipped BOM line {line.id}: item not found or wrong tenant.")
            continue
        try:
            base = float(line.base_consumption or 0)
        except (TypeError, ValueError):
            base = 0.0
        try:
            wastage = float(line.wastage_pct or 0) / 100.0
        except (TypeError, ValueError):
            wastage = 0.0
        qty = quantity * base * (1.0 + wastage)
        qty_str = f"{qty:.4g}".strip()
        if qty_str == "0":
            qty_str = "0"
        unit_price = (item.default_cost or "0").strip() or "0"
        db.add(
            PurchaseOrderItem(
                tenant_id=tenant.id,
                purchase_order_id=po.id,
                item_id=line.item_id,
                quantity=qty_str,
                unit_price=unit_price,
            )
        )
    await db.commit()
    await db.refresh(po)
    return {"id": po.id, "po_code": po.po_code, "warnings": warnings}


GOVERNED_BOM_STATUSES = merch_c.GOVERNED_BOM_STATUSES


async def _get_latest_governed_bom(
    db: AsyncSession,
    *,
    tenant_id: int,
    style_id: int,
) -> Bom | None:
    result = await db.execute(
        select(Bom)
        .where(
            Bom.tenant_id == tenant_id,
            Bom.style_id == style_id,
            Bom.status.in_(GOVERNED_BOM_STATUSES),
        )
        .order_by(Bom.version_no.desc())
        .limit(1)
    )
    return result.scalar_one_or_none()


class MaterialRequirementLineOut(BaseModel):
    item_id: int
    item_code: str
    item_name: str
    uom: str | None
    required_qty: float
    available_qty: float
    shortage_qty: float


class MaterialRequirementOut(BaseModel):
    order_id: int
    order_code: str
    style_id: int
    bom_id: int
    quantity_used: float
    lines: list[MaterialRequirementLineOut]


@router.get("/orders/{order_id}/material-requirement", response_model=MaterialRequirementOut)
async def get_order_material_requirement(
    order_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Explode BOM for the order's style by order quantity; return required vs available stock per item (no persistence)."""
    _ensure_tenant(user, tenant)
    order = await db.get(Order, order_id)
    if not order or order.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Order not found")
    style_id: int | None = None
    if order.quotation_id:
        quotation = await db.get(Quotation, order.quotation_id)
        if quotation and quotation.tenant_id == tenant.id:
            style_id = quotation.style_id
    if not style_id:
        raise HTTPException(
            status_code=400,
            detail="Order has no style. Link a quotation with a style to generate material requirement.",
        )
    order_qty = _to_float_safe(str(order.quantity)) if order.quantity is not None else 0.0
    if order_qty <= 0:
        raise HTTPException(status_code=400, detail="Order quantity must be positive")
    bom = await _get_latest_governed_bom(
        db,
        tenant_id=tenant.id,
        style_id=style_id,
    )
    if not bom:
        raise HTTPException(
            status_code=400,
            detail="No APPROVED/FROZEN BOM found for this order style.",
        )
    bom_lines_result = await db.execute(
        select(BomItem)
        .where(
            BomItem.tenant_id == tenant.id,
            BomItem.bom_id == bom.id,
            BomItem.item_id.isnot(None),
        )
        .order_by(BomItem.id)
    )
    bom_lines = list(bom_lines_result.scalars().all())
    if not bom_lines:
        raise HTTPException(
            status_code=400,
            detail="BOM has no lines linked to inventory items.",
        )
    lines_out: list[MaterialRequirementLineOut] = []
    for line in bom_lines:
        item = await db.get(Item, line.item_id)
        if not item or item.tenant_id != tenant.id:
            continue
        base = _to_float_safe(line.base_consumption)
        wastage = _to_float_safe(line.wastage_pct) / 100.0
        required = order_qty * base * (1.0 + wastage)
        mov_in = await db.execute(
            select(StockMovement.quantity).where(
                StockMovement.tenant_id == tenant.id,
                StockMovement.item_id == line.item_id,
                func.upper(StockMovement.movement_type) == "IN",
            )
        )
        mov_out = await db.execute(
            select(StockMovement.quantity).where(
                StockMovement.tenant_id == tenant.id,
                StockMovement.item_id == line.item_id,
                func.upper(StockMovement.movement_type) == "OUT",
            )
        )
        in_qty = sum(_to_float_safe(q[0]) for q in mov_in.all())
        out_qty = sum(_to_float_safe(q[0]) for q in mov_out.all())
        available = round(in_qty - out_qty, 4)
        shortage = round(max(0.0, required - available), 4)
        unit_name = None
        if item.unit_id:
            unit = await db.get(ItemUnit, item.unit_id)
            if unit and unit.tenant_id != tenant.id:
                unit = None
            if unit:
                unit_name = unit.unit_code
        lines_out.append(
            MaterialRequirementLineOut(
                item_id=line.item_id,
                item_code=item.item_code,
                item_name=item.name,
                uom=unit_name or line.uom,
                required_qty=round(required, 4),
                available_qty=available,
                shortage_qty=shortage,
            )
        )
    return MaterialRequirementOut(
        order_id=order.id,
        order_code=order.order_code,
        style_id=style_id,
        bom_id=bom.id,
        quantity_used=order_qty,
        lines=lines_out,
    )


@router.get("/consumption-plans")
async def list_consumption_plans(
    response: Response,
    order_id: int | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=MAX_PAGE_SIZE),
    offset: int = Query(default=0, ge=0),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    stmt = select(ConsumptionPlan).where(ConsumptionPlan.tenant_id == tenant.id)
    if order_id is not None:
        stmt = stmt.where(ConsumptionPlan.order_id == order_id)
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = int((await db.execute(count_stmt)).scalar() or 0)
    result = await db.execute(stmt.order_by(ConsumptionPlan.created_at.desc()).offset(offset).limit(limit))
    response.headers["X-Total-Count"] = str(total)
    return result.scalars().all()


@router.post("/consumption-plans", status_code=201)
async def create_consumption_plan(
    body: ConsumptionPlanCreate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    order = await db.get(Order, body.order_id)
    if not order or order.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Order not found")
    style_id: int | None = None
    if order.quotation_id:
        quotation = await db.get(Quotation, order.quotation_id)
        if quotation and quotation.tenant_id == tenant.id:
            style_id = quotation.style_id
    if not style_id:
        raise HTTPException(status_code=400, detail="Order has no style linked for BOM-driven plan")
    bom = await _get_latest_governed_bom(db, tenant_id=tenant.id, style_id=style_id)
    if not bom:
        raise HTTPException(status_code=400, detail="No APPROVED/FROZEN BOM found for order style")
    bom_lines = (
        await db.execute(
            select(BomItem).where(
                BomItem.tenant_id == tenant.id,
                BomItem.bom_id == bom.id,
            )
        )
    ).scalars().all()

    row = ConsumptionPlan(tenant_id=tenant.id, **body.model_dump())
    db.add(row)
    await db.flush()
    order_qty = _to_float_safe(str(order.quantity)) if order.quantity is not None else 0.0
    for line in bom_lines:
        base = _to_float_safe(line.base_consumption)
        wastage = _to_float_safe(line.wastage_pct) / 100.0
        required_qty = order_qty * base * (1.0 + wastage)
        db.add(
            ConsumptionPlanItem(
                tenant_id=tenant.id,
                plan_id=row.id,
                item_code=line.item_code,
                required_qty=str(round(required_qty, 4)),
                uom=line.uom,
            )
        )
    await db.refresh(row)
    return row


@router.get("/consumption-plans/{plan_id}")
async def get_consumption_plan(
    plan_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(ConsumptionPlan, plan_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Consumption plan not found")
    items = await db.execute(
        select(ConsumptionPlanItem)
        .where(ConsumptionPlanItem.tenant_id == tenant.id, ConsumptionPlanItem.plan_id == plan_id)
        .order_by(ConsumptionPlanItem.id)
    )
    return {"plan": row, "items": items.scalars().all()}


@router.patch("/consumption-plans/{plan_id}")
async def update_consumption_plan(
    plan_id: int,
    body: ConsumptionPlanUpdate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(ConsumptionPlan, plan_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Consumption plan not found")
    if body.status is not None:
        row.status = body.status
    await db.flush()
    await db.refresh(row)
    return row


@router.delete("/consumption-plans/{plan_id}", status_code=204)
async def delete_consumption_plan(
    plan_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(ConsumptionPlan, plan_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Consumption plan not found")
    await db.delete(row)
    await db.flush()


@router.post("/consumption-plans/{plan_id}/items", status_code=201)
async def create_consumption_plan_item(
    plan_id: int,
    body: ConsumptionPlanItemBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    raise HTTPException(
        status_code=400,
        detail="Consumption plan items are BOM-driven. Use approved BOM changes/change request flow.",
    )


@router.patch("/consumption-plans/{plan_id}/items/{item_id}")
async def update_consumption_plan_item(
    plan_id: int,
    item_id: int,
    body: ConsumptionPlanItemBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    raise HTTPException(
        status_code=400,
        detail="Manual item override is disabled. Use approved BOM changes/change request flow.",
    )


@router.delete("/consumption-plans/{plan_id}/items/{item_id}", status_code=204)
async def delete_consumption_plan_item(
    plan_id: int,
    item_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    raise HTTPException(
        status_code=400,
        detail="Manual deletion is disabled. Use approved BOM changes/change request flow.",
    )


@router.get("/followups")
async def list_followups(
    response: Response,
    order_id: int | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    limit: int = Query(default=200, ge=1, le=MAX_PAGE_SIZE),
    offset: int = Query(default=0, ge=0),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    stmt = select(Followup).where(Followup.tenant_id == tenant.id)
    if order_id is not None:
        stmt = stmt.where(Followup.order_id == order_id)
    if status_filter:
        stmt = stmt.where(Followup.status == status_filter)
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = int((await db.execute(count_stmt)).scalar() or 0)
    result = await db.execute(stmt.order_by(Followup.created_at.desc()).offset(offset).limit(limit))
    response.headers["X-Total-Count"] = str(total)
    return result.scalars().all()


@router.post("/followups", status_code=201)
async def create_followup(
    body: FollowupCreate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    ord_row = await db.get(Order, body.order_id)
    if not ord_row or ord_row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Order not found")
    row = Followup(tenant_id=tenant.id, **body.model_dump())
    db.add(row)
    await db.flush()
    await db.refresh(row)
    return row


@router.patch("/followups/{followup_id}")
async def update_followup(
    followup_id: int,
    body: FollowupUpdate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(Followup, followup_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Followup not found")
    for field in ("title", "due_date", "status", "severity", "notes"):
        value = getattr(body, field)
        if value is not None:
            setattr(row, field, value)
    await db.flush()
    await db.refresh(row)
    return row


@router.delete("/followups/{followup_id}", status_code=204)
async def delete_followup(
    followup_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(Followup, followup_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Followup not found")
    await db.delete(row)
    await db.flush()


# ---------- TNA / Advanced Order Follow-up: templates and action lines ----------

TNA_PHASES = merch_c.TNA_PHASES
TNA_ACTION_STATUSES = merch_c.TNA_ACTION_STATUSES
TNA_APPROVAL_STATUSES = merch_c.TNA_APPROVAL_STATUSES
TNA_SEVERITIES = merch_c.TNA_SEVERITIES
TNA_DEFAULT_TEMPLATE_SEED = merch_c.TNA_DEFAULT_TEMPLATE_SEED


class FollowupActionTemplateOut(BaseModel):
    id: int
    code: str
    name: str
    phase: str
    action_group: str | None
    sequence_no: int
    default_days_before_delivery: int | None
    is_mandatory: bool
    is_active: bool
    buyer_id: int | None


class FollowupActionTemplateCreate(BaseModel):
    code: str
    name: str
    phase: str
    action_group: str | None = None
    sequence_no: int = 0
    default_days_before_delivery: int | None = None
    is_mandatory: bool = False
    is_active: bool = True
    buyer_id: int | None = None

    @field_validator("phase")
    @classmethod
    def _validate_phase(cls, v: str) -> str:
        s = (v or "").strip().lower()
        if s not in set(merch_c.TNA_PHASES):
            raise ValueError(f"Invalid phase. Allowed: {', '.join(merch_c.TNA_PHASES)}")
        return s


class OrderFollowupActionOut(BaseModel):
    id: int
    order_id: int
    order_code: str | None
    delivery_date: date | None
    style_code: str | None
    template_id: int | None
    sequence_no: int
    phase: str
    action_group: str | None
    action_type: str | None
    title: str
    description: str | None
    is_template_generated: bool
    is_mandatory: bool
    is_active: bool
    assigned_to_id: int | None
    planned_date: date | None
    actual_submission_date: date | None
    approval_received_date: date | None
    actual_completion_date: date | None
    resubmission_date: date | None
    status: str
    approval_status: str | None
    is_rejected: bool
    rejection_reason: str | None
    delay_reason: str | None
    severity: str | None
    remarks: str | None
    completed_at: datetime | None
    milestone_type: str | None
    external_id: int | None
    created_at: datetime
    updated_at: datetime


class FollowupActionCommentOut(BaseModel):
    id: int
    user_id: int
    username: str | None
    comment_text: str
    created_at: datetime


class FollowupActionCommentCreate(BaseModel):
    comment_text: str


class OrderFollowupActionCreate(BaseModel):
    order_id: int
    template_id: int | None = None
    sequence_no: int = 0
    phase: str
    action_group: str | None = None
    action_type: str | None = None
    title: str
    description: str | None = None
    is_mandatory: bool = False
    planned_date: date | None = None
    actual_submission_date: date | None = None
    approval_received_date: date | None = None
    resubmission_date: date | None = None
    status: str = "pending"
    approval_status: str | None = None
    is_rejected: bool = False
    rejection_reason: str | None = None
    delay_reason: str | None = None
    severity: str | None = None
    remarks: str | None = None
    assigned_to_id: int | None = None


class OrderFollowupActionUpdate(BaseModel):
    sequence_no: int | None = None
    phase: str | None = None
    action_group: str | None = None
    action_type: str | None = None
    title: str | None = None
    description: str | None = None
    planned_date: date | None = None
    actual_submission_date: date | None = None
    approval_received_date: date | None = None
    actual_completion_date: date | None = None
    resubmission_date: date | None = None
    status: str | None = None
    approval_status: str | None = None
    is_rejected: bool | None = None
    rejection_reason: str | None = None
    delay_reason: str | None = None
    severity: str | None = None
    remarks: str | None = None
    assigned_to_id: int | None = None
    milestone_type: str | None = None
    external_id: int | None = None


class FollowupSummaryOut(BaseModel):
    open_count: int
    overdue_count: int
    due_this_week_count: int
    rejected_count: int
    completed_count: int


class RejectionLogEntryOut(BaseModel):
    id: int
    rejected_at: datetime
    rejection_reason: str | None
    resubmission_date: date | None
    created_at: datetime


class RejectionLogCreate(BaseModel):
    rejection_reason: str | None = None
    resubmission_date: date | None = None


class TnaGenerateRequest(BaseModel):
    order_id: int
    template_ids: list[int] | None = None  # if None, use all active templates


async def _order_context_for_action(
    db: AsyncSession, tenant_id: int, order_id: int
) -> tuple[str | None, date | None, str | None]:
    """Return (order_code, delivery_date, style_code) for an order."""
    order = await db.get(Order, order_id)
    if not order or order.tenant_id != tenant_id:
        return None, None, None
    order_code = order.order_code
    delivery_date = order.delivery_date
    style_code = None
    if order.quotation_id:
        q = await db.get(Quotation, order.quotation_id)
        if q and q.tenant_id == tenant_id and q.style_id:
            style = await db.get(GarmentStyle, q.style_id)
            if style and style.tenant_id != tenant_id:
                style = None
            style_code = style.style_code if style else None
    return order_code, delivery_date, style_code


async def _batch_order_context_map(
    db: AsyncSession, tenant_id: int, order_ids: list[int]
) -> dict[int, tuple[str | None, date | None, str | None]]:
    """Batch-load (order_code, delivery_date, style_code) for many orders (one query set per entity type)."""
    out: dict[int, tuple[str | None, date | None, str | None]] = {}
    if not order_ids:
        return out
    uids = list({oid for oid in order_ids})
    orows = (
        await db.execute(select(Order).where(Order.tenant_id == tenant_id, Order.id.in_(uids)))
    ).scalars().all()
    qids = [o.quotation_id for o in orows if o.quotation_id]
    quotes: dict[int, Quotation] = {}
    if qids:
        qres = await db.execute(
            select(Quotation).where(Quotation.tenant_id == tenant_id, Quotation.id.in_(list({q for q in qids if q})))
        )
        for q in qres.scalars().all():
            quotes[q.id] = q
    sids = [q.style_id for q in quotes.values() if q.style_id]
    styles: dict[int, GarmentStyle] = {}
    if sids:
        sid_set = {s for s in sids if s}
        sres = await db.execute(
            select(GarmentStyle).where(GarmentStyle.tenant_id == tenant_id, GarmentStyle.id.in_(sid_set))
        )
        for s in sres.scalars().all():
            styles[s.id] = s
    for o in orows:
        sc: str | None = None
        if o.quotation_id and o.quotation_id in quotes:
            q = quotes[o.quotation_id]
            if q.style_id and q.style_id in styles:
                sc = styles[q.style_id].style_code
        out[o.id] = (o.order_code, o.delivery_date, sc)
    return out


@router.get("/followup-templates", response_model=list[FollowupActionTemplateOut])
async def list_followup_templates(
    phase: str | None = Query(default=None),
    is_active: bool | None = Query(default=None),
    buyer_id: int | None = Query(default=None, description="Filter: global (buyer_id null) or this buyer"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List TNA action templates for the tenant. Seeds default templates if empty.
    When buyer_id is set, returns templates where buyer_id is null (global) or buyer_id == buyer_id.
    """
    _ensure_tenant(user, tenant)
    stmt = select(FollowupActionTemplate).where(FollowupActionTemplate.tenant_id == tenant.id)
    if phase:
        stmt = stmt.where(FollowupActionTemplate.phase == phase)
    if is_active is not None:
        stmt = stmt.where(FollowupActionTemplate.is_active == is_active)
    if buyer_id is not None:
        stmt = stmt.where(
            or_(FollowupActionTemplate.buyer_id.is_(None), FollowupActionTemplate.buyer_id == buyer_id)
        )
    result = await db.execute(stmt.order_by(FollowupActionTemplate.sequence_no, FollowupActionTemplate.id))
    rows = result.scalars().all()
    if not rows:
        for code, name, ph, default_days, seq in TNA_DEFAULT_TEMPLATE_SEED:
            t = FollowupActionTemplate(
                tenant_id=tenant.id,
                code=code,
                name=name,
                phase=ph,
                sequence_no=seq,
                default_days_before_delivery=default_days,
                is_mandatory=False,
                is_active=True,
            )
            db.add(t)
        await db.commit()
        result = await db.execute(
            select(FollowupActionTemplate).where(FollowupActionTemplate.tenant_id == tenant.id).order_by(
                FollowupActionTemplate.sequence_no, FollowupActionTemplate.id
            )
        )
        rows = result.scalars().all()
    return [
        FollowupActionTemplateOut(
            id=r.id,
            code=r.code,
            name=r.name,
            phase=r.phase,
            action_group=r.action_group,
            sequence_no=r.sequence_no,
            default_days_before_delivery=r.default_days_before_delivery,
            is_mandatory=r.is_mandatory,
            is_active=r.is_active,
            buyer_id=r.buyer_id,
        )
        for r in rows
    ]


@router.get("/followup-templates/{template_id}", response_model=FollowupActionTemplateOut)
async def get_followup_template(
    template_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(FollowupActionTemplate, template_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Template not found")
    return FollowupActionTemplateOut(
        id=row.id, code=row.code, name=row.name, phase=row.phase, action_group=row.action_group,
        sequence_no=row.sequence_no, default_days_before_delivery=row.default_days_before_delivery,
        is_mandatory=row.is_mandatory, is_active=row.is_active, buyer_id=row.buyer_id,
    )


@router.post("/followup-templates", status_code=201, response_model=FollowupActionTemplateOut)
async def create_followup_template(
    body: FollowupActionTemplateCreate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_merch_permission(MERCH_PERMISSION_TNA_MANAGE)),
):
    _ensure_tenant(user, tenant)
    row = FollowupActionTemplate(tenant_id=tenant.id, **body.model_dump())
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return FollowupActionTemplateOut(
        id=row.id, code=row.code, name=row.name, phase=row.phase, action_group=row.action_group,
        sequence_no=row.sequence_no, default_days_before_delivery=row.default_days_before_delivery,
        is_mandatory=row.is_mandatory, is_active=row.is_active, buyer_id=row.buyer_id,
    )


class FollowupActionTemplateUpdate(BaseModel):
    name: str | None = None
    phase: str | None = None
    action_group: str | None = None
    sequence_no: int | None = None
    default_days_before_delivery: int | None = None
    is_mandatory: bool | None = None
    is_active: bool | None = None
    buyer_id: int | None = None

    @field_validator("phase")
    @classmethod
    def _validate_phase_optional(cls, v: str | None) -> str | None:
        if v is None:
            return None
        s = v.strip().lower()
        if s not in set(merch_c.TNA_PHASES):
            raise ValueError(f"Invalid phase. Allowed: {', '.join(merch_c.TNA_PHASES)}")
        return s


@router.patch("/followup-templates/{template_id}", response_model=FollowupActionTemplateOut)
async def update_followup_template(
    template_id: int,
    body: FollowupActionTemplateUpdate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_merch_permission(MERCH_PERMISSION_TNA_MANAGE)),
):
    _ensure_tenant(user, tenant)
    row = await db.get(FollowupActionTemplate, template_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Template not found")
    for f in ("name", "phase", "action_group", "sequence_no", "default_days_before_delivery", "is_mandatory", "is_active", "buyer_id"):
        v = getattr(body, f)
        if v is not None:
            setattr(row, f, v)
    await db.commit()
    await db.refresh(row)
    return FollowupActionTemplateOut(
        id=row.id, code=row.code, name=row.name, phase=row.phase, action_group=row.action_group,
        sequence_no=row.sequence_no, default_days_before_delivery=row.default_days_before_delivery,
        is_mandatory=row.is_mandatory, is_active=row.is_active, buyer_id=row.buyer_id,
    )


@router.delete("/followup-templates/{template_id}", status_code=204)
async def delete_followup_template(
    template_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_merch_permission(MERCH_PERMISSION_TNA_MANAGE)),
):
    _ensure_tenant(user, tenant)
    row = await db.get(FollowupActionTemplate, template_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Template not found")
    await db.delete(row)
    await db.commit()


@router.get("/followup-actions", response_model=list[OrderFollowupActionOut])
async def list_followup_actions(
    order_id: int | None = Query(default=None),
    status: str | None = Query(default=None),
    phase: str | None = Query(default=None),
    assigned_to_id: int | None = Query(default=None),
    due_from: date | None = Query(default=None),
    due_to: date | None = Query(default=None),
    overdue_only: bool = Query(default=False),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    today = date.today()
    stmt = select(OrderFollowupAction).where(OrderFollowupAction.tenant_id == tenant.id)
    if order_id is not None:
        stmt = stmt.where(OrderFollowupAction.order_id == order_id)
    if status:
        stmt = stmt.where(OrderFollowupAction.status == status)
    if phase:
        stmt = stmt.where(OrderFollowupAction.phase == phase)
    if assigned_to_id is not None:
        stmt = stmt.where(OrderFollowupAction.assigned_to_id == assigned_to_id)
    if due_from is not None:
        stmt = stmt.where(OrderFollowupAction.planned_date >= due_from)
    if due_to is not None:
        stmt = stmt.where(OrderFollowupAction.planned_date <= due_to)
    if overdue_only:
        stmt = stmt.where(
            OrderFollowupAction.planned_date.isnot(None),
            OrderFollowupAction.planned_date < today,
            OrderFollowupAction.status.notin_(["completed", "approved", "cancelled"]),
        )
    result = await db.execute(stmt.order_by(OrderFollowupAction.planned_date.asc().nullslast(), OrderFollowupAction.sequence_no, OrderFollowupAction.id))
    actions = result.scalars().all()
    ctx_map = await _batch_order_context_map(db, tenant.id, [a.order_id for a in actions])
    out: list[OrderFollowupActionOut] = []
    for a in actions:
        order_code, delivery_date, style_code = ctx_map.get(
            a.order_id, (None, None, None)
        )
        out.append(
            OrderFollowupActionOut(
                id=a.id,
                order_id=a.order_id,
                order_code=order_code,
                delivery_date=delivery_date,
                style_code=style_code,
                template_id=a.template_id,
                sequence_no=a.sequence_no,
                phase=a.phase,
                action_group=a.action_group,
                action_type=a.action_type,
                title=a.title,
                description=a.description,
                is_template_generated=a.is_template_generated,
                is_mandatory=a.is_mandatory,
                is_active=a.is_active,
                assigned_to_id=a.assigned_to_id,
                planned_date=a.planned_date,
                actual_submission_date=a.actual_submission_date,
                approval_received_date=a.approval_received_date,
                actual_completion_date=a.actual_completion_date,
                resubmission_date=a.resubmission_date,
                status=a.status,
                approval_status=a.approval_status,
                is_rejected=a.is_rejected,
                rejection_reason=a.rejection_reason,
                delay_reason=a.delay_reason,
                severity=a.severity,
                remarks=a.remarks,
                completed_at=a.completed_at,
                milestone_type=a.milestone_type,
                external_id=a.external_id,
                created_at=a.created_at,
                updated_at=a.updated_at,
            )
        )
    return out


@router.get("/followup-actions/summary", response_model=FollowupSummaryOut)
async def get_followup_actions_summary(
    order_id: int | None = Query(default=None),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    today = date.today()
    week_end = today + timedelta(days=7)
    stmt = select(OrderFollowupAction).where(
        OrderFollowupAction.tenant_id == tenant.id,
        OrderFollowupAction.is_active == True,
    )
    if order_id is not None:
        stmt = stmt.where(OrderFollowupAction.order_id == order_id)
    result = await db.execute(stmt)
    rows = result.scalars().all()
    open_statuses = {"pending", "in_progress", "submitted", "rejected", "resubmitted", "on_hold"}
    open_count = sum(1 for r in rows if r.status in open_statuses)
    overdue_count = sum(
        1 for r in rows
        if r.planned_date and r.planned_date < today and r.status in open_statuses
    )
    due_this_week_count = sum(
        1 for r in rows
        if r.planned_date and r.planned_date <= week_end and r.planned_date >= today and r.status in open_statuses
    )
    rejected_count = sum(1 for r in rows if r.is_rejected or r.status == "rejected")
    completed_count = sum(1 for r in rows if r.status in ("completed", "approved"))
    return FollowupSummaryOut(
        open_count=open_count,
        overdue_count=overdue_count,
        due_this_week_count=due_this_week_count,
        rejected_count=rejected_count,
        completed_count=completed_count,
    )


@router.get("/followup-actions/search", response_model=list[OrderFollowupActionOut])
async def search_followup_actions(
    q: str = Query(..., min_length=2),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    # Search in action title, description, and order code
    order_ids: list[int] = []
    order_stmt = select(Order.id).where(Order.tenant_id == tenant.id)
    if q:
        order_stmt = order_stmt.where(Order.order_code.ilike(f"%{q}%"))
    ord_result = await db.execute(order_stmt.limit(50))
    order_ids = list(ord_result.scalars().all())
    conds = [
        OrderFollowupAction.title.ilike(f"%{q}%"),
    ]
    if order_ids:
        conds.append(OrderFollowupAction.order_id.in_(order_ids))
    # Search in description (nullable)
    conds.append(and_(OrderFollowupAction.description.isnot(None), OrderFollowupAction.description.ilike(f"%{q}%")))
    stmt = (
        select(OrderFollowupAction)
        .where(OrderFollowupAction.tenant_id == tenant.id)
        .where(or_(*conds))
    )
    result = await db.execute(stmt.order_by(OrderFollowupAction.planned_date.asc().nullslast()).limit(100))
    actions = result.scalars().all()
    out: list[OrderFollowupActionOut] = []
    for a in actions:
        order_code, delivery_date, style_code = await _order_context_for_action(db, tenant.id, a.order_id)
        out.append(
            OrderFollowupActionOut(
                id=a.id, order_id=a.order_id, order_code=order_code, delivery_date=delivery_date, style_code=style_code,
                template_id=a.template_id, sequence_no=a.sequence_no, phase=a.phase, action_group=a.action_group,
                action_type=a.action_type, title=a.title, description=a.description,
                is_template_generated=a.is_template_generated, is_mandatory=a.is_mandatory, is_active=a.is_active,
                assigned_to_id=a.assigned_to_id, planned_date=a.planned_date, actual_submission_date=a.actual_submission_date,
                approval_received_date=a.approval_received_date, actual_completion_date=a.actual_completion_date,
                resubmission_date=a.resubmission_date, status=a.status, approval_status=a.approval_status,
                is_rejected=a.is_rejected, rejection_reason=a.rejection_reason, delay_reason=a.delay_reason,
                severity=a.severity, remarks=a.remarks, completed_at=a.completed_at, milestone_type=a.milestone_type, external_id=a.external_id, created_at=a.created_at, updated_at=a.updated_at,
            )
        )
    return out


@router.get("/followup-actions/order/{order_id}/timeline", response_model=list[OrderFollowupActionOut])
async def get_followup_actions_timeline(
    order_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    order = await db.get(Order, order_id)
    if not order or order.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Order not found")
    stmt = (
        select(OrderFollowupAction)
        .where(OrderFollowupAction.tenant_id == tenant.id, OrderFollowupAction.order_id == order_id)
        .order_by(OrderFollowupAction.sequence_no, OrderFollowupAction.planned_date.asc().nullslast(), OrderFollowupAction.id)
    )
    result = await db.execute(stmt)
    actions = result.scalars().all()
    order_code, delivery_date, style_code = await _order_context_for_action(db, tenant.id, order_id)
    return [
        OrderFollowupActionOut(
            id=a.id, order_id=a.order_id, order_code=order_code, delivery_date=delivery_date, style_code=style_code,
            template_id=a.template_id, sequence_no=a.sequence_no, phase=a.phase, action_group=a.action_group,
            action_type=a.action_type, title=a.title, description=a.description,
            is_template_generated=a.is_template_generated, is_mandatory=a.is_mandatory, is_active=a.is_active,
            assigned_to_id=a.assigned_to_id, planned_date=a.planned_date, actual_submission_date=a.actual_submission_date,
            approval_received_date=a.approval_received_date, actual_completion_date=a.actual_completion_date,
            resubmission_date=a.resubmission_date, status=a.status, approval_status=a.approval_status,
            is_rejected=a.is_rejected, rejection_reason=a.rejection_reason, delay_reason=a.delay_reason,
            severity=a.severity, remarks=a.remarks, completed_at=a.completed_at, milestone_type=a.milestone_type, external_id=a.external_id, created_at=a.created_at, updated_at=a.updated_at,
        )
        for a in actions
    ]


@router.post("/followup-actions/generate", response_model=list[OrderFollowupActionOut])
async def generate_followup_actions(
    body: TnaGenerateRequest,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Generate TNA action lines for an order from templates. Planned dates = delivery_date - default_days_before_delivery."""
    _ensure_tenant(user, tenant)
    order = await db.get(Order, body.order_id)
    if not order or order.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Order not found")
    delivery = order.delivery_date
    if not delivery:
        raise HTTPException(status_code=400, detail="Order has no delivery date; set delivery date first.")
    if body.template_ids:
        tstmt = select(FollowupActionTemplate).where(
            FollowupActionTemplate.tenant_id == tenant.id,
            FollowupActionTemplate.id.in_(body.template_ids),
            FollowupActionTemplate.is_active == True,
        ).order_by(FollowupActionTemplate.sequence_no)
    else:
        tstmt = select(FollowupActionTemplate).where(
            FollowupActionTemplate.tenant_id == tenant.id,
            FollowupActionTemplate.is_active == True,
        ).order_by(FollowupActionTemplate.sequence_no)
    templates = (await db.execute(tstmt)).scalars().all()
    created: list[OrderFollowupAction] = []
    for seq, t in enumerate(templates, start=1):
        planned = None
        if t.default_days_before_delivery is not None:
            planned = delivery - timedelta(days=t.default_days_before_delivery)
        action = OrderFollowupAction(
            tenant_id=tenant.id,
            order_id=body.order_id,
            template_id=t.id,
            sequence_no=seq,
            phase=t.phase,
            action_group=t.action_group,
            title=t.name,
            is_template_generated=True,
            is_mandatory=t.is_mandatory,
            is_active=True,
            planned_date=planned,
            status="pending",
        )
        db.add(action)
        created.append(action)
    await db.commit()
    for a in created:
        await db.refresh(a)
    order_code, delivery_date, style_code = await _order_context_for_action(db, tenant.id, body.order_id)
    return [
        OrderFollowupActionOut(
            id=a.id, order_id=a.order_id, order_code=order_code, delivery_date=delivery_date, style_code=style_code,
            template_id=a.template_id, sequence_no=a.sequence_no, phase=a.phase, action_group=a.action_group,
            action_type=a.action_type, title=a.title, description=a.description,
            is_template_generated=a.is_template_generated, is_mandatory=a.is_mandatory, is_active=a.is_active,
            assigned_to_id=a.assigned_to_id, planned_date=a.planned_date, actual_submission_date=a.actual_submission_date,
            approval_received_date=a.approval_received_date, actual_completion_date=a.actual_completion_date,
            resubmission_date=a.resubmission_date, status=a.status, approval_status=a.approval_status,
            is_rejected=a.is_rejected, rejection_reason=a.rejection_reason, delay_reason=a.delay_reason,
            severity=a.severity, remarks=a.remarks, completed_at=a.completed_at, milestone_type=a.milestone_type, external_id=a.external_id, created_at=a.created_at, updated_at=a.updated_at,
        )
        for a in created
    ]


@router.get("/followup-actions/overdue", response_model=list[OrderFollowupActionOut])
async def list_overdue_followup_actions(
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    today = date.today()
    stmt = (
        select(OrderFollowupAction)
        .where(
            OrderFollowupAction.tenant_id == tenant.id,
            OrderFollowupAction.is_active == True,
            OrderFollowupAction.planned_date.isnot(None),
            OrderFollowupAction.planned_date < today,
            OrderFollowupAction.status.notin_(["completed", "approved", "cancelled"]),
        )
        .order_by(OrderFollowupAction.planned_date.asc())
    )
    result = await db.execute(stmt)
    actions = result.scalars().all()
    out: list[OrderFollowupActionOut] = []
    for a in actions:
        order_code, delivery_date, style_code = await _order_context_for_action(db, tenant.id, a.order_id)
        out.append(
            OrderFollowupActionOut(
                id=a.id, order_id=a.order_id, order_code=order_code, delivery_date=delivery_date, style_code=style_code,
                template_id=a.template_id, sequence_no=a.sequence_no, phase=a.phase, action_group=a.action_group,
                action_type=a.action_type, title=a.title, description=a.description,
                is_template_generated=a.is_template_generated, is_mandatory=a.is_mandatory, is_active=a.is_active,
                assigned_to_id=a.assigned_to_id, planned_date=a.planned_date, actual_submission_date=a.actual_submission_date,
                approval_received_date=a.approval_received_date, actual_completion_date=a.actual_completion_date,
                resubmission_date=a.resubmission_date, status=a.status, approval_status=a.approval_status,
                is_rejected=a.is_rejected, rejection_reason=a.rejection_reason, delay_reason=a.delay_reason,
                severity=a.severity, remarks=a.remarks, completed_at=a.completed_at, milestone_type=a.milestone_type, external_id=a.external_id, created_at=a.created_at, updated_at=a.updated_at,
            )
        )
    return out


@router.get("/followup-actions/{action_id}/rejection-history", response_model=list[RejectionLogEntryOut])
async def get_followup_action_rejection_history(
    action_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(OrderFollowupAction, action_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Follow-up action not found")
    stmt = (
        select(FollowupActionRejectionLog)
        .where(
            FollowupActionRejectionLog.tenant_id == tenant.id,
            FollowupActionRejectionLog.action_id == action_id,
        )
        .order_by(FollowupActionRejectionLog.rejected_at.desc())
    )
    result = await db.execute(stmt)
    logs = result.scalars().all()
    return [
        RejectionLogEntryOut(
            id=log.id,
            rejected_at=log.rejected_at,
            rejection_reason=log.rejection_reason,
            resubmission_date=log.resubmission_date,
            created_at=log.created_at,
        )
        for log in logs
    ]


@router.post("/followup-actions/{action_id}/rejection-history", status_code=201, response_model=RejectionLogEntryOut)
async def add_followup_action_rejection_log(
    action_id: int,
    body: RejectionLogCreate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(OrderFollowupAction, action_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Follow-up action not found")
    log_entry = FollowupActionRejectionLog(
        tenant_id=tenant.id,
        action_id=action_id,
        rejected_at=datetime.now(timezone.utc),
        rejection_reason=body.rejection_reason,
        resubmission_date=body.resubmission_date,
        created_by_id=user.id,
    )
    db.add(log_entry)
    row.status = "rejected"
    row.is_rejected = True
    if body.rejection_reason is not None:
        row.rejection_reason = body.rejection_reason
    if body.resubmission_date is not None:
        row.resubmission_date = body.resubmission_date
    await db.commit()
    await db.refresh(log_entry)
    return RejectionLogEntryOut(
        id=log_entry.id,
        rejected_at=log_entry.rejected_at,
        rejection_reason=log_entry.rejection_reason,
        resubmission_date=log_entry.resubmission_date,
        created_at=log_entry.created_at,
    )


@router.get("/followup-actions/{action_id}", response_model=OrderFollowupActionOut)
async def get_followup_action(
    action_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(OrderFollowupAction, action_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Follow-up action not found")
    order_code, delivery_date, style_code = await _order_context_for_action(db, tenant.id, row.order_id)
    return OrderFollowupActionOut(
        id=row.id, order_id=row.order_id, order_code=order_code, delivery_date=delivery_date, style_code=style_code,
        template_id=row.template_id, sequence_no=row.sequence_no, phase=row.phase, action_group=row.action_group,
        action_type=row.action_type, title=row.title, description=row.description,
        is_template_generated=row.is_template_generated, is_mandatory=row.is_mandatory, is_active=row.is_active,
        assigned_to_id=row.assigned_to_id, planned_date=row.planned_date, actual_submission_date=row.actual_submission_date,
        approval_received_date=row.approval_received_date, actual_completion_date=row.actual_completion_date,
        resubmission_date=row.resubmission_date, status=row.status, approval_status=row.approval_status,
        is_rejected=row.is_rejected, rejection_reason=row.rejection_reason, delay_reason=row.delay_reason,
        severity=row.severity, remarks=row.remarks, completed_at=row.completed_at, milestone_type=row.milestone_type, external_id=row.external_id, created_at=row.created_at, updated_at=row.updated_at,
    )


@router.get("/followup-actions/{action_id}/comments", response_model=list[FollowupActionCommentOut])
async def get_followup_action_comments(
    action_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(OrderFollowupAction, action_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Follow-up action not found")
    stmt = (
        select(FollowupActionComment, User.username)
        .join(User, FollowupActionComment.user_id == User.id)
        .where(
            FollowupActionComment.tenant_id == tenant.id,
            FollowupActionComment.action_id == action_id,
        )
        .order_by(FollowupActionComment.created_at.desc())
    )
    result = await db.execute(stmt)
    rows = result.all()
    return [
        FollowupActionCommentOut(
            id=c.id,
            user_id=c.user_id,
            username=uname,
            comment_text=c.comment_text,
            created_at=c.created_at,
        )
        for c, uname in rows
    ]


@router.post("/followup-actions/{action_id}/comments", status_code=201, response_model=FollowupActionCommentOut)
async def create_followup_action_comment(
    action_id: int,
    body: FollowupActionCommentCreate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(OrderFollowupAction, action_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Follow-up action not found")
    comment = FollowupActionComment(
        tenant_id=tenant.id,
        action_id=action_id,
        user_id=user.id,
        comment_text=body.comment_text.strip(),
    )
    db.add(comment)
    await db.commit()
    await db.refresh(comment)
    return FollowupActionCommentOut(
        id=comment.id,
        user_id=comment.user_id,
        username=user.username,
        comment_text=comment.comment_text,
        created_at=comment.created_at,
    )


@router.post("/followup-actions", status_code=201, response_model=OrderFollowupActionOut)
async def create_followup_action(
    body: OrderFollowupActionCreate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    order = await db.get(Order, body.order_id)
    if not order or order.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Order not found")
    row = OrderFollowupAction(
        tenant_id=tenant.id,
        order_id=body.order_id,
        template_id=body.template_id,
        sequence_no=body.sequence_no,
        phase=body.phase,
        action_group=body.action_group,
        action_type=body.action_type,
        title=body.title,
        description=body.description,
        is_mandatory=body.is_mandatory,
        planned_date=body.planned_date,
        actual_submission_date=body.actual_submission_date,
        approval_received_date=body.approval_received_date,
        resubmission_date=body.resubmission_date,
        status=body.status,
        approval_status=body.approval_status,
        is_rejected=body.is_rejected,
        rejection_reason=body.rejection_reason,
        delay_reason=body.delay_reason,
        severity=body.severity,
        remarks=body.remarks,
        assigned_to_id=body.assigned_to_id,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    order_code, delivery_date, style_code = await _order_context_for_action(db, tenant.id, row.order_id)
    return OrderFollowupActionOut(
        id=row.id, order_id=row.order_id, order_code=order_code, delivery_date=delivery_date, style_code=style_code,
        template_id=row.template_id, sequence_no=row.sequence_no, phase=row.phase, action_group=row.action_group,
        action_type=row.action_type, title=row.title, description=row.description,
        is_template_generated=row.is_template_generated, is_mandatory=row.is_mandatory, is_active=row.is_active,
        assigned_to_id=row.assigned_to_id, planned_date=row.planned_date, actual_submission_date=row.actual_submission_date,
        approval_received_date=row.approval_received_date, actual_completion_date=row.actual_completion_date,
        resubmission_date=row.resubmission_date, status=row.status, approval_status=row.approval_status,
        is_rejected=row.is_rejected, rejection_reason=row.rejection_reason, delay_reason=row.delay_reason,
        severity=row.severity, remarks=row.remarks, completed_at=row.completed_at, milestone_type=row.milestone_type, external_id=row.external_id, created_at=row.created_at, updated_at=row.updated_at,
    )


@router.patch("/followup-actions/{action_id}", response_model=OrderFollowupActionOut)
async def update_followup_action(
    action_id: int,
    body: OrderFollowupActionUpdate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(OrderFollowupAction, action_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Follow-up action not found")
    for field in (
        "sequence_no", "phase", "action_group", "action_type", "title", "description",
        "planned_date", "actual_submission_date", "approval_received_date", "actual_completion_date", "resubmission_date",
        "status", "approval_status", "is_rejected", "rejection_reason", "delay_reason", "severity", "remarks", "assigned_to_id",
        "milestone_type", "external_id",
    ):
        val = getattr(body, field)
        if val is not None:
            setattr(row, field, val)
    if row.is_rejected and (body.rejection_reason is not None or row.rejection_reason):
        log_entry = FollowupActionRejectionLog(
            tenant_id=tenant.id,
            action_id=row.id,
            rejected_at=datetime.now(timezone.utc),
            rejection_reason=row.rejection_reason,
            resubmission_date=row.resubmission_date,
            created_by_id=user.id,
        )
        db.add(log_entry)
    await db.commit()
    await db.refresh(row)
    from app.modules.orders.pipeline_service import auto_advance_order_pipeline

    await auto_advance_order_pipeline(db, tenant_id=tenant.id, order_id=row.order_id)
    await db.commit()
    order_code, delivery_date, style_code = await _order_context_for_action(db, tenant.id, row.order_id)
    return OrderFollowupActionOut(
        id=row.id, order_id=row.order_id, order_code=order_code, delivery_date=delivery_date, style_code=style_code,
        template_id=row.template_id, sequence_no=row.sequence_no, phase=row.phase, action_group=row.action_group,
        action_type=row.action_type, title=row.title, description=row.description,
        is_template_generated=row.is_template_generated, is_mandatory=row.is_mandatory, is_active=row.is_active,
        assigned_to_id=row.assigned_to_id, planned_date=row.planned_date, actual_submission_date=row.actual_submission_date,
        approval_received_date=row.approval_received_date, actual_completion_date=row.actual_completion_date,
        resubmission_date=row.resubmission_date, status=row.status, approval_status=row.approval_status,
        is_rejected=row.is_rejected, rejection_reason=row.rejection_reason, delay_reason=row.delay_reason,
        severity=row.severity, remarks=row.remarks, completed_at=row.completed_at, milestone_type=row.milestone_type, external_id=row.external_id, created_at=row.created_at, updated_at=row.updated_at,
    )


@router.post("/followup-actions/{action_id}/complete", response_model=OrderFollowupActionOut)
async def complete_followup_action(
    action_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(OrderFollowupAction, action_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Follow-up action not found")
    row.status = "completed"
    row.approval_status = "approved"
    row.is_rejected = False
    row.actual_completion_date = date.today()
    row.completed_at = datetime.now(timezone.utc)
    row.completed_by_id = user.id
    await db.commit()
    await db.refresh(row)
    from app.modules.orders.pipeline_service import auto_advance_order_pipeline

    await auto_advance_order_pipeline(db, tenant_id=tenant.id, order_id=row.order_id)
    await db.commit()
    order_code, delivery_date, style_code = await _order_context_for_action(db, tenant.id, row.order_id)
    return OrderFollowupActionOut(
        id=row.id, order_id=row.order_id, order_code=order_code, delivery_date=delivery_date, style_code=style_code,
        template_id=row.template_id, sequence_no=row.sequence_no, phase=row.phase, action_group=row.action_group,
        action_type=row.action_type, title=row.title, description=row.description,
        is_template_generated=row.is_template_generated, is_mandatory=row.is_mandatory, is_active=row.is_active,
        assigned_to_id=row.assigned_to_id, planned_date=row.planned_date, actual_submission_date=row.actual_submission_date,
        approval_received_date=row.approval_received_date, actual_completion_date=row.actual_completion_date,
        resubmission_date=row.resubmission_date, status=row.status, approval_status=row.approval_status,
        is_rejected=row.is_rejected, rejection_reason=row.rejection_reason, delay_reason=row.delay_reason,
        severity=row.severity, remarks=row.remarks, completed_at=row.completed_at, milestone_type=row.milestone_type, external_id=row.external_id, created_at=row.created_at, updated_at=row.updated_at,
    )


@router.post("/followup-actions/{action_id}/reopen", response_model=OrderFollowupActionOut)
async def reopen_followup_action(
    action_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(OrderFollowupAction, action_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Follow-up action not found")
    row.status = "pending"
    row.approval_status = "pending"
    row.actual_completion_date = None
    row.completed_at = None
    row.completed_by_id = None
    await db.commit()
    await db.refresh(row)
    order_code, delivery_date, style_code = await _order_context_for_action(db, tenant.id, row.order_id)
    return OrderFollowupActionOut(
        id=row.id, order_id=row.order_id, order_code=order_code, delivery_date=delivery_date, style_code=style_code,
        template_id=row.template_id, sequence_no=row.sequence_no, phase=row.phase, action_group=row.action_group,
        action_type=row.action_type, title=row.title, description=row.description,
        is_template_generated=row.is_template_generated, is_mandatory=row.is_mandatory, is_active=row.is_active,
        assigned_to_id=row.assigned_to_id, planned_date=row.planned_date, actual_submission_date=row.actual_submission_date,
        approval_received_date=row.approval_received_date, actual_completion_date=row.actual_completion_date,
        resubmission_date=row.resubmission_date, status=row.status, approval_status=row.approval_status,
        is_rejected=row.is_rejected, rejection_reason=row.rejection_reason, delay_reason=row.delay_reason,
        severity=row.severity, remarks=row.remarks, completed_at=row.completed_at, milestone_type=row.milestone_type, external_id=row.external_id, created_at=row.created_at, updated_at=row.updated_at,
    )


@router.delete("/followup-actions/{action_id}", status_code=204)
async def delete_followup_action(
    action_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    row = await db.get(OrderFollowupAction, action_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Follow-up action not found")
    await db.delete(row)
    await db.commit()


# ---------- Pipeline: stages config (research-based order lifecycle + win probability) ----------

PIPELINE_STAGES = [
    {"stage_key": "inquiry_draft", "label": "Inquiry · Draft", "document_type": "inquiry", "status_value": "DRAFT", "win_probability": 5, "sort_order": 1},
    {"stage_key": "inquiry_submitted", "label": "Inquiry · Submitted", "document_type": "inquiry", "status_value": "SUBMITTED", "win_probability": 15, "sort_order": 2},
    {"stage_key": "inquiry_converted", "label": "Inquiry · Converted", "document_type": "inquiry", "status_value": "CONVERTED", "win_probability": 100, "sort_order": 3},
    {"stage_key": "quotation_draft", "label": "Quotation · Draft", "document_type": "quotation", "status_value": "DRAFT", "win_probability": 20, "sort_order": 4},
    {"stage_key": "quotation_new", "label": "Quotation · New", "document_type": "quotation", "status_value": "NEW", "win_probability": 25, "sort_order": 5},
    {"stage_key": "quotation_submitted", "label": "Quotation · Submitted", "document_type": "quotation", "status_value": "SUBMITTED", "win_probability": 35, "sort_order": 6},
    {"stage_key": "quotation_approved", "label": "Quotation · Approved", "document_type": "quotation", "status_value": "APPROVED", "win_probability": 50, "sort_order": 7},
    {"stage_key": "quotation_sent", "label": "Quotation · Sent", "document_type": "quotation", "status_value": "SENT", "win_probability": 60, "sort_order": 8},
    {"stage_key": "quotation_converted", "label": "Quotation · Converted", "document_type": "quotation", "status_value": "CONVERTED", "win_probability": 100, "sort_order": 9},
    {"stage_key": "order_draft", "label": "Order · Draft", "document_type": "order", "status_value": "DRAFT", "win_probability": 70, "sort_order": 10},
    {"stage_key": "order_new", "label": "Order · New", "document_type": "order", "status_value": "NEW", "win_probability": 80, "sort_order": 11},
    {"stage_key": "order_in_progress", "label": "Order · In Progress", "document_type": "order", "status_value": "IN_PROGRESS", "win_probability": 90, "sort_order": 12},
    {"stage_key": "order_completed", "label": "Order · Completed", "document_type": "order", "status_value": "COMPLETED", "win_probability": 100, "sort_order": 13},
]


class PipelineStageOut(BaseModel):
    stage_key: str
    label: str
    document_type: str
    status_value: str
    win_probability: int
    sort_order: int


class PipelineItemOut(BaseModel):
    document_type: str  # inquiry | quotation | order
    id: int
    code: str
    stage_key: str
    customer_id: int
    customer_name: str
    style_ref: str | None
    style_name: str | None
    quantity: int | None
    total_amount: str | None
    created_at: str
    detail_path: str
    next_status_options: list[str]  # allowed next statuses for "Move to" action


def _inquiry_stage_key(status: str) -> str:
    key = (status or "DRAFT").upper()
    if key == "CONVERTED":
        return "inquiry_converted"
    if key == "SUBMITTED":
        return "inquiry_submitted"
    return "inquiry_draft"


def _quotation_stage_key(status: str) -> str:
    key = (status or "DRAFT").upper()
    for s in PIPELINE_STAGES:
        if s["document_type"] == "quotation" and s["status_value"] == key:
            return s["stage_key"]
    return "quotation_draft"


def _order_stage_key(status: str) -> str:
    key = (status or "DRAFT").upper()
    for s in PIPELINE_STAGES:
        if s["document_type"] == "order" and s["status_value"] == key:
            return s["stage_key"]
    return "order_draft"


@router.get("/pipeline")
async def get_pipeline_summary(
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    inq_count = (
        await db.execute(select(func.count()).select_from(Inquiry).where(Inquiry.tenant_id == tenant.id))
    ).scalar() or 0
    qt_count = (
        await db.execute(select(func.count()).select_from(Quotation).where(Quotation.tenant_id == tenant.id))
    ).scalar() or 0
    ord_count = (
        await db.execute(select(func.count()).select_from(Order).where(Order.tenant_id == tenant.id))
    ).scalar() or 0
    return {
        "inquiries": inq_count,
        "quotations": qt_count,
        "orders": ord_count,
    }


@router.get("/pipeline/full", response_model=dict)
async def get_pipeline_full(
    document_type: str | None = Query(default=None, description="Filter: inquiry, quotation, order"),
    customer_id: int | None = Query(default=None),
    search: str | None = Query(default=None, description="Search code, style_ref"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Advanced pipeline: stages with win probability + all items grouped by stage for Kanban/list views."""
    _ensure_tenant(user, tenant)
    from app.models import Customer

    stages_out = [PipelineStageOut(**s) for s in PIPELINE_STAGES]
    items_out: list[PipelineItemOut] = []

    # Inquiries with customer name
    inq_stmt = select(Inquiry, Customer.name).join(
        Customer, Inquiry.customer_id == Customer.id
    ).where(Inquiry.tenant_id == tenant.id)
    if document_type == "order" or document_type == "quotation":
        inq_stmt = inq_stmt.where(1 == 0)  # exclude
    if customer_id is not None:
        inq_stmt = inq_stmt.where(Inquiry.customer_id == customer_id)
    if search:
        pat = f"%{search}%"
        inq_stmt = inq_stmt.where(
            or_(Inquiry.inquiry_code.ilike(pat), (Inquiry.style_ref or "").ilike(pat))
        )
    inq_result = await db.execute(inq_stmt.order_by(Inquiry.created_at.desc()))
    for inq, cust_name in inq_result.all():
        stage_key = _inquiry_stage_key(inq.status)
        next_options = next_status_options(
            INQUIRY_TRANSITIONS,
            inq.status,
            fallback="DRAFT",
        )
        # Conversion to quotation is a dedicated action, not a direct status update action.
        next_options = [opt for opt in next_options if opt != "CONVERTED"]
        items_out.append(
            PipelineItemOut(
                document_type="inquiry",
                id=inq.id,
                code=inq.inquiry_code,
                stage_key=stage_key,
                customer_id=inq.customer_id,
                customer_name=cust_name or "",
                style_ref=inq.style_ref,
                style_name=None,
                quantity=inq.quantity,
                total_amount=None,
                created_at=inq.created_at.isoformat(),
                detail_path=f"/app/inquiries/{inq.id}",
                next_status_options=next_options,
            )
        )

    # Quotations with customer name
    qt_stmt = select(Quotation, Customer.name).join(
        Customer, Quotation.customer_id == Customer.id
    ).where(Quotation.tenant_id == tenant.id)
    if document_type == "inquiry" or document_type == "order":
        qt_stmt = qt_stmt.where(1 == 0)
    if customer_id is not None:
        qt_stmt = qt_stmt.where(Quotation.customer_id == customer_id)
    if search:
        pat = f"%{search}%"
        qt_stmt = qt_stmt.where(
            or_(Quotation.quotation_code.ilike(pat), (Quotation.style_ref or "").ilike(pat))
        )
    qt_result = await db.execute(qt_stmt.order_by(Quotation.created_at.desc()))
    for qt, cust_name in qt_result.all():
        stage_key = _quotation_stage_key(qt.status)
        next_options = next_status_options(
            QUOTATION_TRANSITIONS,
            qt.status,
            fallback="DRAFT",
        )
        # Conversion to order is a dedicated action, not a direct status update action.
        next_options = [opt for opt in next_options if opt != "CONVERTED"]
        items_out.append(
            PipelineItemOut(
                document_type="quotation",
                id=qt.id,
                code=qt.quotation_code,
                stage_key=stage_key,
                customer_id=qt.customer_id,
                customer_name=cust_name or "",
                style_ref=qt.style_ref,
                style_name=None,
                quantity=qt.projected_quantity,
                total_amount=qt.total_amount,
                created_at=qt.created_at.isoformat(),
                detail_path=f"/app/quotations/{qt.id}",
                next_status_options=next_options,
            )
        )

    # Orders with customer name
    ord_stmt = select(Order, Customer.name).join(
        Customer, Order.customer_id == Customer.id
    ).where(Order.tenant_id == tenant.id)
    if document_type == "inquiry" or document_type == "quotation":
        ord_stmt = ord_stmt.where(1 == 0)
    if customer_id is not None:
        ord_stmt = ord_stmt.where(Order.customer_id == customer_id)
    if search:
        pat = f"%{search}%"
        ord_stmt = ord_stmt.where(
            or_(Order.order_code.ilike(pat), (Order.style_ref or "").ilike(pat))
        )
    ord_result = await db.execute(ord_stmt.order_by(Order.created_at.desc()))
    for ord_row, cust_name in ord_result.all():
        stage_key = _order_stage_key(ord_row.status)
        next_options = next_status_options(
            ORDER_TRANSITIONS,
            ord_row.status,
            fallback="DRAFT",
        )
        items_out.append(
            PipelineItemOut(
                document_type="order",
                id=ord_row.id,
                code=ord_row.order_code,
                stage_key=stage_key,
                customer_id=ord_row.customer_id,
                customer_name=cust_name or "",
                style_ref=ord_row.style_ref,
                style_name=None,
                quantity=ord_row.quantity,
                total_amount=None,
                created_at=ord_row.created_at.isoformat(),
                detail_path=f"/app/orders/{ord_row.id}",
                next_status_options=next_options,
            )
        )

    summary = {
        "inquiries": len([i for i in items_out if i.document_type == "inquiry"]),
        "quotations": len([i for i in items_out if i.document_type == "quotation"]),
        "orders": len([i for i in items_out if i.document_type == "order"]),
    }
    return {
        "stages": [s.model_dump() for s in stages_out],
        "items": [i.model_dump() for i in items_out],
        "summary": summary,
    }


# ---------- Pipeline analytics: month-wise and quarterly (for marketing) ----------


class PipelineAnalyticsBucket(BaseModel):
    """One period (month or quarter) with key metrics."""
    period_key: str  # e.g. "2025-01" or "2025-Q1"
    period_label: str  # e.g. "Jan 2025" or "Q1 2025"
    year: int
    month: int | None  # 1-12 for monthly; None for quarter
    quarter: int | None  # 1-4 for quarterly; None for monthly
    inquiries_received: int
    confirmed_orders_count: int
    confirmed_orders_quantity: int
    inquiry_under_processing: int
    potential_orders_count: int


class PipelineAnalyticsResponse(BaseModel):
    by_month: list[PipelineAnalyticsBucket]
    by_quarter: list[PipelineAnalyticsBucket]
    summary: dict


def _month_key(d: date | datetime) -> tuple[int, int]:
    if isinstance(d, datetime):
        return (d.year, d.month)
    return (d.year, d.month)


def _quarter_key(d: date | datetime) -> tuple[int, int]:
    if isinstance(d, datetime):
        y, m = d.year, d.month
    else:
        y, m = d.year, d.month
    q = (m - 1) // 3 + 1
    return (y, q)


@router.get("/pipeline/analytics", response_model=PipelineAnalyticsResponse)
async def get_pipeline_analytics(
    years_back: int = Query(default=2, ge=0, le=5, description="Years to look back from current"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Month-wise and quarterly pipeline picture for marketing:
    - Inquiries received (by created month/quarter)
    - Confirmed orders by expected delivery date (count + quantity)
    - Inquiry under processing (inquiries not yet led to an order)
    - Potential orders (quotations SENT/APPROVED not yet converted)
    """
    _ensure_tenant(user, tenant)

    today = date.today()
    # Inquiries: id, created_at
    inq_result = await db.execute(
        select(Inquiry.id, Inquiry.created_at).where(Inquiry.tenant_id == tenant.id)
    )
    inquiries = [(r[0], r[1]) for r in inq_result.all() if r[1]]

    # Quotations: id, inquiry_id, status, created_at
    qt_result = await db.execute(
        select(Quotation.id, Quotation.inquiry_id, Quotation.status, Quotation.created_at).where(
            Quotation.tenant_id == tenant.id
        )
    )
    quotations = qt_result.all()

    # Orders: id, quotation_id, status, delivery_date, order_date, quantity
    ord_result = await db.execute(
        select(Order.id, Order.quotation_id, Order.status, Order.delivery_date, Order.order_date, Order.quantity).where(
            Order.tenant_id == tenant.id
        )
    )
    orders = list(ord_result.all())

    # Inquiry IDs that have at least one order (via any quotation)
    quotation_ids_with_order = {o[1] for o in orders if o[1] is not None}
    inquiry_ids_converted: set[int] = set()
    for q in quotations:
        if q[0] in quotation_ids_with_order and q[1] is not None:
            inquiry_ids_converted.add(q[1])

    # Quotation IDs that are converted (have an order)
    converted_quotation_ids = {o[1] for o in orders if o[1] is not None}

    # Build month buckets: (year, month) -> counts
    month_inq = defaultdict(int)
    month_ord_count = defaultdict(int)
    month_ord_qty = defaultdict(int)
    month_inq_under = defaultdict(int)
    month_potential = defaultdict(int)

    for inq_id, created in inquiries:
        if not created:
            continue
        y, m = created.year, created.month
        month_inq[(y, m)] += 1
        if inq_id not in inquiry_ids_converted:
            month_inq_under[(y, m)] += 1

    for o in orders:
        status_val = (o[2] or "").upper()
        if status_val == "DRAFT":
            continue
        deliv = o[3] or o[4]
        if not deliv:
            continue
        y, m = _month_key(deliv)
        month_ord_count[(y, m)] += 1
        month_ord_qty[(y, m)] += (o[5] or 0)

    for q in quotations:
        if q[2] in ("SENT", "APPROVED") and q[0] not in converted_quotation_ids:
            if q[3]:
                y, m = q[3].year, q[3].month
                month_potential[(y, m)] += 1

    # Quarter buckets
    quarter_inq = defaultdict(int)
    quarter_ord_count = defaultdict(int)
    quarter_ord_qty = defaultdict(int)
    quarter_inq_under = defaultdict(int)
    quarter_potential = defaultdict(int)

    for inq_id, created in inquiries:
        if not created:
            continue
        y, q = _quarter_key(created)
        quarter_inq[(y, q)] += 1
        if inq_id not in inquiry_ids_converted:
            quarter_inq_under[(y, q)] += 1

    for o in orders:
        status_val = (o[2] or "").upper()
        if status_val == "DRAFT":
            continue
        deliv = o[3] or o[4]
        if not deliv:
            continue
        y, q = _quarter_key(deliv)
        quarter_ord_count[(y, q)] += 1
        quarter_ord_qty[(y, q)] += (o[5] or 0)

    for q in quotations:
        if q[2] in ("SENT", "APPROVED") and q[0] not in converted_quotation_ids:
            if q[3]:
                y, qq = _quarter_key(q[3])
                quarter_potential[(y, qq)] += 1

    month_names = "Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split()

    all_month_keys = set(month_inq.keys()) | set(month_ord_count.keys()) | set(month_inq_under.keys()) | set(month_potential.keys())
    start_year = today.year - years_back
    # Limit to last N years and up to 6 months ahead for delivery
    end_month_linear = today.year * 12 + today.month + 6
    by_month_out: list[PipelineAnalyticsBucket] = []
    for y in range(start_year, today.year + 2):
        for m in range(1, 13):
            if y * 12 + m < (today.year - years_back) * 12 + 1:
                continue
            if y * 12 + m > end_month_linear:
                continue
            if (y, m) > (today.year, today.month) and (y, m) not in all_month_keys:
                continue
            period_key = f"{y}-{m:02d}"
            period_label = f"{month_names[m - 1]} {y}"
            by_month_out.append(
                PipelineAnalyticsBucket(
                    period_key=period_key,
                    period_label=period_label,
                    year=y,
                    month=m,
                    quarter=None,
                    inquiries_received=month_inq.get((y, m), 0),
                    confirmed_orders_count=month_ord_count.get((y, m), 0),
                    confirmed_orders_quantity=month_ord_qty.get((y, m), 0),
                    inquiry_under_processing=month_inq_under.get((y, m), 0),
                    potential_orders_count=month_potential.get((y, m), 0),
                )
            )
    by_month_out.sort(key=lambda x: (x.year, x.month or 0))
    # Trim to only include months that have data or are recent (last N years)
    cutoff = today - timedelta(days=365 * years_back)
    by_month_out = [b for b in by_month_out if (b.year, b.month or 0) >= (cutoff.year, cutoff.month) or any([
        month_inq.get((b.year, b.month or 0), 0) > 0,
        month_ord_count.get((b.year, b.month or 0), 0) > 0,
        month_inq_under.get((b.year, b.month or 0), 0) > 0,
        month_potential.get((b.year, b.month or 0), 0) > 0,
    ])]

    all_quarter_keys = set(quarter_inq.keys()) | set(quarter_ord_count.keys()) | set(quarter_inq_under.keys()) | set(quarter_potential.keys())
    by_quarter_out: list[PipelineAnalyticsBucket] = []
    for y in range(start_year, today.year + 2):
        for q in range(1, 5):
            if (y, q) not in all_quarter_keys and (y, q) > (today.year, (today.month - 1) // 3 + 1):
                continue
            period_key = f"{y}-Q{q}"
            period_label = f"Q{q} {y}"
            by_quarter_out.append(
                PipelineAnalyticsBucket(
                    period_key=period_key,
                    period_label=period_label,
                    year=y,
                    month=None,
                    quarter=q,
                    inquiries_received=quarter_inq.get((y, q), 0),
                    confirmed_orders_count=quarter_ord_count.get((y, q), 0),
                    confirmed_orders_quantity=quarter_ord_qty.get((y, q), 0),
                    inquiry_under_processing=quarter_inq_under.get((y, q), 0),
                    potential_orders_count=quarter_potential.get((y, q), 0),
                )
            )
    by_quarter_out.sort(key=lambda x: (x.year, x.quarter or 0))

    total_inq = sum(month_inq.values())
    total_ord_count = sum(month_ord_count.values())
    total_ord_qty = sum(month_ord_qty.values())
    total_under = len([i for i in inquiries if i[0] not in inquiry_ids_converted])
    total_potential = len([q for q in quotations if q[2] in ("SENT", "APPROVED") and q[0] not in converted_quotation_ids])

    return PipelineAnalyticsResponse(
        by_month=by_month_out,
        by_quarter=by_quarter_out,
        summary={
            "inquiries_received_total": total_inq,
            "confirmed_orders_total": total_ord_count,
            "confirmed_orders_quantity_total": total_ord_qty,
            "inquiry_under_processing_total": total_under,
            "potential_orders_total": total_potential,
        },
    )


# ---------- Phase E: Wastage reporting ----------


DEFAULT_WASTAGE_THRESHOLD_PCT = merch_c.DEFAULT_WASTAGE_THRESHOLD_PCT


class WastageReportRowOut(BaseModel):
    order_id: int
    order_code: str
    order_date: date | None
    delivery_date: date | None
    buyer_id: int
    buyer_name: str
    style_id: int
    style_code: str
    item_id: int
    item_code: str
    item_name: str
    category: str  # fabric | trim | other
    expected_qty: float
    actual_qty: float
    wastage_pct_vs_bom: float  # (actual - expected) / expected * 100 when expected > 0
    wastage_value: float  # max(0, actual - expected) * unit_cost
    allowed_threshold_pct: float
    threshold_breach: bool


def _wastage_category_from_item(item: Item, category: ItemCategory | None) -> str:
    if not category:
        return "other"
    code = (category.category_code or "").upper()
    if "FABRIC" in code or code.startswith("FAB"):
        return "fabric"
    if "TRIM" in code or "PACK" in code or "ACCESSORY" in code:
        return "trim"
    return "other"


def _resolve_wastage_threshold_pct(
    rules: list[WastageThresholdRule],
    customer_id: int | None,
) -> float:
    """Resolve allowed_pct from threshold rules: buyer match first, then tenant-wide. Default 15%."""
    allowed = DEFAULT_WASTAGE_THRESHOLD_PCT
    for r in rules:
        if r.scope_type == "tenant" and r.scope_id is None:
            allowed = float(r.allowed_pct)
        if r.scope_type == "buyer" and r.scope_id is not None and r.scope_id == customer_id:
            allowed = float(r.allowed_pct)
            break
    return allowed


@router.get("/reports/wastage", response_model=list[WastageReportRowOut])
async def get_wastage_report(
    order_id: int | None = Query(default=None, description="Filter by order"),
    style_id: int | None = Query(default=None, description="Filter by style"),
    buyer_id: int | None = Query(default=None, description="Filter by buyer (customer)"),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    threshold_pct: float | None = Query(default=None, description="Only rows where wastage % above this"),
    above_threshold_only: bool = Query(default=False, description="Only rows above default 15% threshold"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Report actual issued consumption vs BOM expected by order/item. Actual = CONSUMPTION_ISSUE movements; expected = order qty × BOM base × (1 + wastage%). Returns wastage value and threshold breach for profitability control."""
    _ensure_tenant(user, tenant)
    # Load threshold rules for tenant (used to resolve allowed_pct per order)
    thr_result = await db.execute(
        select(WastageThresholdRule).where(WastageThresholdRule.tenant_id == tenant.id)
    )
    threshold_rules = list(thr_result.scalars().all())
    mov_result = await db.execute(
        select(StockMovement.reference_id).where(
            StockMovement.tenant_id == tenant.id,
            StockMovement.reference_type == "CONSUMPTION_ISSUE",
            StockMovement.reference_id.isnot(None),
        ).distinct()
    )
    oids = [x for x in mov_result.scalars().all() if x is not None]
    if not oids:
        return []
    stmt = select(Order).where(Order.tenant_id == tenant.id, Order.id.in_(oids))
    if order_id is not None:
        stmt = stmt.where(Order.id == order_id)
    if buyer_id is not None:
        stmt = stmt.where(Order.customer_id == buyer_id)
    if date_from is not None:
        stmt = stmt.where(Order.order_date >= date_from)
    if date_to is not None:
        stmt = stmt.where(Order.order_date <= date_to)
    orders_result = await db.execute(stmt)
    orders = list(orders_result.scalars().all())
    if not orders:
        return []
    out: list[WastageReportRowOut] = []

    qids = {o.quotation_id for o in orders if o.quotation_id}
    quotations: dict[int, Quotation] = {}
    if qids:
        qr = await db.execute(select(Quotation).where(Quotation.id.in_(qids)))
        quotations = {q.id: q for q in qr.scalars().all()}
    style_ids = {q.style_id for q in quotations.values() if q.style_id}
    styles_map: dict[int, GarmentStyle] = {}
    if style_ids:
        sr = await db.execute(select(GarmentStyle).where(GarmentStyle.id.in_(style_ids)))
        styles_map = {s.id: s for s in sr.scalars().all()}
    cids = {o.customer_id for o in orders if o.customer_id}
    customers: dict[int, Customer] = {}
    if cids:
        cr = await db.execute(select(Customer).where(Customer.id.in_(cids)))
        customers = {c.id: c for c in cr.scalars().all()}

    bom_by_style: dict[int, Bom] = {}
    if style_ids:
        br = await db.execute(
            select(Bom).where(
                Bom.tenant_id == tenant.id,
                Bom.style_id.in_(style_ids),
                Bom.status.in_(GOVERNED_BOM_STATUSES),
            )
        )
        for b in br.scalars().all():
            cur = bom_by_style.get(b.style_id)
            if cur is None or (b.version_no or 0) > (cur.version_no or 0):
                bom_by_style[b.style_id] = b

    bom_ids = [b.id for b in bom_by_style.values()]
    bom_lines_by_bom: dict[int, list[BomItem]] = defaultdict(list)
    if bom_ids:
        blr = await db.execute(
            select(BomItem).where(
                BomItem.tenant_id == tenant.id,
                BomItem.bom_id.in_(bom_ids),
                BomItem.item_id.isnot(None),
            )
        )
        for bi in blr.scalars().all():
            bom_lines_by_bom[bi.bom_id].append(bi)

    all_item_ids = {bi.item_id for lines in bom_lines_by_bom.values() for bi in lines if bi.item_id}
    items_map: dict[int, Item] = {}
    if all_item_ids:
        ir = await db.execute(select(Item).where(Item.tenant_id == tenant.id, Item.id.in_(all_item_ids)))
        items_map = {i.id: i for i in ir.scalars().all()}
    cat_ids = {it.category_id for it in items_map.values() if it.category_id}
    categories: dict[int, ItemCategory] = {}
    if cat_ids:
        catr = await db.execute(select(ItemCategory).where(ItemCategory.id.in_(cat_ids)))
        categories = {c.id: c for c in catr.scalars().all()}

    order_id_list = [o.id for o in orders]
    mov_map: dict[tuple[int, int], list[StockMovement]] = defaultdict(list)
    if order_id_list:
        movr = await db.execute(
            select(StockMovement).where(
                StockMovement.tenant_id == tenant.id,
                StockMovement.reference_type == "CONSUMPTION_ISSUE",
                StockMovement.reference_id.in_(order_id_list),
            )
        )
        for m in movr.scalars().all():
            if m.reference_id is not None and m.item_id is not None:
                mov_map[(m.reference_id, m.item_id)].append(m)

    for order in orders:
        allowed_pct = _resolve_wastage_threshold_pct(threshold_rules, order.customer_id)
        if not order.quotation_id:
            continue
        quotation = quotations.get(order.quotation_id)
        if not quotation or quotation.tenant_id != tenant.id or not quotation.style_id:
            continue
        sid = quotation.style_id
        if style_id is not None and sid != style_id:
            continue
        style = styles_map.get(sid)
        style_code = style.style_code if style else str(sid)
        customer = customers.get(order.customer_id) if order.customer_id else None
        buyer_name = customer.name if customer else f"Customer #{order.customer_id}"
        order_qty = _to_float_safe(str(order.quantity)) if order.quantity is not None else 0.0
        if order_qty <= 0:
            continue
        bom = bom_by_style.get(sid)
        if not bom:
            continue
        for line in bom_lines_by_bom.get(bom.id, []):
            item = items_map.get(line.item_id)
            if not item or item.tenant_id != tenant.id:
                continue
            cat = categories.get(item.category_id) if item.category_id else None
            category = _wastage_category_from_item(item, cat)
            base = _to_float_safe(line.base_consumption)
            wastage = _to_float_safe(line.wastage_pct) / 100.0
            expected = order_qty * base * (1.0 + wastage)
            actual = sum(
                _to_float_safe(m.quantity)
                for m in mov_map.get((order.id, line.item_id), [])
                if (m.movement_type or "").upper() == "OUT"
            )
            if expected <= 0:
                wastage_pct = 0.0
            else:
                wastage_pct = round((actual - expected) / expected * 100.0, 2)
            use_threshold = threshold_pct if threshold_pct is not None else (allowed_pct if above_threshold_only else None)
            if use_threshold is not None and wastage_pct < use_threshold:
                continue
            unit_cost = _to_float_safe(item.default_cost)
            variance_qty = max(0.0, actual - expected)
            wastage_value = round(variance_qty * unit_cost, 2)
            threshold_breach = wastage_pct > allowed_pct
            out.append(
                WastageReportRowOut(
                    order_id=order.id,
                    order_code=order.order_code,
                    order_date=order.order_date,
                    delivery_date=order.delivery_date,
                    buyer_id=order.customer_id,
                    buyer_name=buyer_name,
                    style_id=sid,
                    style_code=style_code,
                    item_id=line.item_id,
                    item_code=item.item_code,
                    item_name=item.name or item.description or "",
                    category=category,
                    expected_qty=round(expected, 4),
                    actual_qty=round(actual, 4),
                    wastage_pct_vs_bom=wastage_pct,
                    wastage_value=wastage_value,
                    allowed_threshold_pct=allowed_pct,
                    threshold_breach=threshold_breach,
                )
            )
    return out


@router.get("/reports/wastage/summary")
async def get_wastage_summary(
    style_id: int | None = Query(default=None),
    buyer_id: int | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Aggregate wastage: KPIs (total value, fabric/trim avg %, above-threshold count) and by_style."""
    _ensure_tenant(user, tenant)
    rows = await get_wastage_report(
        order_id=None,
        style_id=style_id,
        buyer_id=buyer_id,
        date_from=date_from,
        date_to=date_to,
        threshold_pct=None,
        above_threshold_only=False,
        tenant=tenant,
        user=user,
        db=db,
    )
    total_wastage_value = sum(r.wastage_value for r in rows)
    fabric_pcts = [r.wastage_pct_vs_bom for r in rows if r.category == "fabric"]
    trim_pcts = [r.wastage_pct_vs_bom for r in rows if r.category == "trim"]
    fabric_wastage_pct_avg = round(sum(fabric_pcts) / len(fabric_pcts), 2) if fabric_pcts else 0.0
    trim_wastage_pct_avg = round(sum(trim_pcts) / len(trim_pcts), 2) if trim_pcts else 0.0
    above_threshold_orders_count = len({r.order_id for r in rows if r.threshold_breach})
    by_style: dict[int, list[float]] = {}
    for r in rows:
        by_style.setdefault(r.style_id, []).append(r.wastage_pct_vs_bom)
    by_style_list = [
        {
            "style_id": sid,
            "order_item_count": len(pcts),
            "avg_wastage_pct": round(sum(pcts) / len(pcts), 2) if pcts else 0,
            "max_wastage_pct": round(max(pcts), 2) if pcts else 0,
        }
        for sid, pcts in by_style.items()
    ]
    return {
        "total_wastage_value": round(total_wastage_value, 2),
        "fabric_wastage_pct_avg": fabric_wastage_pct_avg,
        "trim_wastage_pct_avg": trim_wastage_pct_avg,
        "above_threshold_orders_count": above_threshold_orders_count,
        "by_style": by_style_list,
        "total_rows": len(rows),
    }


class WastageReasonOut(BaseModel):
    id: int
    code: str
    name: str
    category: str
    recoverable: bool


# Default wastage reason taxonomy (seed when tenant has none). Same as migration 059.
WASTAGE_REASON_SEED: list[tuple[str, str, str, bool]] = [
    ("marker_cutting", "Marker / cutting wastage", "fabric", False),
    ("spreading", "Spreading wastage", "fabric", False),
    ("end_bit_remnant", "End-bit / remnant wastage", "fabric", True),
    ("thread_overconsumption", "Thread overconsumption", "trim", False),
    ("rework_loss", "Rework loss", "process", False),
    ("rejection_loss", "Rejection loss", "process", False),
    ("excess_issue_vs_standard", "Excess issue vs standard", "store", False),
]


@router.get("/reports/wastage/reasons", response_model=list[WastageReasonOut])
async def get_wastage_reasons(
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List wastage reason codes for the tenant (for filters and reason breakdown). Seeds from default taxonomy if empty."""
    _ensure_tenant(user, tenant)
    result = await db.execute(
        select(WastageReason).where(WastageReason.tenant_id == tenant.id).order_by(WastageReason.category, WastageReason.code)
    )
    rows = result.scalars().all()
    if not rows:
        for code, name, category, recoverable in WASTAGE_REASON_SEED:
            r = WastageReason(tenant_id=tenant.id, code=code, name=name, category=category, recoverable=recoverable)
            db.add(r)
        await db.commit()
        result = await db.execute(
            select(WastageReason).where(WastageReason.tenant_id == tenant.id).order_by(WastageReason.category, WastageReason.code)
        )
        rows = result.scalars().all()
    return [
        WastageReasonOut(id=r.id, code=r.code, name=r.name, category=r.category, recoverable=r.recoverable)
        for r in rows
    ]


class WastageTrendSeriesItem(BaseModel):
    label: str
    value: float


class WastageTrendsOut(BaseModel):
    series: list[WastageTrendSeriesItem] | None = None
    by_buyer: list[dict] | None = None  # [{ buyer_id, buyer_name, value }]
    by_material_group: list[dict] | None = None  # [{ category, value }]


@router.get("/reports/wastage/trends", response_model=WastageTrendsOut)
async def get_wastage_trends(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    group_by: str = Query(default="month", description="month | buyer | material_group"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Trends: wastage value grouped by month, buyer, or material group (category)."""
    _ensure_tenant(user, tenant)
    rows = await get_wastage_report(
        order_id=None,
        style_id=None,
        buyer_id=None,
        date_from=date_from,
        date_to=date_to,
        threshold_pct=None,
        above_threshold_only=False,
        tenant=tenant,
        user=user,
        db=db,
    )
    if group_by == "month":
        by_month: dict[str, float] = defaultdict(float)
        for r in rows:
            d = r.order_date
            key = d.strftime("%Y-%m") if d else "unknown"
            by_month[key] += r.wastage_value
        series = [WastageTrendSeriesItem(label=k, value=round(v, 2)) for k, v in sorted(by_month.items())]
        return WastageTrendsOut(series=series)
    if group_by == "buyer":
        by_buyer: dict[int, tuple[str, float]] = {}
        for r in rows:
            if r.buyer_id not in by_buyer:
                by_buyer[r.buyer_id] = (r.buyer_name, 0.0)
            by_buyer[r.buyer_id] = (r.buyer_name, by_buyer[r.buyer_id][1] + r.wastage_value)
        by_buyer_list = [
            {"buyer_id": bid, "buyer_name": name, "value": round(v, 2)}
            for bid, (name, v) in by_buyer.items()
        ]
        return WastageTrendsOut(by_buyer=by_buyer_list)
    if group_by == "material_group":
        by_cat: dict[str, float] = defaultdict(float)
        for r in rows:
            by_cat[r.category] += r.wastage_value
        by_material = [{"category": k, "value": round(v, 2)} for k, v in sorted(by_cat.items())]
        return WastageTrendsOut(by_material_group=by_material)
    return WastageTrendsOut()


class WastageTransactionCreate(BaseModel):
    order_id: int
    item_id: int | None = None
    process_stage: str
    reason_id: int | None = None
    quantity: str = "0"
    unit_cost: str = "0"
    recoverable_value: str = "0"


@router.post("/reports/wastage/transactions")
async def create_wastage_transaction(
    body: WastageTransactionCreate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Record a wastage transaction for an order (process stage and optional reason)."""
    _ensure_tenant(user, tenant)
    order = await db.get(Order, body.order_id)
    if not order or order.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Order not found")
    qty = _to_float_safe(body.quantity)
    cost = _to_float_safe(body.unit_cost)
    rec = _to_float_safe(body.recoverable_value)
    value = max(0.0, qty * cost)
    tx = WastageTransaction(
        tenant_id=tenant.id,
        order_id=body.order_id,
        item_id=body.item_id,
        process_stage=body.process_stage,
        reason_id=body.reason_id,
        quantity=body.quantity,
        unit_cost=body.unit_cost,
        value=str(round(value, 2)),
        recoverable_value=body.recoverable_value,
        created_by_id=user.id,
    )
    db.add(tx)
    await db.commit()
    await db.refresh(tx)
    return {"id": tx.id, "order_id": tx.order_id, "value": value}


@router.get("/reports/wastage/export")
async def export_wastage_report(
    order_id: int | None = Query(default=None),
    style_id: int | None = Query(default=None),
    buyer_id: int | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    format: str = Query(default="xlsx", description="xlsx"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export wastage report as Excel (same filters as report)."""
    _ensure_tenant(user, tenant)
    rows = await get_wastage_report(
        order_id=order_id,
        style_id=style_id,
        buyer_id=buyer_id,
        date_from=date_from,
        date_to=date_to,
        threshold_pct=None,
        above_threshold_only=False,
        tenant=tenant,
        user=user,
        db=db,
    )
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export not available (openpyxl missing)")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=500, detail="Workbook error")
    ws.title = "Wastage detail"
    headers = [
        "Order ID", "Order Code", "Order Date", "Delivery Date", "Buyer", "Style", "Item Code", "Item Name",
        "Category", "Expected Qty", "Actual Qty", "Wastage %", "Wastage Value", "Threshold Breach",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.order_id, r.order_code, r.order_date.isoformat() if r.order_date else "", r.delivery_date.isoformat() if r.delivery_date else "",
            r.buyer_name, r.style_code, r.item_code, r.item_name, r.category,
            r.expected_qty, r.actual_qty, r.wastage_pct_vs_bom, r.wastage_value, "Yes" if r.threshold_breach else "No",
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="wastage_report.xlsx"'},
    )


# ---------- Phase 3: thresholds, saved views, management summary ----------


class WastageThresholdRuleOut(BaseModel):
    id: int
    scope_type: str
    scope_id: int | None
    allowed_pct: float
    critical_pct: float


class WastageThresholdRuleCreate(BaseModel):
    scope_type: str  # tenant | buyer | order_type | material_type
    scope_id: int | None = None
    allowed_pct: float
    critical_pct: float


@router.get("/reports/wastage/thresholds", response_model=list[WastageThresholdRuleOut])
async def list_wastage_thresholds(
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List wastage threshold rules for the tenant (for badge and config)."""
    _ensure_tenant(user, tenant)
    result = await db.execute(
        select(WastageThresholdRule).where(WastageThresholdRule.tenant_id == tenant.id)
    )
    rows = result.scalars().all()
    return [
        WastageThresholdRuleOut(
            id=r.id,
            scope_type=r.scope_type,
            scope_id=r.scope_id,
            allowed_pct=float(r.allowed_pct),
            critical_pct=float(r.critical_pct),
        )
        for r in rows
    ]


@router.post("/reports/wastage/thresholds", status_code=201, response_model=WastageThresholdRuleOut)
async def create_wastage_threshold(
    body: WastageThresholdRuleCreate,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create a wastage threshold rule (tenant, buyer, order_type, or material_type scope)."""
    _ensure_tenant(user, tenant)
    row = WastageThresholdRule(
        tenant_id=tenant.id,
        scope_type=body.scope_type,
        scope_id=body.scope_id,
        allowed_pct=body.allowed_pct,
        critical_pct=body.critical_pct,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return WastageThresholdRuleOut(
        id=row.id,
        scope_type=row.scope_type,
        scope_id=row.scope_id,
        allowed_pct=float(row.allowed_pct),
        critical_pct=float(row.critical_pct),
    )


class WastageSavedViewOut(BaseModel):
    id: int
    name: str
    description: str | None
    filter_json: dict
    is_default: bool
    created_at: datetime | None


class WastageSavedViewBody(BaseModel):
    name: str
    description: str | None = None
    filter_json: dict
    is_default: bool = False


@router.get("/reports/wastage/views", response_model=list[WastageSavedViewOut])
async def list_wastage_views(
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List saved filter views for wastage report (current user)."""
    _ensure_tenant(user, tenant)
    result = await db.execute(
        select(WastageSavedView).where(
            WastageSavedView.tenant_id == tenant.id,
            WastageSavedView.user_id == user.id,
        ).order_by(WastageSavedView.name.asc())
    )
    rows = result.scalars().all()
    return [
        WastageSavedViewOut(
            id=r.id,
            name=r.name,
            description=r.description,
            filter_json=r.filter_json or {},
            is_default=r.is_default,
            created_at=r.created_at,
        )
        for r in rows
    ]


@router.post("/reports/wastage/views", status_code=201, response_model=WastageSavedViewOut)
async def create_wastage_view(
    body: WastageSavedViewBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Save current filter state as a named wastage report view."""
    _ensure_tenant(user, tenant)
    if body.is_default:
        default_rows = (await db.execute(
            select(WastageSavedView).where(
                WastageSavedView.tenant_id == tenant.id,
                WastageSavedView.user_id == user.id,
                WastageSavedView.is_default == True,
            )
        )).scalars().all()
        for r in default_rows:
            r.is_default = False
    row = WastageSavedView(
        tenant_id=tenant.id,
        user_id=user.id,
        name=body.name,
        description=body.description,
        filter_json=body.filter_json,
        is_default=body.is_default,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return WastageSavedViewOut(
        id=row.id,
        name=row.name,
        description=row.description,
        filter_json=row.filter_json or {},
        is_default=row.is_default,
        created_at=row.created_at,
    )


@router.delete("/reports/wastage/views/{view_id}", status_code=204)
async def delete_wastage_view(
    view_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete a saved wastage report view."""
    _ensure_tenant(user, tenant)
    row = await db.get(WastageSavedView, view_id)
    if not row or row.tenant_id != tenant.id or row.user_id != user.id:
        raise HTTPException(status_code=404, detail="Saved view not found")
    await db.delete(row)
    await db.commit()


class WastageManagementSummaryOut(BaseModel):
    top_orders: list[dict]  # [{ order_id, order_code, buyer_name, total_wastage_value }]
    top_materials: list[dict]  # [{ item_id, item_code, item_name, total_wastage_value }]
    top_reasons: list[dict]  # [{ reason_code, reason_name, value, count }] from wastage_transaction
    mom_change: dict  # { current_total, previous_total, current_above_threshold, previous_above_threshold }
    suggested_actions: list[str]


@router.get("/reports/wastage/management-summary", response_model=WastageManagementSummaryOut)
async def get_wastage_management_summary(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Management summary: top 10 orders, top 10 materials, top reasons, month-over-month, suggested actions."""
    _ensure_tenant(user, tenant)
    rows = await get_wastage_report(
        order_id=None,
        style_id=None,
        buyer_id=None,
        date_from=date_from,
        date_to=date_to,
        threshold_pct=None,
        above_threshold_only=False,
        tenant=tenant,
        user=user,
        db=db,
    )
    # Top 10 orders by total wastage value
    order_totals: dict[int, tuple[str, str, float]] = {}
    for r in rows:
        if r.order_id not in order_totals:
            order_totals[r.order_id] = (r.order_code, r.buyer_name, 0.0)
        order_totals[r.order_id] = (r.order_code, r.buyer_name, order_totals[r.order_id][2] + r.wastage_value)
    top_orders = [
        {"order_id": oid, "order_code": code, "buyer_name": name, "total_wastage_value": round(v, 2)}
        for oid, (code, name, v) in sorted(order_totals.items(), key=lambda x: -x[1][2])[:10]
    ]
    # Top 10 materials (items) by total wastage value
    item_totals: dict[int, tuple[str, str, float]] = {}
    for r in rows:
        if r.item_id not in item_totals:
            item_totals[r.item_id] = (r.item_code, r.item_name, 0.0)
        item_totals[r.item_id] = (r.item_code, r.item_name, item_totals[r.item_id][2] + r.wastage_value)
    top_materials = [
        {"item_id": iid, "item_code": code, "item_name": name, "total_wastage_value": round(v, 2)}
        for iid, (code, name, v) in sorted(item_totals.items(), key=lambda x: -x[1][2])[:10]
    ]
    # Top reasons from wastage_transaction (aggregate by reason; left join for null reason_id)
    reason_result = await db.execute(
        select(
            WastageTransaction.reason_id,
            WastageReason.code,
            WastageReason.name,
            WastageTransaction.value,
        )
        .select_from(WastageTransaction)
        .outerjoin(WastageReason, WastageTransaction.reason_id == WastageReason.id)
        .where(WastageTransaction.tenant_id == tenant.id)
    )
    reason_agg: dict[str, tuple[str, float, int]] = {}
    for r in reason_result.scalars().all():
        code = (r.code or "") if r.code else ""
        name = (r.name or "Unknown") if r.name else "Unknown"
        val = _to_float_safe(r.value)
        key = code or f"reason_{r.reason_id or 0}"
        if key not in reason_agg:
            reason_agg[key] = (name, 0.0, 0)
        reason_agg[key] = (reason_agg[key][0], reason_agg[key][1] + val, reason_agg[key][2] + 1)
    top_reasons = [
        {"reason_code": k, "reason_name": v[0], "value": round(v[1], 2), "count": v[2]}
        for k, v in sorted(reason_agg.items(), key=lambda x: -x[1][1])[:10]
    ]
    # Month-over-month: current period vs previous period (same length)
    current_total = sum(r.wastage_value for r in rows)
    current_above = len({r.order_id for r in rows if r.threshold_breach})
    period_start = date_from
    period_end = date_to
    if period_start and period_end:
        delta = (period_end - period_start).days + 1
        prev_end = period_start - timedelta(days=1)
        prev_start = prev_end - timedelta(days=delta - 1)
        prev_rows = await get_wastage_report(
            order_id=None, style_id=None, buyer_id=None,
            date_from=prev_start, date_to=prev_end,
            threshold_pct=None, above_threshold_only=False,
            tenant=tenant, user=user, db=db,
        )
        previous_total = sum(r.wastage_value for r in prev_rows)
        previous_above = len({r.order_id for r in prev_rows if r.threshold_breach})
    else:
        previous_total = 0.0
        previous_above = 0
    mom_change = {
        "current_total": round(current_total, 2),
        "previous_total": round(previous_total, 2),
        "current_above_threshold": current_above,
        "previous_above_threshold": previous_above,
    }
    # Suggested actions
    suggested_actions: list[str] = []
    if current_above > 0:
        suggested_actions.append(f"Review {current_above} order(s) with wastage above threshold.")
    if top_orders and float(top_orders[0].get("total_wastage_value", 0)) > 0:
        suggested_actions.append(f"Highest loss order: {top_orders[0].get('order_code', '')} – consider process review.")
    if top_reasons:
        suggested_actions.append(f"Top wastage reason: {top_reasons[0].get('reason_name', '')} – target corrective action.")
    if not suggested_actions:
        suggested_actions.append("No high-wastage areas identified for the selected period.")
    return WastageManagementSummaryOut(
        top_orders=top_orders,
        top_materials=top_materials,
        top_reasons=top_reasons,
        mom_change=mom_change,
        suggested_actions=suggested_actions,
    )


@router.post("/reports/wastage/refresh-summary")
async def refresh_wastage_order_summary(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Recompute and store wastage order summaries for the tenant (optional date range). Used for management view and KPIs."""
    _ensure_tenant(user, tenant)
    rows = await get_wastage_report(
        order_id=None,
        style_id=None,
        buyer_id=None,
        date_from=date_from,
        date_to=date_to,
        threshold_pct=None,
        above_threshold_only=False,
        tenant=tenant,
        user=user,
        db=db,
    )
    order_agg: dict[int, tuple[float, float, float, float, float, bool]] = {}
    for r in rows:
        if r.order_id not in order_agg:
            order_agg[r.order_id] = (0.0, 0.0, 0.0, 0.0, False)
        exp, act, trim_val, total_val, breach = order_agg[r.order_id]
        exp += r.expected_qty
        act += r.actual_qty
        if r.category == "trim":
            trim_val += r.wastage_value
        total_val += r.wastage_value
        breach = breach or r.threshold_breach
        order_agg[r.order_id] = (exp, act, trim_val, total_val, breach)
    for oid, (exp, act, trim_val, total_val, breach) in order_agg.items():
        await db.execute(delete(WastageOrderSummary).where(
            WastageOrderSummary.tenant_id == tenant.id,
            WastageOrderSummary.order_id == oid,
        ))
        var_pct = (act - exp) / exp * 100.0 if exp > 0 else 0.0
        summary = WastageOrderSummary(
            tenant_id=tenant.id,
            order_id=oid,
            period_start=date_from,
            period_end=date_to,
            planned_fabric_cons=str(round(exp, 4)),
            actual_fabric_cons=str(round(act, 4)),
            fabric_variance_pct=str(round(var_pct, 2)),
            trim_wastage_value=str(round(trim_val, 2)),
            total_wastage_value=str(round(total_val, 2)),
            above_threshold=breach,
        )
        db.add(summary)
    await db.commit()
    return {"updated_orders": len(order_agg)}


class WastageOrderDetailBomLine(BaseModel):
    item_id: int
    item_code: str
    item_name: str
    category: str
    base_consumption: float
    wastage_pct: float
    expected_qty: float
    actual_qty: float
    variance_qty: float
    wastage_pct_vs_bom: float
    wastage_value: float
    threshold_breach: bool


class WastageReasonBreakdownItem(BaseModel):
    reason_id: int | None
    reason_code: str
    reason_name: str
    value: float
    quantity: float


class WastageProcessStageBreakdownItem(BaseModel):
    process_stage: str
    value: float
    quantity: float


class WastageOrderDetailOut(BaseModel):
    order_id: int
    order_code: str
    order_date: date | None
    delivery_date: date | None
    buyer_id: int
    buyer_name: str
    style_id: int
    style_code: str
    quantity: int | None
    bom_lines: list[WastageOrderDetailBomLine]
    total_expected_value: float
    total_actual_value: float
    total_wastage_value: float
    linked_alert_ids: list[int]
    reason_breakdown: list[WastageReasonBreakdownItem] = []
    process_stage_breakdown: list[WastageProcessStageBreakdownItem] = []


@router.get("/reports/wastage/order/{order_id}", response_model=WastageOrderDetailOut)
async def get_wastage_order_detail(
    order_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Detail for one order: BOM lines with planned/actual/variance and linked alert IDs."""
    _ensure_tenant(user, tenant)
    order = await db.get(Order, order_id)
    if not order or order.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Order not found")
    customer = await db.get(Customer, order.customer_id)
    buyer_name = customer.name if customer else f"Customer #{order.customer_id}"
    if not order.quotation_id:
        raise HTTPException(status_code=404, detail="Order has no quotation")
    quotation = await db.get(Quotation, order.quotation_id)
    if not quotation or quotation.tenant_id != tenant.id or not quotation.style_id:
        raise HTTPException(status_code=404, detail="Quotation or style not found")
    sid = quotation.style_id
    style = await db.get(GarmentStyle, sid)
    style_code = style.style_code if style else str(sid)
    order_qty = _to_float_safe(str(order.quantity)) if order.quantity is not None else 0.0
    bom = await _get_latest_governed_bom(
        db,
        tenant_id=tenant.id,
        style_id=sid,
    )
    if not bom:
        raise HTTPException(status_code=400, detail="No APPROVED/FROZEN BOM for style")
    lines_result = await db.execute(
        select(BomItem).where(
            BomItem.tenant_id == tenant.id,
            BomItem.bom_id == bom.id,
            BomItem.item_id.isnot(None),
        )
    )
    total_expected_value = 0.0
    total_actual_value = 0.0
    bom_lines: list[WastageOrderDetailBomLine] = []
    for line in lines_result.scalars().all():
        item = await db.get(Item, line.item_id)
        if not item or item.tenant_id != tenant.id:
            continue
        cat = await db.get(ItemCategory, item.category_id) if item.category_id else None
        category = _wastage_category_from_item(item, cat)
        base = _to_float_safe(line.base_consumption)
        wastage_pct_val = _to_float_safe(line.wastage_pct)
        expected = order_qty * base * (1.0 + wastage_pct_val / 100.0)
        act_result = await db.execute(
            select(StockMovement).where(
                StockMovement.tenant_id == tenant.id,
                StockMovement.reference_type == "CONSUMPTION_ISSUE",
                StockMovement.reference_id == order.id,
                StockMovement.item_id == line.item_id,
            )
        )
        actual = sum(
            _to_float_safe(m.quantity) for m in act_result.scalars().all() if (m.movement_type or "").upper() == "OUT"
        )
        variance_qty = actual - expected
        wastage_pct_vs_bom = round((variance_qty / expected * 100.0), 2) if expected > 0 else 0.0
        unit_cost = _to_float_safe(item.default_cost)
        wastage_value = round(max(0.0, variance_qty) * unit_cost, 2)
        total_expected_value += expected * unit_cost
        total_actual_value += actual * unit_cost
        bom_lines.append(
            WastageOrderDetailBomLine(
                item_id=line.item_id,
                item_code=item.item_code,
                item_name=item.name or item.description or "",
                category=category,
                base_consumption=base,
                wastage_pct=wastage_pct_val,
                expected_qty=round(expected, 4),
                actual_qty=round(actual, 4),
                variance_qty=round(variance_qty, 4),
                wastage_pct_vs_bom=wastage_pct_vs_bom,
                wastage_value=wastage_value,
                threshold_breach=wastage_pct_vs_bom > DEFAULT_WASTAGE_THRESHOLD_PCT,
            )
        )
    total_wastage_value = round(total_actual_value - total_expected_value, 2)
    if total_wastage_value < 0:
        total_wastage_value = 0.0
    alert_ids_result = await db.execute(
        select(AlertRelatedEntity.alert_id).where(
            AlertRelatedEntity.tenant_id == tenant.id,
            AlertRelatedEntity.entity_type == "order",
            AlertRelatedEntity.entity_id == order_id,
        )
    )
    linked_alert_ids = list({r[0] for r in alert_ids_result.scalars().all()})

    # Reason and process-stage breakdown from wastage_transaction
    tx_result = await db.execute(
        select(WastageTransaction, WastageReason).outerjoin(
            WastageReason, WastageTransaction.reason_id == WastageReason.id
        ).where(
            WastageTransaction.tenant_id == tenant.id,
            WastageTransaction.order_id == order_id,
        )
    )
    tx_rows = tx_result.all()
    reason_agg: dict[int | None, tuple[str, str, float, float]] = {}
    stage_agg: dict[str, tuple[float, float]] = {}
    for tx, reason in tx_rows:
        val = _to_float_safe(tx.value)
        qty = _to_float_safe(tx.quantity)
        code = reason.code if reason else ""
        name = reason.name if reason else "Unknown"
        rid = tx.reason_id
        if rid not in reason_agg:
            reason_agg[rid] = (code, name, 0.0, 0.0)
        reason_agg[rid] = (code, name, reason_agg[rid][2] + val, reason_agg[rid][3] + qty)
        stage = tx.process_stage or "unknown"
        if stage not in stage_agg:
            stage_agg[stage] = (0.0, 0.0)
        stage_agg[stage] = (stage_agg[stage][0] + val, stage_agg[stage][1] + qty)
    reason_breakdown = [
        WastageReasonBreakdownItem(reason_id=rid, reason_code=c, reason_name=n, value=round(v, 2), quantity=round(q, 4))
        for rid, (c, n, v, q) in reason_agg.items()
    ]
    process_stage_breakdown = [
        WastageProcessStageBreakdownItem(process_stage=st, value=round(v, 2), quantity=round(q, 4))
        for st, (v, q) in stage_agg.items()
    ]

    return WastageOrderDetailOut(
        order_id=order.id,
        order_code=order.order_code,
        order_date=order.order_date,
        delivery_date=order.delivery_date,
        buyer_id=order.customer_id,
        buyer_name=buyer_name,
        style_id=sid,
        style_code=style_code,
        quantity=order.quantity,
        bom_lines=bom_lines,
        total_expected_value=round(total_expected_value, 2),
        total_actual_value=round(total_actual_value, 2),
        total_wastage_value=total_wastage_value,
        linked_alert_ids=linked_alert_ids,
        reason_breakdown=reason_breakdown,
        process_stage_breakdown=process_stage_breakdown,
    )


# ---------- Advanced Critical Alerts (Phase 1: persisted alert engine) ----------

class AlertStatusUpdateBody(BaseModel):
    status: str = Field(
        ...,
        description="acknowledged | in_progress | waiting_on_buyer | waiting_on_supplier | resolved | closed | escalated | dismissed",
    )

    @field_validator("status")
    @classmethod
    def _status_allowed(cls, v: str) -> str:
        allowed = {
            "acknowledged",
            "in_progress",
            "waiting_on_buyer",
            "waiting_on_supplier",
            "resolved",
            "closed",
            "escalated",
            "dismissed",
            "open",
            "snoozed",
        }
        s = (v or "").strip().lower()
        if s not in allowed:
            raise ValueError(f"Invalid alert status. Allowed: {', '.join(sorted(allowed))}")
        return s


class AlertSnoozeBody(BaseModel):
    snoozed_until: datetime


class AlertAssignBody(BaseModel):
    assigned_to_id: int | None


class AlertMutationOut(BaseModel):
    id: int
    status: str
    assigned_to_id: int | None
    acknowledged_at: datetime | None
    acknowledged_by_id: int | None
    resolved_at: datetime | None
    resolved_by_id: int | None
    snoozed_until: datetime | None
    escalated_at: datetime | None
    escalation_level: int | None
    updated_at: datetime | None


def _alert_mutation_out(row: AlertInstance) -> AlertMutationOut:
    return AlertMutationOut(
        id=row.id,
        status=row.status,
        assigned_to_id=row.assigned_to_id,
        acknowledged_at=row.acknowledged_at,
        acknowledged_by_id=row.acknowledged_by_id,
        resolved_at=row.resolved_at,
        resolved_by_id=row.resolved_by_id,
        snoozed_until=row.snoozed_until,
        escalated_at=row.escalated_at,
        escalation_level=row.escalation_level,
        updated_at=row.updated_at,
    )


def _alert_priority_score(alert: AlertInstance, now: datetime) -> int:
    severity_weight = {
        "critical": 100,
        "high": 70,
        "medium": 40,
        "low": 20,
        "informational": 10,
    }.get((alert.severity or "").lower(), 10)
    age_hours = 0
    if alert.created_at:
        age_hours = max(0, int((now - alert.created_at).total_seconds() // 3600))
    escalation_weight = int(alert.escalation_level or 0) * 15
    return severity_weight + min(age_hours, 240) + escalation_weight


def _alert_sla_bucket(alert: AlertInstance, now: datetime) -> str:
    if (alert.status or "").lower() in {"resolved", "closed"}:
        return "met"
    age_hours = 0
    if alert.created_at:
        age_hours = max(0, int((now - alert.created_at).total_seconds() // 3600))
    sev = (alert.severity or "").lower()
    breach_hours = {"critical": 24, "high": 48, "medium": 72, "low": 120, "informational": 168}.get(sev, 72)
    return "breach" if age_hours > breach_hours else "at_risk"


class AlertDefinitionOut(BaseModel):
    id: int
    rule_key: str
    name: str
    description: str | None
    severity_default: str
    entity_type: str
    is_system: bool
    is_enabled: bool
    config_json: dict | None = None


class AlertDefinitionPatch(BaseModel):
    is_enabled: bool | None = None
    config_json: dict | None = None


@router.get("/alert-definitions", response_model=list[AlertDefinitionOut])
async def list_alert_definitions(
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Tenant-scoped alert rule definitions (enable/disable and JSON config per rule)."""
    from app.modules.merch.alert_engine import ensure_definitions_for_tenant as _seed_defs

    _ensure_tenant(user, tenant)
    await _seed_defs(db, tenant.id)
    result = await db.execute(
        select(AlertDefinition).where(AlertDefinition.tenant_id == tenant.id).order_by(AlertDefinition.rule_key)
    )
    rows = result.scalars().all()
    return [
        AlertDefinitionOut(
            id=r.id,
            rule_key=r.rule_key,
            name=r.name,
            description=r.description,
            severity_default=r.severity_default,
            entity_type=r.entity_type,
            is_system=r.is_system,
            is_enabled=r.is_enabled,
            config_json=r.config_json if isinstance(r.config_json, dict) else None,
        )
        for r in rows
    ]


@router.patch("/alert-definitions/{definition_id}", response_model=AlertDefinitionOut)
async def patch_alert_definition(
    definition_id: int,
    body: AlertDefinitionPatch,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_merch_permission(MERCH_PERMISSION_ALERT_DEFINITIONS)),
):
    _ensure_tenant(user, tenant)
    row = await db.get(AlertDefinition, definition_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Alert definition not found")
    if body.is_enabled is not None:
        row.is_enabled = body.is_enabled
    if body.config_json is not None:
        row.config_json = body.config_json
    await db.flush()
    await db.refresh(row)
    return AlertDefinitionOut(
        id=row.id,
        rule_key=row.rule_key,
        name=row.name,
        description=row.description,
        severity_default=row.severity_default,
        entity_type=row.entity_type,
        is_system=row.is_system,
        is_enabled=row.is_enabled,
        config_json=row.config_json if isinstance(row.config_json, dict) else None,
    )


@router.get("/alerts")
async def list_alerts(
    severity: str | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    alert_type: str | None = Query(default=None, alias="alert_type"),
    entity_type: str | None = Query(default=None, alias="entity_type"),
    entity_id: int | None = Query(default=None, alias="entity_id"),
    order_id: int | None = Query(default=None),
    assigned_to_id: int | None = Query(default=None),
    min_priority_score: int | None = Query(default=None, ge=0),
    sla_bucket: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=MAX_PAGE_SIZE, description="Max rows per page (Finding #3)"),
    sort: str = Query(default="-created_at"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List persisted alerts with filters and pagination."""
    _ensure_tenant(user, tenant)
    stmt = select(AlertInstance).where(AlertInstance.tenant_id == tenant.id)
    if entity_type:
        stmt = stmt.join(AlertDefinition, AlertInstance.definition_id == AlertDefinition.id).where(
            AlertDefinition.tenant_id == tenant.id,
            AlertDefinition.entity_type == entity_type.strip().lower(),
        )
    if entity_id is not None:
        rel_sub = select(AlertRelatedEntity.alert_id).where(
            AlertRelatedEntity.tenant_id == tenant.id,
            AlertRelatedEntity.entity_id == entity_id,
        )
        if entity_type:
            rel_sub = rel_sub.where(AlertRelatedEntity.entity_type == entity_type.strip().lower())
        stmt = stmt.where(AlertInstance.id.in_(rel_sub))
    if severity:
        stmt = stmt.where(AlertInstance.severity == severity.lower())
    if status_filter:
        stmt = stmt.where(AlertInstance.status == status_filter.lower())
    if alert_type:
        stmt = stmt.where(AlertInstance.alert_type == alert_type)
    if assigned_to_id is not None:
        stmt = stmt.where(AlertInstance.assigned_to_id == assigned_to_id)
    if order_id is not None:
        sub = select(AlertRelatedEntity.alert_id).where(
            AlertRelatedEntity.tenant_id == tenant.id,
            AlertRelatedEntity.entity_type == "order",
            AlertRelatedEntity.entity_id == order_id,
        )
        stmt = stmt.where(AlertInstance.id.in_(sub))
    # Exclude snoozed that are still in future (use timezone-aware now to match TIMESTAMPTZ created_at)
    now = datetime.now(timezone.utc)
    stmt = stmt.where(
        or_(
            AlertInstance.snoozed_until.is_(None),
            AlertInstance.snoozed_until <= now,
        )
    )
    needs_python_pagination = (
        min_priority_score is not None
        or (sla_bucket is not None and sla_bucket.strip() != "")
        or sort.lstrip("-") == "priority_score"
    )
    total_result = await db.execute(select(func.count()).select_from(stmt.subquery()))
    total = total_result.scalar() or 0
    sort_col = AlertInstance.created_at
    if sort.lstrip("-") == "created_at":
        sort_col = AlertInstance.created_at
    elif sort.lstrip("-") == "updated_at":
        sort_col = AlertInstance.updated_at
    elif sort.lstrip("-") == "severity":
        sort_col = AlertInstance.severity
    if sort.startswith("-"):
        stmt = stmt.order_by(sort_col.desc())
    else:
        stmt = stmt.order_by(sort_col.asc())
    if not needs_python_pagination:
        stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(stmt)
    rows = result.scalars().all()
    # Enrich with order_id / order_code from related_entity
    items = []
    normalized_sla_bucket = (sla_bucket or "").strip().lower() or None
    for r in rows:
        primary_rel_result = await db.execute(
            select(AlertRelatedEntity).where(
                AlertRelatedEntity.alert_id == r.id,
                AlertRelatedEntity.role == "primary",
            ).limit(1)
        )
        primary_rel = primary_rel_result.scalar_one_or_none()
        order_code = None
        link_order_id = None
        link_entity_type = primary_rel.entity_type if primary_rel else None
        link_entity_id = primary_rel.entity_id if primary_rel else None
        if primary_rel and primary_rel.entity_type == "order":
            link_order_id = primary_rel.entity_id
            order_row = await db.get(Order, primary_rel.entity_id)
            if order_row and order_row.tenant_id == tenant.id:
                order_code = order_row.order_code
        priority_score = _alert_priority_score(r, now)
        item_sla_bucket = _alert_sla_bucket(r, now)
        item = {
            "id": r.id,
            "natural_key": r.natural_key,
            "title": r.title,
            "description": r.description,
            "severity": r.severity,
            "status": r.status,
            "alert_type": r.alert_type,
            "assigned_to_id": r.assigned_to_id,
            "entity_type": link_entity_type,
            "entity_id": link_entity_id,
            "order_id": link_order_id,
            "order_code": order_code,
            "reason_text": r.reason_text,
            "recommended_action": r.recommended_action,
            "priority_score": priority_score,
            "sla_bucket": item_sla_bucket,
            "snoozed_until": r.snoozed_until.isoformat() if r.snoozed_until else None,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        }
        if min_priority_score is not None and priority_score < min_priority_score:
            continue
        if normalized_sla_bucket and normalized_sla_bucket != item_sla_bucket:
            continue
        items.append(item)
    if sort.lstrip("-") == "priority_score":
        items.sort(key=lambda x: x["priority_score"], reverse=sort.startswith("-"))
    if needs_python_pagination:
        total = len(items)
        start = (page - 1) * page_size
        end = start + page_size
        items = items[start:end]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/alerts/summary")
async def get_alerts_summary(
    severity: str | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    alert_type: str | None = Query(default=None, alias="alert_type"),
    entity_type: str | None = Query(default=None, alias="entity_type"),
    entity_id: int | None = Query(default=None, alias="entity_id"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """KPI counts for alert center (critical, high, medium, low, total)."""
    _ensure_tenant(user, tenant)
    now = datetime.now(timezone.utc)
    stmt = select(AlertInstance.severity, func.count()).where(
        AlertInstance.tenant_id == tenant.id,
        or_(
            AlertInstance.snoozed_until.is_(None),
            AlertInstance.snoozed_until <= now,
        ),
    )
    if severity:
        stmt = stmt.where(AlertInstance.severity == severity)
    if status_filter:
        stmt = stmt.where(AlertInstance.status == status_filter)
    if alert_type:
        stmt = stmt.where(AlertInstance.alert_type == alert_type.strip())
    if entity_type:
        stmt = stmt.join(AlertDefinition, AlertInstance.definition_id == AlertDefinition.id).where(
            AlertDefinition.tenant_id == tenant.id,
            AlertDefinition.entity_type == entity_type.strip().lower(),
        )
    if entity_id is not None:
        rel_sub = select(AlertRelatedEntity.alert_id).where(
            AlertRelatedEntity.tenant_id == tenant.id,
            AlertRelatedEntity.entity_id == entity_id,
        )
        if entity_type:
            rel_sub = rel_sub.where(AlertRelatedEntity.entity_type == entity_type.strip().lower())
        stmt = stmt.where(AlertInstance.id.in_(rel_sub))
    stmt = stmt.group_by(AlertInstance.severity)
    by_sev = await db.execute(stmt)
    counts = {row[0]: row[1] for row in by_sev.all()}
    total = sum(counts.values())
    return {
        "by_severity": {
            "critical": counts.get("critical", 0),
            "high": counts.get("high", 0),
            "medium": counts.get("medium", 0),
            "low": counts.get("low", 0),
            "informational": counts.get("informational", 0),
        },
        "total": total,
    }


# Static paths must be registered before /alerts/{alert_id} or "views" is parsed as alert_id (422).
@router.get("/alerts/views")
async def list_alert_views(
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List saved filter views for current user."""
    _ensure_tenant(user, tenant)
    result = await db.execute(
        select(AlertSavedView)
        .where(
            AlertSavedView.tenant_id == tenant.id,
            AlertSavedView.user_id == user.id,
        )
        .order_by(AlertSavedView.name.asc())
    )
    rows = result.scalars().all()
    return [
        {
            "id": r.id,
            "name": r.name,
            "description": r.description,
            "filter_json": r.filter_json,
            "is_default": r.is_default,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in rows
    ]


@router.get("/alerts/{alert_id}")
async def get_alert_detail(
    alert_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Single alert for detail drawer."""
    _ensure_tenant(user, tenant)
    row = await db.get(AlertInstance, alert_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Alert not found")
    rels = await db.execute(
        select(AlertRelatedEntity).where(
            AlertRelatedEntity.alert_id == alert_id,
            AlertRelatedEntity.tenant_id == tenant.id,
        )
    )
    related = rels.scalars().all()
    order_id = None
    order_code = None
    for rel in related:
        if rel.entity_type == "order":
            order_id = rel.entity_id
            o = await db.get(Order, rel.entity_id)
            if o and o.tenant_id == tenant.id:
                order_code = o.order_code
            break
    now = datetime.now(timezone.utc)
    return {
        "id": row.id,
        "natural_key": row.natural_key,
        "title": row.title,
        "description": row.description,
        "severity": row.severity,
        "status": row.status,
        "alert_type": row.alert_type,
        "assigned_to_id": row.assigned_to_id,
        "order_id": order_id,
        "order_code": order_code,
        "reason_text": row.reason_text,
        "recommended_action": row.recommended_action,
        "priority_score": _alert_priority_score(row, now),
        "sla_bucket": _alert_sla_bucket(row, now),
        "snoozed_until": row.snoozed_until.isoformat() if row.snoozed_until else None,
        "escalated_at": row.escalated_at.isoformat() if row.escalated_at else None,
        "escalation_level": row.escalation_level,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        "resolved_at": row.resolved_at.isoformat() if row.resolved_at else None,
    }


@router.patch("/alerts/{alert_id}/status", response_model=AlertMutationOut)
async def update_alert_status(
    alert_id: int,
    body: AlertStatusUpdateBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update alert status (acknowledged, in_progress, resolved, etc.)."""
    _ensure_tenant(user, tenant)
    row = await db.get(AlertInstance, alert_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Alert not found")
    old_status = row.status
    row.status = body.status
    if body.status in ("resolved", "closed"):
        row.resolved_at = datetime.now(timezone.utc)
        row.resolved_by_id = user.id
    elif body.status == "acknowledged":
        row.acknowledged_at = datetime.now(timezone.utc)
        row.acknowledged_by_id = user.id
    hist = AlertHistory(
        tenant_id=tenant.id,
        alert_id=alert_id,
        user_id=user.id,
        action="status_change",
        field_name="status",
        old_value=old_status,
        new_value=row.status,
    )
    db.add(hist)
    await db.flush()
    await db.refresh(row)
    return _alert_mutation_out(row)


@router.post("/alerts/{alert_id}/snooze", response_model=AlertMutationOut)
async def snooze_alert(
    alert_id: int,
    body: AlertSnoozeBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Snooze alert until given datetime."""
    _ensure_tenant(user, tenant)
    row = await db.get(AlertInstance, alert_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Alert not found")
    row.status = "snoozed"
    row.snoozed_until = body.snoozed_until
    hist = AlertHistory(
        tenant_id=tenant.id,
        alert_id=alert_id,
        user_id=user.id,
        action="snoozed",
        field_name="snoozed_until",
        new_value=body.snoozed_until.isoformat() if body.snoozed_until else None,
    )
    db.add(hist)
    await db.flush()
    await db.refresh(row)
    return _alert_mutation_out(row)


@router.post("/alerts/{alert_id}/assign", response_model=AlertMutationOut)
async def assign_alert(
    alert_id: int,
    body: AlertAssignBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_merch_permission(MERCH_PERMISSION_ALERT_ASSIGN)),
):
    """Assign alert to a user."""
    _ensure_tenant(user, tenant)
    row = await db.get(AlertInstance, alert_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Alert not found")
    if body.assigned_to_id is not None:
        assignee = await db.get(User, body.assigned_to_id)
        if not assignee or assignee.tenant_id != tenant.id:
            raise HTTPException(status_code=404, detail="Assignee user not found in this tenant")
    old_val = str(row.assigned_to_id) if row.assigned_to_id else None
    row.assigned_to_id = body.assigned_to_id
    hist = AlertHistory(
        tenant_id=tenant.id,
        alert_id=alert_id,
        user_id=user.id,
        action="assigned",
        field_name="assigned_to_id",
        old_value=old_val,
        new_value=str(body.assigned_to_id) if body.assigned_to_id else None,
    )
    db.add(hist)
    await db.flush()
    await db.refresh(row)
    return _alert_mutation_out(row)


async def _run_scan_background(tenant_id: int) -> None:
    """Background task: run alert scan in its own DB session (avoids request timeout)."""
    from app.database import AsyncSessionLocal
    from app.modules.merch.alert_engine import run_scan
    async with AsyncSessionLocal() as db:
        try:
            await run_scan(db, tenant_id, trigger="manual")
            await db.commit()
        except Exception:
            await db.rollback()
            raise


@router.post("/alerts/scan")
async def run_alerts_scan(
    background_tasks: BackgroundTasks,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _: None = Depends(require_merch_permission(MERCH_PERMISSION_ALERT_SCAN)),
):
    """Start alert rule scan for current tenant (runs in background; returns immediately)."""
    _ensure_tenant(user, tenant)
    background_tasks.add_task(_run_scan_background, tenant.id)
    return JSONResponse(
        status_code=202,
        content={"status": "accepted", "message": "Scan started in background. List will update shortly."},
    )


class AlertCommentBody(BaseModel):
    body: str
    is_internal: bool = False


class AlertEscalateBody(BaseModel):
    to_level: int = 1
    assigned_to_id: int | None = None
    reason: str | None = None


class AlertSavedViewBody(BaseModel):
    name: str
    description: str | None = None
    filter_json: dict
    is_default: bool = False


@router.get("/alerts/{alert_id}/comments")
async def list_alert_comments(
    alert_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List comments for an alert (lazy-loaded in drawer)."""
    _ensure_tenant(user, tenant)
    row = await db.get(AlertInstance, alert_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Alert not found")
    result = await db.execute(
        select(AlertComment)
        .where(AlertComment.alert_id == alert_id, AlertComment.tenant_id == tenant.id)
        .order_by(AlertComment.created_at.asc())
    )
    rows = result.scalars().all()
    return [
        {
            "id": r.id,
            "user_id": r.user_id,
            "body": r.body,
            "is_internal": r.is_internal,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in rows
    ]


@router.post("/alerts/{alert_id}/comments", status_code=201)
async def add_alert_comment(
    alert_id: int,
    body: AlertCommentBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Add a comment to an alert."""
    _ensure_tenant(user, tenant)
    row = await db.get(AlertInstance, alert_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Alert not found")
    comment = AlertComment(
        tenant_id=tenant.id,
        alert_id=alert_id,
        user_id=user.id,
        body=body.body,
        is_internal=body.is_internal,
    )
    db.add(comment)
    await db.flush()
    hist = AlertHistory(
        tenant_id=tenant.id,
        alert_id=alert_id,
        user_id=user.id,
        action="comment",
        new_value=str(comment.id),
    )
    db.add(hist)
    await db.flush()
    await db.refresh(comment)
    return {
        "id": comment.id,
        "user_id": comment.user_id,
        "body": comment.body,
        "is_internal": comment.is_internal,
        "created_at": comment.created_at.isoformat() if comment.created_at else None,
    }


@router.get("/alerts/{alert_id}/history")
async def list_alert_history(
    alert_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Timeline/history for an alert (lazy-loaded in drawer)."""
    _ensure_tenant(user, tenant)
    row = await db.get(AlertInstance, alert_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Alert not found")
    result = await db.execute(
        select(AlertHistory)
        .where(AlertHistory.alert_id == alert_id, AlertHistory.tenant_id == tenant.id)
        .order_by(AlertHistory.created_at.desc())
    )
    rows = result.scalars().all()
    return [
        {
            "id": h.id,
            "user_id": h.user_id,
            "action": h.action,
            "field_name": h.field_name,
            "old_value": h.old_value,
            "new_value": h.new_value,
            "created_at": h.created_at.isoformat() if h.created_at else None,
        }
        for h in rows
    ]


@router.post("/alerts/{alert_id}/escalate", response_model=AlertMutationOut)
async def escalate_alert(
    alert_id: int,
    body: AlertEscalateBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Escalate alert to a level and optionally assign."""
    _ensure_tenant(user, tenant)
    row = await db.get(AlertInstance, alert_id)
    if not row or row.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Alert not found")
    if body.assigned_to_id is not None:
        assignee = await db.get(User, body.assigned_to_id)
        if not assignee or assignee.tenant_id != tenant.id:
            raise HTTPException(status_code=404, detail="Assignee user not found in this tenant")
    from_level = row.escalation_level
    row.status = "escalated"
    row.escalated_at = datetime.now(timezone.utc)
    row.escalation_level = body.to_level
    if body.assigned_to_id is not None:
        row.assigned_to_id = body.assigned_to_id
    log = AlertEscalationLog(
        tenant_id=tenant.id,
        alert_id=alert_id,
        from_level=from_level,
        to_level=body.to_level,
        assigned_to_id=body.assigned_to_id,
        reason=body.reason,
        created_by_id=user.id,
    )
    db.add(log)
    hist = AlertHistory(
        tenant_id=tenant.id,
        alert_id=alert_id,
        user_id=user.id,
        action="escalated",
        field_name="escalation_level",
        old_value=str(from_level) if from_level is not None else None,
        new_value=str(body.to_level),
    )
    db.add(hist)
    await db.flush()
    await db.refresh(row)
    return _alert_mutation_out(row)


@router.post("/alerts/views", status_code=201)
async def create_alert_view(
    body: AlertSavedViewBody,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Save current filter state as a named view."""
    _ensure_tenant(user, tenant)
    if body.is_default:
        default_rows = (await db.execute(
            select(AlertSavedView).where(
                AlertSavedView.tenant_id == tenant.id,
                AlertSavedView.user_id == user.id,
                AlertSavedView.is_default == True,
            )
        )).scalars().all()
        for r in default_rows:
            r.is_default = False
    row = AlertSavedView(
        tenant_id=tenant.id,
        user_id=user.id,
        name=body.name,
        description=body.description,
        filter_json=body.filter_json,
        is_default=body.is_default,
    )
    db.add(row)
    await db.flush()
    await db.refresh(row)
    return {
        "id": row.id,
        "name": row.name,
        "description": row.description,
        "filter_json": row.filter_json,
        "is_default": row.is_default,
        "created_at": row.created_at.isoformat() if row.created_at else None,
    }


@router.delete("/alerts/views/{view_id}", status_code=204)
async def delete_alert_view(
    view_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete a saved view."""
    _ensure_tenant(user, tenant)
    row = await db.get(AlertSavedView, view_id)
    if not row or row.tenant_id != tenant.id or row.user_id != user.id:
        raise HTTPException(status_code=404, detail="Saved view not found")
    await db.delete(row)
    await db.flush()


@router.get("/critical-alerts")
async def get_critical_alerts(
    wastage_threshold_pct: float | None = Query(default=15.0, description="Include wastage alerts above this %"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _ensure_tenant(user, tenant)
    overdue = await db.execute(
        select(Followup).where(
            and_(
                Followup.tenant_id == tenant.id,
                Followup.status != "DONE",
                Followup.due_date.is_not(None),
                Followup.due_date < date.today(),
            )
        )
    )
    rows = overdue.scalars().all()
    alerts: list[dict] = [
        {
            "id": f"followup-{r.id}",
            "severity": "critical" if (date.today() - r.due_date).days > 7 else "warning",
            "category": "Order Follow-up",
            "title": r.title,
            "description": f"Order #{r.order_id} overdue by {(date.today() - r.due_date).days} day(s)",
            "order_id": r.order_id,
        }
        for r in rows
        if r.due_date is not None
    ]
    # Phase E: add wastage alerts (actual vs BOM above threshold)
    wastage_rows = await get_wastage_report(
        order_id=None,
        style_id=None,
        date_from=None,
        date_to=None,
        threshold_pct=wastage_threshold_pct,
        tenant=tenant,
        user=user,
        db=db,
    )
    for r in wastage_rows:
        alerts.append({
            "id": f"wastage-{r.order_id}-{r.item_id}",
            "severity": "warning",
            "category": "High Wastage",
            "title": f"Order {r.order_code} · {r.item_code}",
            "description": f"Wastage vs BOM: {r.wastage_pct_vs_bom:+.1f}% (expected {r.expected_qty}, actual {r.actual_qty})",
            "order_id": r.order_id,
            "style_id": r.style_id,
            "item_id": r.item_id,
        })
    return {
        "summary": {
            "critical": len([a for a in alerts if a["severity"] == "critical"]),
            "warning": len([a for a in alerts if a["severity"] == "warning"]),
            "total": len(alerts),
        },
        "alerts": alerts,
    }


# ---------- Consumption Reconciliation (BOM-based planned vs StockMovement actuals) ----------


class ConsumptionReconOrderOut(BaseModel):
    id: int
    order_code: str
    style_code: str
    quantity: int | None


class ConsumptionReconItemOut(BaseModel):
    item_id: int
    item_code: str
    item_name: str
    material_type: str
    uom: str | None
    planned_qty: float
    actual_qty: float
    variance: float
    variance_pct: float
    unit_cost: float | None = None
    planned_cost: float | None = None
    actual_cost: float | None = None
    cost_variance: float | None = None
    last_issued_at: str | None = None
    movement_count: int = 0
    # Three-layer (quoted vs BOM vs actual); optional for backward compatibility
    quoted_consumption_per_unit: float | None = None
    bom_net_consumption_per_unit: float | None = None
    bom_gross_consumption_per_unit: float | None = None
    wastage_pct: float | None = None
    process_loss_pct: float | None = None
    quoted_planned_qty: float | None = None
    quoted_planned_cost: float | None = None
    bom_planned_cost: float | None = None
    quoted_vs_bom_variance_pct: float | None = None
    bom_vs_actual_variance_pct: float | None = None
    quoted_vs_actual_variance_pct: float | None = None
    planned_wastage_qty: float | None = None
    planned_process_loss_qty: float | None = None
    planned_loss_vs_actual_loss: float | None = None
    cost_impact_quoted_vs_bom: float | None = None
    cost_impact_bom_vs_actual: float | None = None
    cost_impact_quoted_vs_actual: float | None = None


class ConsumptionReconSummaryOut(BaseModel):
    total_planned: float
    total_actual: float
    variance: float
    overall_variance_pct: float
    items_exceeding_tolerance: int
    total_planned_cost: float = 0.0
    total_actual_cost: float = 0.0
    cost_variance: float = 0.0
    cost_variance_pct: float = 0.0
    total_quoted_planned_qty: float = 0.0
    total_quoted_planned_cost: float = 0.0
    total_bom_planned_cost: float = 0.0
    quoted_vs_bom_cost_variance: float = 0.0
    quoted_vs_actual_cost_variance: float = 0.0


class ConsumptionReconResponse(BaseModel):
    order: ConsumptionReconOrderOut
    items: list[ConsumptionReconItemOut]
    summary: ConsumptionReconSummaryOut
    bom_version: str | None = None
    bom_status: str | None = None
    order_status: str | None = None
    consumption_plan_status: str | None = None


class ConsumptionReconMovementOut(BaseModel):
    movement_id: int
    movement_date: str | None
    quantity: float
    warehouse_name: str | None
    issued_by: str | None
    reference_code: str | None = None
    notes: str | None


class ConsumptionReconMovementsResponse(BaseModel):
    item_id: int
    item_code: str
    item_name: str
    planned_qty: float
    total_issued: float
    movements: list[ConsumptionReconMovementOut]


def _empty_recon_summary() -> ConsumptionReconSummaryOut:
    return ConsumptionReconSummaryOut(
        total_planned=0.0,
        total_actual=0.0,
        variance=0.0,
        overall_variance_pct=0.0,
        items_exceeding_tolerance=0,
        total_planned_cost=0.0,
        total_actual_cost=0.0,
        cost_variance=0.0,
        cost_variance_pct=0.0,
        total_quoted_planned_qty=0.0,
        total_quoted_planned_cost=0.0,
        total_bom_planned_cost=0.0,
        quoted_vs_bom_cost_variance=0.0,
        quoted_vs_actual_cost_variance=0.0,
    )


async def _get_active_order_governed_bom(
    db: AsyncSession,
    *,
    tenant_id: int,
    order_id: int,
) -> Bom | None:
    r = await db.execute(
        select(Bom).where(
            Bom.tenant_id == tenant_id,
            Bom.order_id == order_id,
            Bom.is_active.is_(True),
            Bom.status.in_(GOVERNED_BOM_STATUSES),
        )
    )
    return r.scalars().first()


async def _consumption_plan_status_for_order(
    db: AsyncSession, tenant_id: int, order_id: int
) -> str | None:
    r = await db.execute(
        select(ConsumptionPlan.status)
        .where(
            ConsumptionPlan.tenant_id == tenant_id,
            ConsumptionPlan.order_id == order_id,
        )
        .order_by(ConsumptionPlan.id.desc())
        .limit(1)
    )
    row = r.first()
    return row[0] if row else None


def _recon_aggregate_status(overall_pct: float, items_exceeding: int, tolerance_pct: float) -> str:
    if items_exceeding > 0:
        return "exceeds"
    if abs(overall_pct) <= 2.0:
        return "on_target"
    if abs(overall_pct) <= tolerance_pct:
        return "minor"
    return "exceeds"


async def _get_consumption_recon_data(
    order_id: int,
    tolerance_pct: float,
    tenant: Tenant,
    db: AsyncSession,
) -> ConsumptionReconResponse | None:
    """Compute BOM-based planned vs StockMovement actuals for one order. Returns None if order not found."""
    order = await db.get(Order, order_id)
    if not order or order.tenant_id != tenant.id:
        return None
    plan_status = await _consumption_plan_status_for_order(db, tenant.id, order.id)
    style_code = str(order_id)
    order_qty = _to_float_safe(str(order.quantity)) if order.quantity is not None else 0.0
    if not order.quotation_id or order_qty <= 0:
        return ConsumptionReconResponse(
            order=ConsumptionReconOrderOut(
                id=order.id,
                order_code=order.order_code or "",
                style_code=style_code,
                quantity=order.quantity,
            ),
            items=[],
            summary=_empty_recon_summary(),
            bom_version=None,
            bom_status=None,
            order_status=order.status,
            consumption_plan_status=plan_status,
        )
    quotation = await db.get(Quotation, order.quotation_id)
    if not quotation or quotation.tenant_id != tenant.id or not quotation.style_id:
        return ConsumptionReconResponse(
            order=ConsumptionReconOrderOut(
                id=order.id,
                order_code=order.order_code or "",
                style_code=style_code,
                quantity=order.quantity,
            ),
            items=[],
            summary=_empty_recon_summary(),
            bom_version=None,
            bom_status=None,
            order_status=order.status,
            consumption_plan_status=plan_status,
        )
    sid = quotation.style_id
    style = await db.get(GarmentStyle, sid)
    style_code = style.style_code if style else str(sid)
    order_governed = await _get_active_order_governed_bom(
        db, tenant_id=tenant.id, order_id=order.id
    )
    if order_governed:
        bom = order_governed
        use_order_bom = True
    else:
        bom = await _get_latest_governed_bom(
            db,
            tenant_id=tenant.id,
            style_id=sid,
        )
        use_order_bom = False
    if not bom:
        return ConsumptionReconResponse(
            order=ConsumptionReconOrderOut(
                id=order.id,
                order_code=order.order_code or "",
                style_code=style_code,
                quantity=order.quantity,
            ),
            items=[],
            summary=_empty_recon_summary(),
            bom_version=None,
            bom_status=None,
            order_status=order.status,
            consumption_plan_status=plan_status,
        )
    lines_result = await db.execute(
        select(BomItem).where(
            BomItem.tenant_id == tenant.id,
            BomItem.bom_id == bom.id,
            BomItem.item_id.isnot(None),
        )
    )
    items_out: list[ConsumptionReconItemOut] = []
    total_planned = 0.0
    total_actual = 0.0
    total_planned_cost = 0.0
    total_actual_cost = 0.0
    total_quoted_qty = 0.0
    total_quoted_cost = 0.0
    total_bom_cost = 0.0
    items_exceeding = 0
    for line in lines_result.scalars().all():
        item = await db.get(Item, line.item_id)
        if not item or item.tenant_id != tenant.id:
            continue
        cat = await db.get(ItemCategory, item.category_id) if item.category_id else None
        material_type = _wastage_category_from_item(item, cat)
        quoted_cons = None
        quoted_planned_qty = None
        bom_net = None
        bom_gross = None
        w_pct = None
        pl_pct = None
        if use_order_bom and line.bom_gross_consumption_per_unit is not None:
            bom_gross = float(line.bom_gross_consumption_per_unit)
            planned_qty = order_qty * bom_gross
            if line.quoted_consumption_per_unit is not None:
                quoted_cons = float(line.quoted_consumption_per_unit)
                quoted_planned_qty = order_qty * quoted_cons
            if line.bom_net_consumption_per_unit is not None:
                bom_net = float(line.bom_net_consumption_per_unit)
            w_pct = _to_float_safe(line.wastage_pct)
            if line.process_loss_pct is not None:
                pl_pct = float(line.process_loss_pct)
        else:
            base = _to_float_safe(line.base_consumption)
            wastage = _to_float_safe(line.wastage_pct) / 100.0
            planned_qty = order_qty * base * (1.0 + wastage)
        act_result = await db.execute(
            select(StockMovement).where(
                StockMovement.tenant_id == tenant.id,
                StockMovement.reference_type == "CONSUMPTION_ISSUE",
                StockMovement.reference_id == order.id,
                StockMovement.item_id == line.item_id,
            )
        )
        movements = act_result.scalars().all()
        out_movements = [m for m in movements if (m.movement_type or "").upper() == "OUT"]
        actual_qty = sum(_to_float_safe(m.quantity) for m in out_movements)
        movement_count = len(out_movements)
        cost_qty_weighted = 0.0
        cost_sum = 0.0
        last_dt: datetime | None = None
        for m in out_movements:
            q = _to_float_safe(m.quantity)
            uc_m = _to_float_safe(m.unit_cost) if m.unit_cost else 0.0
            if uc_m > 0:
                cost_qty_weighted += q
                cost_sum += q * uc_m
            cand = None
            if m.movement_date:
                cand = datetime.combine(m.movement_date, time.min)
            elif m.created_at:
                cand = m.created_at
            if cand and (last_dt is None or cand > last_dt):
                last_dt = cand
        base_uc = _to_float_safe(item.default_cost)
        if cost_qty_weighted > 0:
            unit_cost = cost_sum / cost_qty_weighted
        else:
            unit_cost = base_uc
        bom_price = (
            float(line.bom_expected_unit_price)
            if use_order_bom and line.bom_expected_unit_price is not None
            else unit_cost
        )
        quoted_price = (
            float(line.quoted_unit_price)
            if use_order_bom and line.quoted_unit_price is not None
            else None
        )
        planned_cost = planned_qty * bom_price
        quoted_planned_cost = None
        if quoted_planned_qty is not None and quoted_price is not None:
            quoted_planned_cost = quoted_planned_qty * quoted_price
        actual_cost_computed = 0.0
        for m in out_movements:
            q = _to_float_safe(m.quantity)
            uc_line = _to_float_safe(m.unit_cost) if m.unit_cost else unit_cost
            actual_cost_computed += q * uc_line
        cost_variance = actual_cost_computed - planned_cost
        last_issued_at = last_dt.isoformat() if last_dt else None
        variance = actual_qty - planned_qty
        variance_pct = (variance / planned_qty * 100.0) if planned_qty > 0 else 0.0
        if abs(variance_pct) > tolerance_pct:
            items_exceeding += 1
        uom = None
        if item.unit_id:
            u = await db.get(ItemUnit, item.unit_id)
            uom = u.code if u else None
        total_planned += planned_qty
        total_actual += actual_qty
        total_planned_cost += planned_cost
        total_actual_cost += actual_cost_computed
        if quoted_planned_qty is not None:
            total_quoted_qty += quoted_planned_qty
        if quoted_planned_cost is not None:
            total_quoted_cost += quoted_planned_cost
        total_bom_cost += planned_cost

        qvb = None
        qva = None
        bva = None
        if quoted_cons is not None and bom_gross is not None and quoted_cons > 0:
            qvb = (bom_gross - quoted_cons) / quoted_cons * 100.0
        if quoted_planned_qty and quoted_planned_qty > 0:
            qva = (actual_qty - quoted_planned_qty) / quoted_planned_qty * 100.0
        if planned_qty > 0:
            bva = (actual_qty - planned_qty) / planned_qty * 100.0
        pwq = float(line.wastage_qty) if use_order_bom and line.wastage_qty is not None else None
        plq = float(line.process_loss_qty) if use_order_bom and line.process_loss_qty is not None else None
        planned_loss_total = (pwq or 0) + (plq or 0)
        actual_loss_proxy = max(0.0, actual_qty - (order_qty * (bom_net or 0))) if bom_net else None
        pl_vs_al = (
            (planned_loss_total - actual_loss_proxy)
            if actual_loss_proxy is not None and use_order_bom
            else None
        )
        c_q_b = (
            (planned_cost - quoted_planned_cost)
            if quoted_planned_cost is not None and use_order_bom
            else None
        )
        c_b_a = actual_cost_computed - planned_cost
        c_q_a = (
            (actual_cost_computed - quoted_planned_cost)
            if quoted_planned_cost is not None and use_order_bom
            else None
        )

        items_out.append(
            ConsumptionReconItemOut(
                item_id=line.item_id,
                item_code=item.item_code or "",
                item_name=item.name or item.description or "",
                material_type=material_type,
                uom=uom,
                planned_qty=round(planned_qty, 4),
                actual_qty=round(actual_qty, 4),
                variance=round(variance, 4),
                variance_pct=round(variance_pct, 2),
                unit_cost=round(unit_cost, 4),
                planned_cost=round(planned_cost, 4),
                actual_cost=round(actual_cost_computed, 4),
                cost_variance=round(cost_variance, 4),
                last_issued_at=last_issued_at,
                movement_count=movement_count,
                quoted_consumption_per_unit=round(quoted_cons, 6) if quoted_cons is not None else None,
                bom_net_consumption_per_unit=round(bom_net, 6) if bom_net is not None else None,
                bom_gross_consumption_per_unit=round(bom_gross, 6) if bom_gross is not None else None,
                wastage_pct=round(w_pct, 4) if w_pct is not None else None,
                process_loss_pct=round(pl_pct, 4) if pl_pct is not None else None,
                quoted_planned_qty=round(quoted_planned_qty, 4) if quoted_planned_qty is not None else None,
                quoted_planned_cost=round(quoted_planned_cost, 4) if quoted_planned_cost is not None else None,
                bom_planned_cost=round(planned_cost, 4) if use_order_bom else None,
                quoted_vs_bom_variance_pct=round(qvb, 2) if qvb is not None else None,
                bom_vs_actual_variance_pct=round(bva, 2) if bva is not None else None,
                quoted_vs_actual_variance_pct=round(qva, 2) if qva is not None else None,
                planned_wastage_qty=round(pwq, 4) if pwq is not None else None,
                planned_process_loss_qty=round(plq, 4) if plq is not None else None,
                planned_loss_vs_actual_loss=round(pl_vs_al, 4) if pl_vs_al is not None else None,
                cost_impact_quoted_vs_bom=round(c_q_b, 4) if c_q_b is not None else None,
                cost_impact_bom_vs_actual=round(c_b_a, 4),
                cost_impact_quoted_vs_actual=round(c_q_a, 4) if c_q_a is not None else None,
            )
        )
    overall_pct = (total_actual - total_planned) / total_planned * 100.0 if total_planned > 0 else 0.0
    cost_var = total_actual_cost - total_planned_cost
    cost_var_pct = (cost_var / total_planned_cost * 100.0) if total_planned_cost > 0 else 0.0
    qv_b_c = total_bom_cost - total_quoted_cost
    qv_a_c = total_actual_cost - total_quoted_cost
    return ConsumptionReconResponse(
        order=ConsumptionReconOrderOut(
            id=order.id,
            order_code=order.order_code or "",
            style_code=style_code,
            quantity=order.quantity,
        ),
        items=items_out,
        summary=ConsumptionReconSummaryOut(
            total_planned=round(total_planned, 4),
            total_actual=round(total_actual, 4),
            variance=round(total_actual - total_planned, 4),
            overall_variance_pct=round(overall_pct, 2),
            items_exceeding_tolerance=items_exceeding,
            total_planned_cost=round(total_planned_cost, 4),
            total_actual_cost=round(total_actual_cost, 4),
            cost_variance=round(cost_var, 4),
            cost_variance_pct=round(cost_var_pct, 2),
            total_quoted_planned_qty=round(total_quoted_qty, 4),
            total_quoted_planned_cost=round(total_quoted_cost, 4),
            total_bom_planned_cost=round(total_bom_cost, 4),
            quoted_vs_bom_cost_variance=round(qv_b_c, 4),
            quoted_vs_actual_cost_variance=round(qv_a_c, 4),
        ),
        bom_version=str(bom.version_no),
        bom_status=bom.status,
        order_status=order.status,
        consumption_plan_status=plan_status,
    )


CONSUMPTION_RECON_DASHBOARD_MAX_ORDERS = 500


class ConsumptionReconDashboardOrderRow(BaseModel):
    order_id: int
    order_code: str
    style_code: str
    style_id: int | None = None
    buyer_name: str | None
    order_qty: int | None
    total_planned: float
    total_actual: float
    variance: float
    overall_variance_pct: float
    items_exceeding_tolerance: int
    total_items: int
    worst_item_name: str | None
    worst_item_variance_pct: float
    status: str


class ConsumptionReconCategoryBreakdown(BaseModel):
    material_type: str
    total_planned: float
    total_actual: float
    variance_pct: float


class ConsumptionReconDashboardSummary(BaseModel):
    total_orders: int
    orders_on_target: int
    orders_minor: int
    orders_exceeding: int
    avg_variance_pct: float
    total_planned_qty: float
    total_actual_qty: float


class ConsumptionReconDashboardResponse(BaseModel):
    orders: list[ConsumptionReconDashboardOrderRow]
    summary: ConsumptionReconDashboardSummary
    category_breakdown: list[ConsumptionReconCategoryBreakdown]
    total_count: int


class ConsumptionReconTrendPoint(BaseModel):
    period: str
    orders_count: int
    avg_variance_pct: float
    total_planned: float
    total_actual: float
    exceeding_count: int


class ConsumptionReconTrendsResponse(BaseModel):
    points: list[ConsumptionReconTrendPoint]
    tolerance_pct: float


async def _orders_base_query_for_recon(
    tenant_id: int,
    buyer_id: int | None,
    style_id: int | None,
    date_from: date | None,
    date_to: date | None,
):
    q = select(Order).where(Order.tenant_id == tenant_id)
    if buyer_id is not None:
        q = q.where(Order.customer_id == buyer_id)
    if style_id is not None:
        q = q.join(Quotation, Order.quotation_id == Quotation.id).where(Quotation.style_id == style_id)
    if date_from is not None:
        start = datetime.combine(date_from, time.min)
        q = q.where(Order.created_at >= start)
    if date_to is not None:
        end = datetime.combine(date_to, time.max)
        q = q.where(Order.created_at <= end)
    return q.order_by(Order.created_at.desc()).limit(CONSUMPTION_RECON_DASHBOARD_MAX_ORDERS)


def _worst_variance_item(items: list[ConsumptionReconItemOut]) -> tuple[str | None, float]:
    if not items:
        return None, 0.0
    worst = max(items, key=lambda x: abs(x.variance_pct))
    label = f"{worst.item_code} {worst.item_name}".strip()
    return label, worst.variance_pct


@router.get("/consumption-reconciliation/dashboard", response_model=ConsumptionReconDashboardResponse)
async def get_consumption_reconciliation_dashboard(
    buyer_id: int | None = Query(default=None),
    style_id: int | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    material_type: str | None = Query(default=None),
    tolerance_pct: float = Query(default=5.0),
    limit: int = Query(default=20, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    sort_by: str = Query(default="overall_variance_pct"),
    sort_dir: str = Query(default="desc"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Multi-order consumption reconciliation summary (scans up to 500 most recent matching orders)."""
    _ensure_tenant(user, tenant)
    q = await _orders_base_query_for_recon(tenant.id, buyer_id, style_id, date_from, date_to)
    ores = await db.execute(q)
    orders_list = ores.scalars().unique().all()
    rows_full: list[ConsumptionReconDashboardOrderRow] = []
    cat_acc: dict[str, tuple[float, float]] = defaultdict(lambda: (0.0, 0.0))
    for order in orders_list:
        data = await _get_consumption_recon_data(order.id, tolerance_pct, tenant, db)
        if data is None:
            continue
        worst_name, worst_pct = _worst_variance_item(data.items)
        st = _recon_aggregate_status(
            data.summary.overall_variance_pct,
            data.summary.items_exceeding_tolerance,
            tolerance_pct,
        )
        if status_filter and status_filter.strip().lower() != st:
            continue
        if material_type and material_type.strip():
            mt = material_type.strip().lower()
            if not any(it.material_type.lower() == mt for it in data.items):
                continue
        cust = await db.get(Customer, order.customer_id)
        buyer_name = None
        if cust:
            buyer_name = (cust.trade_name or cust.legal_entity_name or cust.name or "").strip() or None
        sid_row: int | None = None
        if order.quotation_id:
            qo = await db.get(Quotation, order.quotation_id)
            if qo and qo.style_id:
                sid_row = qo.style_id
        rows_full.append(
            ConsumptionReconDashboardOrderRow(
                order_id=order.id,
                order_code=data.order.order_code,
                style_code=data.order.style_code,
                style_id=sid_row,
                buyer_name=buyer_name,
                order_qty=data.order.quantity,
                total_planned=data.summary.total_planned,
                total_actual=data.summary.total_actual,
                variance=data.summary.variance,
                overall_variance_pct=data.summary.overall_variance_pct,
                items_exceeding_tolerance=data.summary.items_exceeding_tolerance,
                total_items=len(data.items),
                worst_item_name=worst_name,
                worst_item_variance_pct=round(worst_pct, 2),
                status=st,
            )
        )
        for it in data.items:
            p0, a0 = cat_acc[it.material_type]
            cat_acc[it.material_type] = (p0 + it.planned_qty, a0 + it.actual_qty)
    sort_key = sort_by if sort_by in (
        "order_code",
        "overall_variance_pct",
        "total_planned",
        "total_actual",
        "variance",
    ) else "overall_variance_pct"
    reverse = sort_dir.lower() != "asc"

    def sort_val(r: ConsumptionReconDashboardOrderRow) -> float | str:
        v = getattr(r, sort_key, r.overall_variance_pct)
        return v if isinstance(v, (int, float)) else str(v)

    rows_full.sort(key=sort_val, reverse=reverse)
    total_count = len(rows_full)
    page_rows = rows_full[offset : offset + limit]
    on_target = sum(1 for r in rows_full if r.status == "on_target")
    minor = sum(1 for r in rows_full if r.status == "minor")
    exceeds = sum(1 for r in rows_full if r.status == "exceeds")
    avg_var = (
        sum(r.overall_variance_pct for r in rows_full) / len(rows_full) if rows_full else 0.0
    )
    tp = sum(r.total_planned for r in rows_full)
    ta = sum(r.total_actual for r in rows_full)
    breakdown: list[ConsumptionReconCategoryBreakdown] = []
    for mtype, (cp, ca) in sorted(cat_acc.items()):
        vp = ((ca - cp) / cp * 100.0) if cp > 0 else 0.0
        breakdown.append(
            ConsumptionReconCategoryBreakdown(
                material_type=mtype,
                total_planned=round(cp, 4),
                total_actual=round(ca, 4),
                variance_pct=round(vp, 2),
            )
        )
    return ConsumptionReconDashboardResponse(
        orders=page_rows,
        summary=ConsumptionReconDashboardSummary(
            total_orders=len(rows_full),
            orders_on_target=on_target,
            orders_minor=minor,
            orders_exceeding=exceeds,
            avg_variance_pct=round(avg_var, 2),
            total_planned_qty=round(tp, 4),
            total_actual_qty=round(ta, 4),
        ),
        category_breakdown=breakdown,
        total_count=total_count,
    )


@router.get("/consumption-reconciliation/trends", response_model=ConsumptionReconTrendsResponse)
async def get_consumption_reconciliation_trends(
    months: int = Query(default=6, ge=1, le=24),
    buyer_id: int | None = Query(default=None),
    style_id: int | None = Query(default=None),
    tolerance_pct: float = Query(default=5.0),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Monthly aggregates of consumption reconciliation variance."""
    _ensure_tenant(user, tenant)
    today = date.today()
    start_month = date(today.year, today.month, 1)
    for _ in range(months - 1):
        if start_month.month == 1:
            start_month = date(start_month.year - 1, 12, 1)
        else:
            start_month = date(start_month.year, start_month.month - 1, 1)
    points: list[ConsumptionReconTrendPoint] = []
    cur = start_month
    while cur <= today.replace(day=1):
        if cur.month == 12:
            next_m = date(cur.year + 1, 1, 1)
        else:
            next_m = date(cur.year, cur.month + 1, 1)
        q = select(Order).where(Order.tenant_id == tenant.id)
        q = q.where(Order.created_at >= datetime.combine(cur, time.min))
        q = q.where(Order.created_at < datetime.combine(next_m, time.min))
        if buyer_id is not None:
            q = q.where(Order.customer_id == buyer_id)
        if style_id is not None:
            q = q.join(Quotation, Order.quotation_id == Quotation.id).where(Quotation.style_id == style_id)
        ores = await db.execute(q)
        month_orders = ores.scalars().unique().all()
        var_list: list[float] = []
        tp_sum = 0.0
        ta_sum = 0.0
        ex_count = 0
        for order in month_orders:
            data = await _get_consumption_recon_data(order.id, tolerance_pct, tenant, db)
            if data is None or not data.items:
                continue
            var_list.append(data.summary.overall_variance_pct)
            tp_sum += data.summary.total_planned
            ta_sum += data.summary.total_actual
            if data.summary.items_exceeding_tolerance > 0:
                ex_count += 1
        label = f"{cur.year:04d}-{cur.month:02d}"
        n = len(var_list)
        avg_v = sum(var_list) / n if n else 0.0
        points.append(
            ConsumptionReconTrendPoint(
                period=label,
                orders_count=n,
                avg_variance_pct=round(avg_v, 2),
                total_planned=round(tp_sum, 4),
                total_actual=round(ta_sum, 4),
                exceeding_count=ex_count,
            )
        )
        cur = next_m
    return ConsumptionReconTrendsResponse(points=points, tolerance_pct=tolerance_pct)


@router.get("/consumption-reconciliation/dashboard/export")
async def get_consumption_reconciliation_dashboard_export(
    buyer_id: int | None = Query(default=None),
    style_id: int | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    material_type: str | None = Query(default=None),
    tolerance_pct: float = Query(default=5.0),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Excel export of dashboard reconciliation rows (same filters as dashboard, all matching rows)."""
    _ensure_tenant(user, tenant)
    body = await get_consumption_reconciliation_dashboard(
        buyer_id=buyer_id,
        style_id=style_id,
        date_from=date_from,
        date_to=date_to,
        status_filter=status_filter,
        material_type=material_type,
        tolerance_pct=tolerance_pct,
        limit=CONSUMPTION_RECON_DASHBOARD_MAX_ORDERS,
        offset=0,
        sort_by="overall_variance_pct",
        sort_dir="desc",
        tenant=tenant,
        user=user,
        db=db,
    )
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export not available (openpyxl missing)")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=500, detail="Workbook error")
    ws.title = "Summary"
    ws.append(["Consumption reconciliation dashboard"])
    ws.append(["Total orders (in scan)", body.summary.total_orders])
    ws.append(["On target", body.summary.orders_on_target])
    ws.append(["Minor", body.summary.orders_minor])
    ws.append(["Exceeding", body.summary.orders_exceeding])
    ws.append([])
    ws.append(
        [
            "Order",
            "Style",
            "Buyer",
            "Qty",
            "Planned",
            "Actual",
            "Var %",
            "Worst item",
            "Status",
        ]
    )
    for r in body.orders:
        ws.append(
            [
                r.order_code,
                r.style_code,
                r.buyer_name or "",
                r.order_qty or "",
                r.total_planned,
                r.total_actual,
                r.overall_variance_pct,
                r.worst_item_name or "",
                r.status,
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="consumption_recon_dashboard.xlsx"'},
    )


@router.get(
    "/consumption-reconciliation/{order_id}/movements/{item_id}",
    response_model=ConsumptionReconMovementsResponse,
)
async def get_consumption_reconciliation_movements(
    order_id: int,
    item_id: int,
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """CONSUMPTION_ISSUE stock movements for one order line item."""
    _ensure_tenant(user, tenant)
    order = await db.get(Order, order_id)
    if not order or order.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Order not found")
    item = await db.get(Item, item_id)
    if not item or item.tenant_id != tenant.id:
        raise HTTPException(status_code=404, detail="Item not found")
    act_result = await db.execute(
        select(StockMovement)
        .where(
            StockMovement.tenant_id == tenant.id,
            StockMovement.reference_type == "CONSUMPTION_ISSUE",
            StockMovement.reference_id == order.id,
            StockMovement.item_id == item_id,
        )
        .order_by(StockMovement.created_at.desc())
    )
    movements_raw = [m for m in act_result.scalars().all() if (m.movement_type or "").upper() == "OUT"]
    total_issued = sum(_to_float_safe(m.quantity) for m in movements_raw)
    out_list: list[ConsumptionReconMovementOut] = []
    for m in movements_raw:
        wh_name = None
        if m.warehouse_id:
            wh = await db.get(Warehouse, m.warehouse_id)
            wh_name = wh.name if wh else None
        issuer = None
        if m.created_by_user_id:
            u = await db.get(User, m.created_by_user_id)
            if u:
                parts = [u.first_name or "", u.last_name or ""]
                issuer = " ".join(p for p in parts if p).strip() or u.username
        md = None
        if m.movement_date:
            md = m.movement_date.isoformat()
        elif m.created_at:
            md = m.created_at.isoformat()
        out_list.append(
            ConsumptionReconMovementOut(
                movement_id=m.id,
                movement_date=md,
                quantity=round(_to_float_safe(m.quantity), 4),
                warehouse_name=wh_name,
                issued_by=issuer,
                reference_code=m.lot_number,
                notes=m.notes,
            )
        )
    planned_qty = 0.0
    if order.quotation_id:
        quotation = await db.get(Quotation, order.quotation_id)
        if quotation and quotation.style_id:
            bom = await _get_active_order_governed_bom(
                db, tenant_id=tenant.id, order_id=order.id
            )
            if not bom:
                bom = await _get_latest_governed_bom(db, tenant_id=tenant.id, style_id=quotation.style_id)
            if bom:
                bres = await db.execute(
                    select(BomItem).where(
                        BomItem.tenant_id == tenant.id,
                        BomItem.bom_id == bom.id,
                        BomItem.item_id == item_id,
                    )
                )
                bline = bres.scalars().first()
                if bline:
                    oq = _to_float_safe(str(order.quantity)) if order.quantity else 0.0
                    if bline.bom_gross_consumption_per_unit is not None:
                        planned_qty = oq * float(bline.bom_gross_consumption_per_unit)
                    else:
                        base = _to_float_safe(bline.base_consumption)
                        wastage = _to_float_safe(bline.wastage_pct) / 100.0
                        pl = (
                            float(bline.process_loss_pct) / 100.0
                            if bline.process_loss_pct is not None
                            else 0.0
                        )
                        planned_qty = oq * base * (1.0 + wastage + pl)
    return ConsumptionReconMovementsResponse(
        item_id=item_id,
        item_code=item.item_code or "",
        item_name=item.name or item.description or "",
        planned_qty=round(planned_qty, 4),
        total_issued=round(total_issued, 4),
        movements=out_list,
    )


@router.get("/consumption-reconciliation/{order_id}", response_model=ConsumptionReconResponse)
async def get_consumption_reconciliation(
    order_id: int,
    tolerance_pct: float = Query(default=5.0, description="Tolerance % for exceeding count (e.g. 5)"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """BOM-based planned vs StockMovement actuals for one order. Planned = order_qty × BOM base × (1 + wastage%); actual = CONSUMPTION_ISSUE OUT movements."""
    _ensure_tenant(user, tenant)
    data = await _get_consumption_recon_data(order_id, tolerance_pct, tenant, db)
    if data is None:
        raise HTTPException(status_code=404, detail="Order not found")
    return data


@router.get("/consumption-reconciliation/{order_id}/export")
async def get_consumption_reconciliation_export(
    order_id: int,
    format: str = Query(default="xlsx", description="xlsx"),
    tolerance_pct: float = Query(default=5.0, description="Tolerance % for exceeding count"),
    tenant: Tenant = Depends(require_tenant),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export consumption reconciliation for one order as Excel (order info + items + summary)."""
    _ensure_tenant(user, tenant)
    data = await _get_consumption_recon_data(order_id, tolerance_pct, tenant, db)
    if data is None:
        raise HTTPException(status_code=404, detail="Order not found")
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export not available (openpyxl missing)")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=500, detail="Workbook error")
    ws.title = "Reconciliation"
    ws.append(["Order", data.order.order_code, "Style", data.order.style_code, "Qty", data.order.quantity or ""])
    ws.append([])
    ws.append(
        [
            "Item",
            "Type",
            "Unit",
            "Planned",
            "Actual",
            "Variance",
            "Variance %",
            "Planned cost",
            "Actual cost",
            "Cost var.",
            "Movements",
        ]
    )
    for i in data.items:
        ws.append([
            f"{i.item_code} {i.item_name}".strip(),
            i.material_type,
            i.uom or "",
            i.planned_qty,
            i.actual_qty,
            i.variance,
            i.variance_pct,
            i.planned_cost if i.planned_cost is not None else "",
            i.actual_cost if i.actual_cost is not None else "",
            i.cost_variance if i.cost_variance is not None else "",
            i.movement_count,
        ])
    ws.append([])
    ws.append(["Total planned", data.summary.total_planned])
    ws.append(["Total actual", data.summary.total_actual])
    ws.append(["Variance", data.summary.variance])
    ws.append(["Overall variance %", data.summary.overall_variance_pct])
    ws.append(["Items exceeding tolerance", data.summary.items_exceeding_tolerance])
    ws.append(["Total planned cost", data.summary.total_planned_cost])
    ws.append(["Total actual cost", data.summary.total_actual_cost])
    ws.append(["Cost variance", data.summary.cost_variance])
    ws.append(["Cost variance %", data.summary.cost_variance_pct])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_code = "".join(c if c.isalnum() or c in "-_" else "_" for c in data.order.order_code)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="consumption_recon_order_{safe_code}.xlsx"'},
    )
